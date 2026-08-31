import os, uuid, re, unicodedata, hashlib, mimetypes, json
from decimal import Decimal, InvalidOperation
from io import BytesIO
from datetime import datetime, date, timedelta
import calendar
from flask import Blueprint, render_template, request, redirect, url_for, flash, abort, current_app, send_from_directory, send_file
from flask_login import login_required, current_user
from sqlalchemy import func, and_, or_
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, KeepTogether, Image as RLImage, PageBreak
from pypdf import PdfReader, PdfWriter
from PIL import Image as PILImage
from .extensions import db
from .models import User, Employee, EmployeeWorkSchedule, WeekendDuty, TimeClock, MedicalCertificate, MedicalCertificateAllowance, AuthThrottle, SecurityEvent, Request, Document, DocumentSignatureFlow, PayrollEmployeeConfig, PayrollDependent, PayrollLegalParameter, PayrollRubric, PayrollCompetence, PayrollManualEvent, PayrollEmployeeCalculation, PayrollCalculationItem, PayrollClosure, AuditLog, BankHourAdjustment, Vacation, VacationSchedule, TimePeriodClosure, TimeReportAcknowledgement, TimeReportFinalization, Payslip, ROLE_ADMIN, ROLE_MANAGER, ROLE_EMPLOYEE
from .security import roles_required, can_manage_employee, log_action, log_security_event, client_ip
from .timezone import now_local, today_local

bp = Blueprint("rh", __name__, url_prefix="/rh")
ALLOWED_EXTENSIONS = {"pdf", "png", "jpg", "jpeg", "webp"}

MAX_IMAGE_PIXELS = 30_000_000
PIN_MAX_FAILURES = 5
PIN_BLOCK_MINUTES = 15


def _safe_original_filename(filename, fallback="arquivo"):
    name = secure_filename(filename or "")[:180]
    return name or fallback


def _chmod_private(path):
    """No Linux/Render, restringe o arquivo ao usuário do processo."""
    try:
        os.chmod(path, 0o600)
    except OSError:
        pass


def _validate_pdf_stream(file_storage):
    """Valida assinatura e estrutura PDF antes de persistir o upload."""
    stream = file_storage.stream
    stream.seek(0)
    header = stream.read(5)
    stream.seek(0)
    if header != b"%PDF-":
        raise ValueError("O conteúdo do arquivo não corresponde a um PDF válido.")
    try:
        reader = PdfReader(stream)
        _ = len(reader.pages)
        if len(reader.pages) < 1:
            raise ValueError
    except Exception:
        stream.seek(0)
        raise ValueError("O PDF enviado está corrompido ou não pôde ser validado.")
    stream.seek(0)


def _validate_image_stream(file_storage):
    """Decodifica a imagem para impedir arquivos executáveis disfarçados de imagem."""
    stream = file_storage.stream
    stream.seek(0)
    try:
        image = PILImage.open(stream)
        width, height = image.size
        if width <= 0 or height <= 0 or width * height > MAX_IMAGE_PIXELS:
            raise ValueError
        fmt = (image.format or "").upper()
        if fmt not in {"JPEG", "PNG", "WEBP"}:
            raise ValueError
        image.verify()
    except Exception:
        stream.seek(0)
        raise ValueError("A imagem enviada é inválida ou excede os limites de segurança.")
    stream.seek(0)


def _validate_uploaded_content(file_storage, ext):
    ext = (ext or "").lower()
    if ext == "pdf":
        _validate_pdf_stream(file_storage)
    elif ext in {"png", "jpg", "jpeg", "webp"}:
        _validate_image_stream(file_storage)
    else:
        raise ValueError("Tipo de arquivo não permitido.")


def _pin_throttle_key(emp):
    raw = f"pin|{emp.id}|{client_ip()}|{current_app.config['SECRET_KEY']}"
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def _pin_is_blocked(emp):
    row = AuthThrottle.query.filter_by(key_hash=_pin_throttle_key(emp)).first()
    if not row:
        return False
    now = now_local()
    if row.blocked_until and row.blocked_until > now:
        return True
    if row.blocked_until and row.blocked_until <= now:
        db.session.delete(row)
        db.session.commit()
    return False


def _pin_failure(emp):
    key = _pin_throttle_key(emp)
    now = now_local()
    row = AuthThrottle.query.filter_by(key_hash=key).first()
    if not row:
        row = AuthThrottle(key_hash=key, failures=0, window_started_at=now)
        db.session.add(row)
    if not row.window_started_at or now - row.window_started_at > timedelta(minutes=PIN_BLOCK_MINUTES):
        row.failures = 0
        row.window_started_at = now
    row.failures += 1
    row.last_failure_at = now

    severity = "warning"
    event_type = "pin_failed"
    if row.failures >= PIN_MAX_FAILURES:
        row.blocked_until = now + timedelta(minutes=PIN_BLOCK_MINUTES)
        severity = "critical"
        event_type = "pin_blocked"

    log_security_event(
        event_type,
        severity=severity,
        user=emp.user,
        employee=emp,
        details=(
            f"Tentativa de PIN recusada. Falhas na janela atual: {row.failures}. "
            + (
                f"Bloqueado até {row.blocked_until.strftime('%d/%m/%Y %H:%M:%S')}."
                if row.blocked_until else
                "Ainda não bloqueado."
            )
        ),
    )
    db.session.commit()


def _pin_success(emp):
    row = AuthThrottle.query.filter_by(key_hash=_pin_throttle_key(emp)).first()
    if row:
        db.session.delete(row)
        db.session.commit()







def _normalize_person_name(value):
    value = unicodedata.normalize("NFKD", value or "")
    value = "".join(ch for ch in value if not unicodedata.combining(ch))
    value = value.lower()
    value = re.sub(r"[^a-z0-9]+", " ", value)
    return " ".join(value.split())


def _match_employee_from_text(text, employees):
    """
    Identifica o colaborador pelo nome completo encontrado no texto de uma página.
    Retorna (employee, status).
    """
    normalized_text = _normalize_person_name(text)
    if not normalized_text:
        return None, "sem_texto"

    document_tokens = set(normalized_text.split())
    candidates = []

    for emp in employees:
        normalized_name = _normalize_person_name(emp.full_name)
        name_tokens = [token for token in normalized_name.split() if token]
        if not name_tokens:
            continue

        if all(token in document_tokens for token in name_tokens):
            candidates.append((len(name_tokens), len(normalized_name), emp))

    if not candidates:
        return None, "nao_encontrado"

    candidates.sort(key=lambda item: (item[0], item[1]), reverse=True)
    best = candidates[0]

    if len(candidates) > 1 and candidates[1][0:2] == best[0:2]:
        return None, "ambiguo"

    return best[2], "automatico_pdf"


def _split_payslip_pdf_by_employee(file_storage, employees):
    """
    Lê um PDF consolidado página por página.

    Retorna:
    - groups: {employee_id: {"employee": Employee, "pages": [PageObject], "page_numbers": [int]}}
    - unmatched: [{"page": int, "reason": str}]
    """
    try:
        file_storage.stream.seek(0)
        reader = PdfReader(file_storage.stream)
    except Exception:
        try:
            file_storage.stream.seek(0)
        except Exception:
            pass
        raise ValueError("Não foi possível abrir o PDF consolidado.")

    groups = {}
    unmatched = []

    for index, page in enumerate(reader.pages, start=1):
        try:
            text = page.extract_text() or ""
        except Exception:
            text = ""

        emp, status = _match_employee_from_text(text, employees)

        if not emp:
            unmatched.append({"page": index, "reason": status})
            continue

        bucket = groups.setdefault(
            emp.id,
            {"employee": emp, "pages": [], "page_numbers": []}
        )
        bucket["pages"].append(page)
        bucket["page_numbers"].append(index)

    try:
        file_storage.stream.seek(0)
    except Exception:
        pass

    return groups, unmatched


def _save_employee_payslip_pages(emp, year, month, pages, original_name, matched_by):
    """
    Gera um PDF individual somente com as páginas associadas ao colaborador.
    Substitui o holerite existente da mesma competência.
    """
    writer = PdfWriter()
    for page in pages:
        writer.add_page(page)

    stored = f"holerite_{emp.id}_{year}_{month:02d}_{uuid.uuid4().hex}.pdf"
    output_path = os.path.join(current_app.config["UPLOAD_FOLDER"], stored)

    with open(output_path, "wb") as output_file:
        writer.write(output_file)
    _chmod_private(output_path)

    existing = Payslip.query.filter_by(
        employee_id=emp.id,
        year=year,
        month=month,
    ).first()

    old_stored = None
    individual_original = f"Holerite - {emp.full_name} - {month:02d}-{year}.pdf"

    if existing:
        old_stored = existing.stored_name
        existing.original_name = individual_original
        existing.stored_name = stored
        existing.matched_by = matched_by
        existing.uploaded_at = now_local()
        existing.uploaded_by = current_user.id
        existing.employee_viewed_at = None
        item = existing
    else:
        item = Payslip(
            employee_id=emp.id,
            year=year,
            month=month,
            original_name=individual_original,
            stored_name=stored,
            matched_by=matched_by,
            uploaded_by=current_user.id,
        )
        db.session.add(item)

    db.session.flush()

    if old_stored and old_stored != stored:
        old_path = os.path.join(current_app.config["UPLOAD_FOLDER"], old_stored)
        try:
            if os.path.isfile(old_path):
                os.remove(old_path)
        except OSError:
            pass

    return item


def _parse_competence(value):
    try:
        year, month = map(int, (value or "").split("-"))
        if year < 2000 or month < 1 or month > 12:
            raise ValueError
        return year, month
    except Exception:
        raise ValueError("Informe uma competência válida no formato mês/ano.")


def _save_payslip_file(file):
    if not file or not file.filename:
        raise ValueError("Selecione um arquivo de holerite.")
    filename = _safe_original_filename(file.filename, "holerite.pdf")
    if "." not in filename or filename.rsplit(".", 1)[1].lower() != "pdf":
        raise ValueError("Os holerites devem ser enviados em PDF.")

    _validate_pdf_stream(file)

    stored = f"holerite_{uuid.uuid4().hex}.pdf"
    path = os.path.join(current_app.config["UPLOAD_FOLDER"], stored)
    file.save(path)
    _chmod_private(path)
    return filename, stored


def _upsert_payslip(emp, year, month, file, matched_by):
    original, stored = _save_payslip_file(file)
    existing = Payslip.query.filter_by(
        employee_id=emp.id, year=year, month=month
    ).first()

    old_stored = None
    if existing:
        old_stored = existing.stored_name
        existing.original_name = original
        existing.stored_name = stored
        existing.matched_by = matched_by
        existing.uploaded_at = now_local()
        existing.uploaded_by = current_user.id
        existing.employee_viewed_at = None
        item = existing
    else:
        item = Payslip(
            employee_id=emp.id,
            year=year,
            month=month,
            original_name=original,
            stored_name=stored,
            matched_by=matched_by,
            uploaded_by=current_user.id,
        )
        db.session.add(item)

    db.session.flush()

    if old_stored and old_stored != stored:
        old_path = os.path.join(current_app.config["UPLOAD_FOLDER"], old_stored)
        try:
            if os.path.isfile(old_path):
                os.remove(old_path)
        except OSError:
            pass

    return item


def _time_report_signature_code(ack, emp):
    raw = (
        f"{ack.id}|{emp.id}|{emp.user_id}|{ack.year}|{ack.month}|"
        f"{ack.acknowledged_at.isoformat()}"
    )
    digest = hashlib.sha256(raw.encode("utf-8")).hexdigest().upper()
    return f"RH-{ack.year}{ack.month:02d}-{digest[:20]}"


def _expected_daily_minutes(emp):
    """
    Jornada diária de referência.
    Quando o RH cadastrou um intervalo individual, desconta exatamente esse período.
    Na ausência, mantém a regra anterior de 1h para jornadas superiores a 6h.
    """
    if emp.standard_start and emp.standard_end:
        start = datetime.combine(today_local(), emp.standard_start)
        end = datetime.combine(today_local(), emp.standard_end)
        minutes = int((end - start).total_seconds() // 60)

        schedule = emp.work_schedule
        if (
            schedule
            and schedule.interval_start
            and schedule.interval_end
            and schedule.interval_end > schedule.interval_start
        ):
            interval_start_dt = datetime.combine(today_local(), schedule.interval_start)
            interval_end_dt = datetime.combine(today_local(), schedule.interval_end)
            interval_minutes = int((interval_end_dt - interval_start_dt).total_seconds() // 60)
            minutes -= max(interval_minutes, 0)
        elif minutes > 6 * 60:
            minutes -= 60

        return max(minutes, 0)

    return int(round(float(emp.weekly_hours or 0) * 60 / 5))


def _certificate_allowance_by_day(emp, range_start, range_end):
    """
    Distribui o total de horas abonadas de cada atestado pelos dias úteis cobertos,
    limitado à jornada diária prevista. O abono justifica jornada; não cria crédito
    de banco de horas.
    """
    expected = _expected_daily_minutes(emp)
    result = {}

    allowances = (MedicalCertificateAllowance.query
        .filter(MedicalCertificateAllowance.employee_id == emp.id)
        .all())

    for allowance in allowances:
        cert = allowance.certificate
        if not cert or cert.status == "devolvido":
            continue

        remaining = max(int(allowance.minutes or 0), 0)
        if remaining <= 0:
            continue

        cert_start = cert.start_date
        cert_end = cert.start_date + timedelta(days=max(int(cert.days or 1) - 1, 0))
        current = cert_start

        # A distribuição começa no início do atestado, mesmo que o relatório solicitado
        # comece no mês seguinte. Isso preserva a ordem correta em atestados que cruzam meses.
        while current <= cert_end and remaining > 0:
            if current.weekday() < 5 and current >= emp.admission_date:
                allocation = min(remaining, expected)
                if range_start <= current <= range_end:
                    result[current] = result.get(current, 0) + allocation
                remaining -= allocation
            current += timedelta(days=1)

    return result


def _month_clock_summary(emp, year, month):
    first = date(year, month, 1)
    last = date(year, month, calendar.monthrange(year, month)[1])
    rows = (TimeClock.query.filter(
        TimeClock.employee_id == emp.id,
        func.date(TimeClock.punched_at) >= first,
        func.date(TimeClock.punched_at) <= last
    ).order_by(TimeClock.punched_at.asc()).all())

    grouped = {}
    for row in rows:
        grouped.setdefault(row.punched_at.date(), []).append(row)

    expected = _expected_daily_minutes(emp)
    allowance_by_day = _certificate_allowance_by_day(emp, first, last)

    worked = 0
    excused = 0
    excused_applied = 0
    expected_total = 0
    incomplete = 0
    balance = 0

    for day in range(1, last.day + 1):
        d = date(year, month, day)
        if d.weekday() >= 5 or d < emp.admission_date or d > today_local():
            continue

        day_rows = grouped.get(d, [])
        expected_total += expected

        wm = _worked_minutes_for_day(day_rows)
        worked += wm

        registered_excused = int(allowance_by_day.get(d, 0) or 0)
        excused += registered_excused

        # O atestado cobre somente o déficit da jornada. Nunca gera saldo positivo.
        applied_excused = min(registered_excused, max(expected - wm, 0))
        excused_applied += applied_excused

        kinds = {x.kind for x in day_rows}
        accounted = wm + applied_excused

        if accounted < expected:
            if not day_rows or not {"entrada", "saida"}.issubset(kinds):
                incomplete += 1

        balance += accounted - expected

    return {
        "worked": worked,
        "excused": excused,
        "excused_applied": excused_applied,
        "expected": expected_total,
        "balance": balance,
        "incomplete": incomplete,
        "rows": rows,
    }

def _add_years(value, years):
    """Adiciona anos preservando mês/dia; 29/02 vira 28/02 quando necessário."""
    target_year = value.year + years
    try:
        return value.replace(year=target_year)
    except ValueError:
        return value.replace(year=target_year, day=28)

def _vacation_entitlement(emp):
    """Resumo operacional de férias com período aquisitivo/concessivo e programação."""
    today = today_local()
    admission = emp.admission_date

    completed_periods = 0
    while _add_years(admission, completed_periods + 1) <= today:
        completed_periods += 1

    current_start = _add_years(admission, completed_periods)
    current_next_anniversary = _add_years(admission, completed_periods + 1)
    current_end = current_next_anniversary - timedelta(days=1)

    if completed_periods > 0:
        last_acq_start = _add_years(admission, completed_periods - 1)
        last_acq_end = _add_years(admission, completed_periods) - timedelta(days=1)
        concession_start = last_acq_end + timedelta(days=1)
        concession_end = _add_years(concession_start, 1) - timedelta(days=1)
    else:
        last_acq_start = None
        last_acq_end = None
        concession_start = None
        concession_end = None

    earned = completed_periods * 30
    vacations = Vacation.query.filter_by(employee_id=emp.id).all()
    used = sum(int(v.days or 0) for v in vacations)

    schedules = VacationSchedule.query.filter_by(employee_id=emp.id, status="planned").all()
    scheduled_days = sum(int(v.days or 0) for v in schedules)

    available = max(earned - used, 0)
    available_after_schedule = max(available - scheduled_days, 0)

    return {
        "completed_periods": completed_periods,
        "earned": earned,
        "used": used,
        "available": available,
        "scheduled_days": scheduled_days,
        "available_after_schedule": available_after_schedule,
        "current_acquisition_start": current_start,
        "current_acquisition_end": current_end,
        "next_entitlement_date": current_next_anniversary,
        "last_acquisition_start": last_acq_start,
        "last_acquisition_end": last_acq_end,
        "concession_start": concession_start,
        "concession_end": concession_end,
    }

def _pdf_text(value):
    """Escape basic XML-sensitive characters used by ReportLab Paragraph."""
    if value is None:
        return ""
    return str(value).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

def _month_name_pt(month):
    names = ["", "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
             "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
    return names[month] if 1 <= month <= 12 else str(month)

def _kind_label(kind):
    return {
        "entrada": "Entrada",
        "saida_intervalo": "Saída intervalo",
        "retorno": "Retorno",
        "saida": "Saída",
    }.get(kind, (kind or "").replace("_", " ").title())


def _format_minutes(total_minutes):
    """Formata minutos com sinal opcional em HH:MM."""
    total_minutes = int(round(total_minutes or 0))
    sign = "-" if total_minutes < 0 else ""
    total_minutes = abs(total_minutes)
    hours, minutes = divmod(total_minutes, 60)
    return f"{sign}{hours:02d}:{minutes:02d}"

def _worked_minutes_for_day(day_rows):
    """Calcula horas efetivamente registradas no dia a partir das marcações principais."""
    by_kind = {"entrada": [], "saida_intervalo": [], "retorno": [], "saida": []}
    for row in sorted(day_rows, key=lambda r: r.punched_at):
        if row.kind in by_kind:
            by_kind[row.kind].append(row)
    entrada = by_kind["entrada"][0].punched_at if by_kind["entrada"] else None
    saida_intervalo = by_kind["saida_intervalo"][0].punched_at if by_kind["saida_intervalo"] else None
    retorno = by_kind["retorno"][0].punched_at if by_kind["retorno"] else None
    saida = by_kind["saida"][0].punched_at if by_kind["saida"] else None
    total_seconds = 0
    if entrada and saida_intervalo and saida_intervalo >= entrada:
        total_seconds += (saida_intervalo - entrada).total_seconds()
    if retorno and saida and saida >= retorno:
        total_seconds += (saida - retorno).total_seconds()
    # Jornada sem intervalo registrado: usa entrada -> saída apenas quando não há marcações de intervalo.
    if entrada and saida and not saida_intervalo and not retorno and saida >= entrada:
        total_seconds = (saida - entrada).total_seconds()
    return int(round(total_seconds / 60))

def _draw_pdf_footer(canvas, doc):
    canvas.saveState()
    canvas.setFont("Helvetica", 8)
    canvas.drawString(15 * mm, 8 * mm, "Portal RH - Espelho mensal de registros de ponto")
    canvas.drawRightString(282 * mm, 8 * mm, f"Página {doc.page}")
    canvas.restoreState()

def parse_date(v): return datetime.strptime(v, "%Y-%m-%d").date() if v else None
def parse_time(v):
    if not v:
        return None
    value = v.strip()
    for fmt in ("%H:%M:%S", "%H:%M"):
        try:
            return datetime.strptime(value, fmt).time()
        except ValueError:
            continue
    raise ValueError(f"Horário inválido: {value}")


def parse_money(value, default="0"):
    raw = (value or default).strip().replace(".", "").replace(",", ".")
    try:
        return Decimal(raw).quantize(Decimal("0.01"))
    except (InvalidOperation, AttributeError):
        raise ValueError("Informe um valor monetário válido.")


PAYROLL_2026_LEGAL_PARAMETERS = [
    # Salário mínimo
    ("minimum_wage", "Salário mínimo nacional", "1621.00", "money",
     "Decreto nº 12.797/2025", "https://www.planalto.gov.br/ccivil_03/_ato2023-2026/2025/decreto/d12797.htm",
     "Vigência a partir de 01/01/2026."),

    # INSS empregado - limites superiores e alíquotas progressivas
    ("inss_band_1_limit", "INSS 1ª faixa - limite", "1621.00", "money",
     "Portaria Interministerial MPS/MF nº 13/2026", "https://www.gov.br/inss/pt-br/direitos-e-deveres/inscricao-e-contribuicao/tabela-de-contribuicao-mensal", None),
    ("inss_band_1_rate", "INSS 1ª faixa - alíquota", "7.5", "percent",
     "Portaria Interministerial MPS/MF nº 13/2026", "https://www.gov.br/inss/pt-br/direitos-e-deveres/inscricao-e-contribuicao/tabela-de-contribuicao-mensal", None),
    ("inss_band_2_limit", "INSS 2ª faixa - limite", "2902.84", "money",
     "Portaria Interministerial MPS/MF nº 13/2026", "https://www.gov.br/inss/pt-br/direitos-e-deveres/inscricao-e-contribuicao/tabela-de-contribuicao-mensal", None),
    ("inss_band_2_rate", "INSS 2ª faixa - alíquota", "9", "percent",
     "Portaria Interministerial MPS/MF nº 13/2026", "https://www.gov.br/inss/pt-br/direitos-e-deveres/inscricao-e-contribuicao/tabela-de-contribuicao-mensal", None),
    ("inss_band_3_limit", "INSS 3ª faixa - limite", "4354.27", "money",
     "Portaria Interministerial MPS/MF nº 13/2026", "https://www.gov.br/inss/pt-br/direitos-e-deveres/inscricao-e-contribuicao/tabela-de-contribuicao-mensal", None),
    ("inss_band_3_rate", "INSS 3ª faixa - alíquota", "12", "percent",
     "Portaria Interministerial MPS/MF nº 13/2026", "https://www.gov.br/inss/pt-br/direitos-e-deveres/inscricao-e-contribuicao/tabela-de-contribuicao-mensal", None),
    ("inss_band_4_limit", "INSS 4ª faixa / teto de contribuição", "8475.55", "money",
     "Portaria Interministerial MPS/MF nº 13/2026", "https://www.gov.br/inss/pt-br/direitos-e-deveres/inscricao-e-contribuicao/tabela-de-contribuicao-mensal", None),
    ("inss_band_4_rate", "INSS 4ª faixa - alíquota", "14", "percent",
     "Portaria Interministerial MPS/MF nº 13/2026", "https://www.gov.br/inss/pt-br/direitos-e-deveres/inscricao-e-contribuicao/tabela-de-contribuicao-mensal", None),

    # IRRF mensal 2026
    ("irrf_band_1_limit", "IRRF 1ª faixa - limite", "2428.80", "money",
     "Lei nº 15.191/2025 e Lei nº 15.270/2025", "https://www.gov.br/receitafederal/pt-br/assuntos/meu-imposto-de-renda/tabelas", None),
    ("irrf_band_2_limit", "IRRF 2ª faixa - limite", "2826.65", "money",
     "Lei nº 15.191/2025 e Lei nº 15.270/2025", "https://www.gov.br/receitafederal/pt-br/assuntos/meu-imposto-de-renda/tabelas", None),
    ("irrf_band_2_rate", "IRRF 2ª faixa - alíquota", "7.5", "percent",
     "Lei nº 15.191/2025", "https://www.gov.br/receitafederal/pt-br/assuntos/meu-imposto-de-renda/tabelas", None),
    ("irrf_band_2_deduction", "IRRF 2ª faixa - parcela a deduzir", "182.16", "money",
     "Lei nº 15.191/2025", "https://www.gov.br/receitafederal/pt-br/assuntos/meu-imposto-de-renda/tabelas", None),
    ("irrf_band_3_limit", "IRRF 3ª faixa - limite", "3751.05", "money",
     "Lei nº 15.191/2025", "https://www.gov.br/receitafederal/pt-br/assuntos/meu-imposto-de-renda/tabelas", None),
    ("irrf_band_3_rate", "IRRF 3ª faixa - alíquota", "15", "percent",
     "Lei nº 15.191/2025", "https://www.gov.br/receitafederal/pt-br/assuntos/meu-imposto-de-renda/tabelas", None),
    ("irrf_band_3_deduction", "IRRF 3ª faixa - parcela a deduzir", "394.16", "money",
     "Lei nº 15.191/2025", "https://www.gov.br/receitafederal/pt-br/assuntos/meu-imposto-de-renda/tabelas", None),
    ("irrf_band_4_limit", "IRRF 4ª faixa - limite", "4664.68", "money",
     "Lei nº 15.191/2025", "https://www.gov.br/receitafederal/pt-br/assuntos/meu-imposto-de-renda/tabelas", None),
    ("irrf_band_4_rate", "IRRF 4ª faixa - alíquota", "22.5", "percent",
     "Lei nº 15.191/2025", "https://www.gov.br/receitafederal/pt-br/assuntos/meu-imposto-de-renda/tabelas", None),
    ("irrf_band_4_deduction", "IRRF 4ª faixa - parcela a deduzir", "675.49", "money",
     "Lei nº 15.191/2025", "https://www.gov.br/receitafederal/pt-br/assuntos/meu-imposto-de-renda/tabelas", None),
    ("irrf_band_5_rate", "IRRF 5ª faixa - alíquota", "27.5", "percent",
     "Lei nº 15.191/2025", "https://www.gov.br/receitafederal/pt-br/assuntos/meu-imposto-de-renda/tabelas", None),
    ("irrf_band_5_deduction", "IRRF 5ª faixa - parcela a deduzir", "908.73", "money",
     "Lei nº 15.191/2025", "https://www.gov.br/receitafederal/pt-br/assuntos/meu-imposto-de-renda/tabelas", None),
    ("irrf_dependent_deduction", "IRRF - dedução mensal por dependente", "189.59", "money",
     "Receita Federal - tabela mensal", "https://www27.receita.fazenda.gov.br/simulador-irpf/", None),
    ("irrf_simplified_discount", "IRRF - desconto simplificado mensal", "607.20", "money",
     "Lei nº 15.270/2025", "https://www.gov.br/receitafederal/pt-br/assuntos/meu-imposto-de-renda/tabelas/exemplos-de-aplicacao-da-lei-15-270-2025", None),
    ("irrf_reduction_zero_limit", "IRRF 2026 - renda mensal com redução integral até", "5000.00", "money",
     "Lei nº 15.270/2025", "https://www.gov.br/receitafederal/pt-br/assuntos/meu-imposto-de-renda/tabelas/exemplos-de-aplicacao-da-lei-15-270-2025", None),
    ("irrf_reduction_end_limit", "IRRF 2026 - limite de renda para redução parcial", "7350.00", "money",
     "Lei nº 15.270/2025", "https://www.gov.br/receitafederal/pt-br/assuntos/meu-imposto-de-renda/tabelas/exemplos-de-aplicacao-da-lei-15-270-2025", None),
    ("irrf_reduction_formula_constant", "IRRF 2026 - constante da redução parcial", "978.62", "money",
     "Lei nº 15.270/2025", "https://www.gov.br/receitafederal/pt-br/assuntos/meu-imposto-de-renda/tabelas/exemplos-de-aplicacao-da-lei-15-270-2025", None),
    ("irrf_reduction_formula_factor", "IRRF 2026 - fator da redução parcial", "0.133145", "factor",
     "Lei nº 15.270/2025", "https://www.gov.br/receitafederal/pt-br/assuntos/meu-imposto-de-renda/tabelas/exemplos-de-aplicacao-da-lei-15-270-2025", None),

    # Salário-família
    ("salary_family_income_limit", "Salário-família - limite de remuneração", "1980.38", "money",
     "Portaria Interministerial MPS/MF nº 13/2026", "https://www.gov.br/inss/pt-br/direitos-e-deveres/salario-familia/valor-limite-para-direito-ao-salario-familia", None),
    ("salary_family_quota", "Salário-família - cota por dependente elegível", "67.54", "money",
     "Portaria Interministerial MPS/MF nº 13/2026", "https://www.gov.br/inss/pt-br/direitos-e-deveres/salario-familia/valor-limite-para-direito-ao-salario-familia", None),
]


def _ensure_payroll_2026_parameters():
    effective = date(2026, 1, 1)
    changed = False
    for code, description, value, value_type, legal_reference, source_url, notes in PAYROLL_2026_LEGAL_PARAMETERS:
        exists = PayrollLegalParameter.query.filter_by(code=code, effective_from=effective).first()
        if exists:
            continue
        db.session.add(PayrollLegalParameter(
            code=code,
            description=description,
            value=Decimal(value),
            value_type=value_type,
            effective_from=effective,
            legal_reference=legal_reference,
            source_url=source_url,
            notes=notes,
            created_by=current_user.id if current_user.is_authenticated else None,
        ))
        changed = True
    if changed:
        db.session.commit()


def _ensure_default_payroll_rubrics():
    defaults = [
        ("SALARIO", "Salário-base", "earning", True, True, True, None),
        ("HE50", "Hora extra 50%", "earning", True, True, True, Decimal("50")),
        ("HE65", "Hora extra 65%", "earning", True, True, True, Decimal("65")),
        ("HE100", "Hora extra 100%", "earning", True, True, True, Decimal("100")),
        ("ADNOT", "Adicional noturno", "earning", True, True, True, Decimal("20")),
        ("INSAL10", "Insalubridade 10%", "earning", True, True, True, Decimal("10")),
        ("INSAL20", "Insalubridade 20%", "earning", True, True, True, Decimal("20")),
        ("INSAL40", "Insalubridade 40%", "earning", True, True, True, Decimal("40")),
        ("PERIC", "Periculosidade 30%", "earning", True, True, True, Decimal("30")),
        ("SALFAM", "Salário-família", "earning", False, False, False, None),
        ("INSS", "INSS empregado", "deduction", False, False, False, None),
        ("IRRF", "IRRF", "deduction", False, False, False, None),
        ("VT", "Vale-transporte", "deduction", False, False, False, None),
        ("VRVA", "Vale-refeição / alimentação", "deduction", False, False, False, None),
        ("PLANO", "Plano de saúde", "deduction", False, False, False, None),
        ("PENSAO", "Pensão alimentícia", "deduction", False, False, False, None),
        ("FALTA", "Faltas / ausências descontáveis", "deduction", True, True, True, None),
        ("DSR_HE", "DSR sobre horas extras", "earning", True, True, True, None),
    ]
    changed = False
    for code, description, nature, inss, fgts, irrf, pct in defaults:
        existing = PayrollRubric.query.filter_by(code=code).first()
        if existing:
            # Corrige rubricas-padrão essenciais também em bancos criados pela V9.3.
            if code == "FALTA":
                existing.inss_incidence = True
                existing.fgts_incidence = True
                existing.irrf_incidence = True
                changed = True
            continue
        db.session.add(PayrollRubric(
            code=code,
            description=description,
            nature=nature,
            inss_incidence=inss,
            fgts_incidence=fgts,
            irrf_incidence=irrf,
            default_percentage=pct,
            created_by=current_user.id if current_user.is_authenticated else None,
        ))
        changed = True
    if changed:
        db.session.commit()


def _approved_future_bank_minutes(employee_id, from_date=None):
    """Horas de banco aprovadas para uso futuro: reservadas, mas ainda não debitadas."""
    from_date = from_date or today_local()
    rows = Request.query.filter(
        Request.employee_id == employee_id,
        Request.request_type == "bank_use",
        Request.status == "approved",
        Request.request_date > from_date
    ).all()
    return sum(int(r.minutes or 0) for r in rows)

def _apply_due_bank_uses(employee):
    """Efetiva débitos aprovados cuja data de utilização chegou, uma única vez."""
    due = Request.query.filter(
        Request.employee_id == employee.id,
        Request.request_type == "bank_use",
        Request.status == "approved",
        Request.request_date <= today_local(),
        Request.bank_effect_applied == False
    ).order_by(Request.request_date.asc()).all()
    changed = False
    for item in due:
        employee.bank_minutes -= int(item.minutes or 0)
        item.bank_effect_applied = True
        item.bank_effect_applied_at = now_local()
        log_action("efetivou utilização de banco de horas", "request", item.id,
                   f"{employee.full_name}: débito de {int(item.minutes or 0)} min na data {item.request_date.strftime('%d/%m/%Y')}")
        changed = True
    if changed:
        db.session.commit()
    return changed

def _bank_summary(employee):
    _apply_due_bank_uses(employee)
    scheduled = _approved_future_bank_minutes(employee.id)
    return {
        "current": int(employee.bank_minutes or 0),
        "scheduled": scheduled,
        "available": int(employee.bank_minutes or 0) - scheduled,
    }

def allowed_file(name):
    return "." in name and name.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS

def save_upload(file):
    if not file or not file.filename or not allowed_file(file.filename):
        raise ValueError("Envie um arquivo PDF ou imagem válida.")

    original = _safe_original_filename(file.filename)
    if "." not in original:
        raise ValueError("Arquivo sem extensão válida.")
    ext = original.rsplit(".", 1)[1].lower()

    _validate_uploaded_content(file, ext)

    stored = f"{uuid.uuid4().hex}.{ext}"
    path = os.path.join(current_app.config["UPLOAD_FOLDER"], stored)
    file.save(path)
    _chmod_private(path)
    return original, stored


PROFILE_PHOTO_EXTENSIONS = {"png", "jpg", "jpeg", "webp"}

def save_profile_photo(file):
    if not file or not file.filename:
        return None
    filename = _safe_original_filename(file.filename, "foto")
    if "." not in filename:
        raise ValueError("A foto deve ser JPG, JPEG, PNG ou WEBP.")
    ext = filename.rsplit(".", 1)[1].lower()
    if ext not in PROFILE_PHOTO_EXTENSIONS:
        raise ValueError("A foto deve ser JPG, JPEG, PNG ou WEBP.")

    _validate_image_stream(file)

    stored = f"profile_{uuid.uuid4().hex}.{ext}"
    path = os.path.join(current_app.config["UPLOAD_FOLDER"], stored)
    file.save(path)
    _chmod_private(path)
    return stored

def visible_employees():
    if current_user.role == ROLE_ADMIN:
        return Employee.query.order_by(Employee.full_name).all()
    if current_user.role == ROLE_MANAGER and current_user.employee:
        ids = [current_user.employee.id] + [e.id for e in current_user.employee.team]
        return Employee.query.filter(Employee.id.in_(ids)).order_by(Employee.full_name).all()
    return [current_user.employee] if current_user.employee else []


def _notification_items(limit=None):
    """Gera notificações operacionais a partir do estado atual do Portal, sem duplicar dados."""
    if not current_user.is_authenticated:
        return []

    today = today_local()
    items = []

    def add(kind, title, description, endpoint, priority="normal", **values):
        items.append({
            "kind": kind,
            "title": title,
            "description": description,
            "url": url_for(endpoint, **values),
            "priority": priority,
        })

    if current_user.role == ROLE_ADMIN:
        pending_requests = Request.query.filter_by(status="pending").count()
        if pending_requests:
            add("approval", f"{pending_requests} solicitação(ões) aguardando aprovação",
                "Banco de horas, horas extras ou ajustes aguardam decisão.",
                "rh.approvals", "high")

        certs = MedicalCertificate.query.filter_by(status="recebido").count()
        if certs:
            add("certificate", f"{certs} atestado(s) aguardando conferência",
                "Há documentos recebidos ainda não processados pelo RH.",
                "rh.certificates_manage", "high")

        signed_documents = DocumentSignatureFlow.query.filter(
            DocumentSignatureFlow.signed_at.isnot(None),
            DocumentSignatureFlow.finalized_at.is_(None),
            DocumentSignatureFlow.cancelled_at.is_(None),
        ).count()
        if signed_documents:
            add("document_signature", f"{signed_documents} documento(s) aguardando validação do RH",
                "Os colaboradores já assinaram e aguardam o arquivamento final.",
                "rh.document_signatures_manage", "high")

        active_employees = Employee.query.filter_by(is_active=True).all()
        incomplete = 0
        for emp in active_employees:
            count = TimeClock.query.filter(
                TimeClock.employee_id == emp.id,
                func.date(TimeClock.punched_at) == today,
            ).count()
            if count not in (0, 4):
                incomplete += 1
        if incomplete:
            add("clock", f"{incomplete} marcação(ões) incompleta(s) hoje",
                "Confira os registros de ponto antes do fechamento do dia.",
                "rh.time_records", "high")

        open_month = 0
        awaiting_signature = 0
        awaiting_validation = 0
        for emp in active_employees:
            closure = TimePeriodClosure.query.filter_by(
                employee_id=emp.id, year=today.year, month=today.month
            ).first()
            if not closure or closure.status != "closed":
                open_month += 1
                continue
            ack = TimeReportAcknowledgement.query.filter_by(
                employee_id=emp.id, year=today.year, month=today.month
            ).first()
            if not ack:
                awaiting_signature += 1
                continue
            final = TimeReportFinalization.query.filter_by(
                employee_id=emp.id, year=today.year, month=today.month
            ).first()
            if not final:
                awaiting_validation += 1

        if open_month:
            add("closing", f"{open_month} competência(s) ainda aberta(s)",
                f"Competência {today.strftime('%m/%Y')} ainda precisa ser fechada.",
                "rh.time_closing", "normal", month=today.strftime("%Y-%m"))
        if awaiting_signature:
            add("signature", f"{awaiting_signature} fechamento(s) aguardando assinatura",
                "Os colaboradores ainda precisam conferir e assinar seus relatórios.",
                "rh.time_closing", "normal", month=today.strftime("%Y-%m"))
        if awaiting_validation:
            add("validation", f"{awaiting_validation} assinatura(s) para validar",
                "O colaborador já assinou; falta a validação final do RH.",
                "rh.time_closing", "high", month=today.strftime("%Y-%m"))

        access_pending = Employee.query.filter(
            Employee.is_active.is_(True),
            Employee.point_pin_hash.is_(None),
        ).count()
        if access_pending:
            add("access", f"{access_pending} colaborador(es) sem PIN de ponto",
                "Configure o PIN para liberar marcações e assinaturas eletrônicas.",
                "rh.pending_center", "normal")

    elif current_user.employee:
        emp = current_user.employee

        closures = TimePeriodClosure.query.filter_by(
            employee_id=emp.id, status="closed"
        ).order_by(TimePeriodClosure.year.desc(), TimePeriodClosure.month.desc()).all()
        unack = [
            c for c in closures
            if not TimeReportAcknowledgement.query.filter_by(
                employee_id=emp.id, year=c.year, month=c.month
            ).first()
        ]
        if unack:
            latest = unack[0]
            add("signature", f"Seu fechamento de {latest.month:02d}/{latest.year} aguarda assinatura",
                "Abra o espelho mensal, confira os dados e assine com seu PIN.",
                "main.dashboard", "high")

        pending_documents = (
            DocumentSignatureFlow.query
            .join(Document)
            .filter(
                Document.employee_id == emp.id,
                DocumentSignatureFlow.signed_at.is_(None),
                DocumentSignatureFlow.cancelled_at.is_(None),
            )
            .count()
        )
        if pending_documents:
            add("document_signature", f"{pending_documents} documento(s) aguardando sua assinatura",
                "Abra, confira e assine eletronicamente com seu PIN de ponto.",
                "rh.employee_detail", "high", employee_id=emp.id)

        unread_payslips = Payslip.query.filter(
            Payslip.employee_id == emp.id,
            Payslip.employee_viewed_at.is_(None),
        ).count()
        if unread_payslips:
            add("payslip", f"{unread_payslips} holerite(s) ainda não visualizado(s)",
                "Há novos documentos de pagamento disponíveis para consulta.",
                "rh.payslips", "normal")

        pending_requests = Request.query.filter_by(
            employee_id=emp.id, status="pending"
        ).count()
        if pending_requests:
            add("request", f"{pending_requests} solicitação(ões) em análise",
                "Acompanhe o andamento das suas solicitações.",
                "rh.requests_page", "normal")

        next_vacation = VacationSchedule.query.filter(
            VacationSchedule.employee_id == emp.id,
            VacationSchedule.status == "planned",
            VacationSchedule.planned_start >= today,
        ).order_by(VacationSchedule.planned_start.asc()).first()
        if next_vacation and (next_vacation.planned_start - today).days <= 30:
            add("vacation", "Suas férias estão próximas",
                f"Início previsto em {next_vacation.planned_start.strftime('%d/%m/%Y')}.",
                "rh.employee_detail", "normal", employee_id=emp.id)

    priority_order = {"high": 0, "normal": 1}
    items.sort(key=lambda item: priority_order.get(item["priority"], 9))
    return items[:limit] if limit else items


@bp.app_context_processor
def inject_portal_notifications():
    if not current_user.is_authenticated:
        return {"portal_notification_count": 0, "portal_notifications_preview": []}
    items = _notification_items()
    return {
        "portal_notification_count": len(items),
        "portal_notifications_preview": items[:5],
    }


@bp.route("/notifications")
@login_required
def notifications_center():
    return render_template(
        "notifications.html",
        notifications=_notification_items(),
    )


@bp.route("/calendar")
@login_required
@roles_required(ROLE_ADMIN)
def hr_calendar():
    raw_month = (request.args.get("month") or today_local().strftime("%Y-%m")).strip()
    try:
        year, month = [int(x) for x in raw_month.split("-", 1)]
        if month < 1 or month > 12:
            raise ValueError
    except (ValueError, TypeError):
        year, month = today_local().year, today_local().month

    first_day = date(year, month, 1)
    last_day = date(year, month, calendar.monthrange(year, month)[1])

    events_by_day = {day: [] for day in range(1, last_day.day + 1)}

    vacations = VacationSchedule.query.filter(
        VacationSchedule.status == "planned",
        VacationSchedule.planned_start <= last_day,
    ).order_by(VacationSchedule.planned_start.asc()).all()
    for item in vacations:
        return_date = item.planned_return or (item.planned_start + timedelta(days=max(item.days - 1, 0)))
        start = max(item.planned_start, first_day)
        end = min(return_date, last_day)
        cursor = start
        while cursor <= end:
            events_by_day[cursor.day].append({
                "type": "vacation",
                "label": f"Férias · {item.employee.full_name}",
            })
            cursor += timedelta(days=1)

    active_employees = Employee.query.filter_by(is_active=True).order_by(Employee.full_name).all()
    for emp in active_employees:
        if emp.birth_date and emp.birth_date.month == month:
            day = min(emp.birth_date.day, last_day.day)
            events_by_day[day].append({"type": "birthday", "label": f"Aniversário · {emp.full_name}"})
        if emp.admission_date and emp.admission_date.month == month:
            day = min(emp.admission_date.day, last_day.day)
            years = year - emp.admission_date.year
            if years >= 1:
                events_by_day[day].append({
                    "type": "admission",
                    "label": f"{years} ano(s) de casa · {emp.full_name}",
                })

    weeks = calendar.Calendar(firstweekday=0).monthdayscalendar(year, month)
    previous_month = (first_day - timedelta(days=1)).strftime("%Y-%m")
    next_month = (last_day + timedelta(days=1)).strftime("%Y-%m")

    return render_template(
        "hr_calendar.html",
        year=year,
        month=month,
        month_value=f"{year:04d}-{month:02d}",
        month_label=_month_name_pt(month),
        weeks=weeks,
        events_by_day=events_by_day,
        previous_month=previous_month,
        next_month=next_month,
        today=today_local(),
    )


@bp.route("/reports")
@login_required
@roles_required(ROLE_ADMIN)
def reports_center():
    return render_template("reports_center.html", today=today_local())


@bp.route("/reports/employees.xlsx")
@login_required
@roles_required(ROLE_ADMIN)
def report_employees_xlsx():
    employees = Employee.query.order_by(Employee.full_name.asc()).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Colaboradores"
    ws.append([
        "Nome", "CPF", "Matrícula", "Cargo", "Setor", "Projeto",
        "Admissão", "Contrato", "Carga semanal", "Status", "E-mail"
    ])
    for emp in employees:
        ws.append([
            emp.full_name,
            emp.cpf,
            emp.registration or "",
            emp.job_title,
            emp.department,
            emp.project,
            emp.admission_date.strftime("%d/%m/%Y") if emp.admission_date else "",
            emp.contract_type or "",
            emp.weekly_hours,
            "Ativo" if emp.is_active else "Inativo",
            emp.user.email if emp.user else "",
        ])
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    log_action("exportou relatório de colaboradores", "report", None, f"{len(employees)} registros")
    db.session.commit()
    return send_file(
        output,
        as_attachment=True,
        download_name=f"colaboradores-{today_local().strftime('%Y-%m-%d')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@bp.route("/reports/bank.xlsx")
@login_required
@roles_required(ROLE_ADMIN)
def report_bank_xlsx():
    employees = Employee.query.filter_by(is_active=True).order_by(Employee.full_name.asc()).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Banco de Horas"
    ws.append(["Colaborador", "Cargo", "Setor", "Projeto", "Saldo em minutos", "Saldo HH:MM"])
    for emp in employees:
        minutes = int(emp.bank_minutes or 0)
        sign = "-" if minutes < 0 else ""
        absolute = abs(minutes)
        hhmm = f"{sign}{absolute // 60:02d}:{absolute % 60:02d}"
        ws.append([emp.full_name, emp.job_title, emp.department, emp.project, minutes, hhmm])
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    log_action("exportou relatório de banco de horas", "report", None, f"{len(employees)} colaboradores ativos")
    db.session.commit()
    return send_file(
        output,
        as_attachment=True,
        download_name=f"banco-de-horas-{today_local().strftime('%Y-%m-%d')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )



@bp.route("/pending-center")
@login_required
@roles_required(ROLE_ADMIN)
def pending_center():
    """Caixa de entrada operacional do RH, consolidando pendências do Portal."""
    today = today_local()
    active_employees = Employee.query.filter_by(is_active=True).order_by(Employee.full_name).all()
    employee_ids = [e.id for e in active_employees]

    pending_requests = (
        Request.query
        .filter_by(status="pending")
        .order_by(Request.requested_at.asc())
        .all()
    )

    pending_certificates = (
        MedicalCertificate.query
        .filter_by(status="recebido")
        .order_by(MedicalCertificate.uploaded_at.asc())
        .all()
    )

    incomplete_punches = []
    for emp in active_employees:
        today_rows = (
            TimeClock.query
            .filter(
                TimeClock.employee_id == emp.id,
                func.date(TimeClock.punched_at) == today,
            )
            .order_by(TimeClock.punched_at.asc())
            .all()
        )
        if today_rows and len(today_rows) != 4:
            incomplete_punches.append({
                "employee": emp,
                "count": len(today_rows),
                "last": today_rows[-1] if today_rows else None,
            })

    open_closures = []
    awaiting_signatures = []
    awaiting_rh_validation = []

    for emp in active_employees:
        closure = TimePeriodClosure.query.filter_by(
            employee_id=emp.id,
            year=today.year,
            month=today.month,
        ).first()

        if not closure or closure.status != "closed":
            open_closures.append({"employee": emp, "closure": closure})
            continue

        ack = TimeReportAcknowledgement.query.filter_by(
            employee_id=emp.id,
            year=today.year,
            month=today.month,
        ).first()

        if not ack:
            awaiting_signatures.append({"employee": emp, "closure": closure})
            continue

        finalization = TimeReportFinalization.query.filter_by(
            employee_id=emp.id,
            year=today.year,
            month=today.month,
        ).first()
        if not finalization:
            awaiting_rh_validation.append({
                "employee": emp,
                "closure": closure,
                "ack": ack,
            })

    access_pending = []
    for emp in active_employees:
        reasons = []
        if not emp.point_pin_hash:
            reasons.append("PIN de ponto não configurado")
        if emp.user and emp.user.must_change_password:
            reasons.append("troca da senha provisória pendente")
        if reasons:
            access_pending.append({"employee": emp, "reasons": reasons})

    unread_payslips = (
        Payslip.query
        .filter(
            Payslip.employee_id.in_(employee_ids),
            Payslip.employee_viewed_at.is_(None),
        )
        .order_by(Payslip.year.desc(), Payslip.month.desc(), Payslip.uploaded_at.asc())
        .all()
    ) if employee_ids else []

    next_30 = today + timedelta(days=30)
    upcoming_vacations = (
        VacationSchedule.query
        .filter(
            VacationSchedule.employee_id.in_(employee_ids),
            VacationSchedule.status == "planned",
            VacationSchedule.planned_start >= today,
            VacationSchedule.planned_start <= next_30,
        )
        .order_by(VacationSchedule.planned_start.asc())
        .all()
    ) if employee_ids else []

    document_signatures_pending_rh = (
        DocumentSignatureFlow.query
        .filter(
            DocumentSignatureFlow.signed_at.isnot(None),
            DocumentSignatureFlow.finalized_at.is_(None),
            DocumentSignatureFlow.cancelled_at.is_(None),
        )
        .order_by(DocumentSignatureFlow.signed_at.asc())
        .all()
    )

    document_signatures_pending_employee = (
        DocumentSignatureFlow.query
        .filter(
            DocumentSignatureFlow.signed_at.is_(None),
            DocumentSignatureFlow.cancelled_at.is_(None),
        )
        .order_by(DocumentSignatureFlow.requested_at.asc())
        .all()
    )

    total_actionable = (
        len(pending_requests)
        + len(pending_certificates)
        + len(incomplete_punches)
        + len(open_closures)
        + len(awaiting_signatures)
        + len(awaiting_rh_validation)
        + len(access_pending)
        + len(document_signatures_pending_rh)
    )

    return render_template(
        "pending_center.html",
        today=today,
        pending_requests=pending_requests,
        pending_certificates=pending_certificates,
        incomplete_punches=incomplete_punches,
        open_closures=open_closures,
        awaiting_signatures=awaiting_signatures,
        awaiting_rh_validation=awaiting_rh_validation,
        access_pending=access_pending,
        unread_payslips=unread_payslips,
        upcoming_vacations=upcoming_vacations,
        document_signatures_pending_rh=document_signatures_pending_rh,
        document_signatures_pending_employee=document_signatures_pending_employee,
        total_actionable=total_actionable,
    )




def _money(value):
    return Decimal(value or 0).quantize(Decimal("0.01"))


def _param(code, ref_date):
    row = (
        PayrollLegalParameter.query
        .filter(
            PayrollLegalParameter.code == code,
            PayrollLegalParameter.effective_from <= ref_date,
            or_(PayrollLegalParameter.effective_to.is_(None), PayrollLegalParameter.effective_to >= ref_date),
        )
        .order_by(PayrollLegalParameter.effective_from.desc())
        .first()
    )
    if not row:
        raise ValueError(f"Parâmetro legal não encontrado para {code} em {ref_date.strftime('%m/%Y')}.")
    return Decimal(row.value)


def _payroll_period(year, month):
    start = date(year, month, 1)
    end = date(year, month, calendar.monthrange(year, month)[1])
    return start, end


def _payroll_hourly_rate(emp, salary):
    weekly = Decimal(str(emp.weekly_hours or 44))
    divisor = weekly * Decimal("5")
    if divisor <= 0:
        divisor = Decimal("220")
    return (Decimal(salary) / divisor).quantize(Decimal("0.000001")), divisor


def _calc_inss_2026(base, ref_date):
    base = max(Decimal("0"), Decimal(base))
    limits = [
        _param("inss_band_1_limit", ref_date),
        _param("inss_band_2_limit", ref_date),
        _param("inss_band_3_limit", ref_date),
        _param("inss_band_4_limit", ref_date),
    ]
    rates = [
        _param("inss_band_1_rate", ref_date) / 100,
        _param("inss_band_2_rate", ref_date) / 100,
        _param("inss_band_3_rate", ref_date) / 100,
        _param("inss_band_4_rate", ref_date) / 100,
    ]
    capped = min(base, limits[-1])
    previous = Decimal("0")
    total = Decimal("0")
    detail = []
    for idx, (limit, rate) in enumerate(zip(limits, rates), 1):
        portion = min(capped, limit) - previous
        if portion > 0:
            contribution = (portion * rate).quantize(Decimal("0.01"))
            total += contribution
            detail.append({"band": idx, "base": portion, "rate": rate * 100, "amount": contribution})
        previous = limit
        if capped <= limit:
            break
    return total.quantize(Decimal("0.01")), detail


def _calc_irrf_2026(taxable_income, inss_amount, dependent_count, pension_deduction, ref_date):
    taxable_income = max(Decimal("0"), Decimal(taxable_income))
    dep_deduction = _param("irrf_dependent_deduction", ref_date) * Decimal(int(dependent_count or 0))
    simplified = _param("irrf_simplified_discount", ref_date)
    legal_deductions = Decimal(inss_amount) + dep_deduction + max(Decimal("0"), Decimal(pension_deduction or 0))
    chosen_deduction = max(simplified, legal_deductions)
    deduction_method = "simplificado" if simplified >= legal_deductions else "deduções legais"
    base = max(Decimal("0"), taxable_income - chosen_deduction)

    l1 = _param("irrf_band_1_limit", ref_date)
    l2 = _param("irrf_band_2_limit", ref_date)
    l3 = _param("irrf_band_3_limit", ref_date)
    l4 = _param("irrf_band_4_limit", ref_date)
    if base <= l1:
        rate, deduction = Decimal("0"), Decimal("0")
    elif base <= l2:
        rate, deduction = _param("irrf_band_2_rate", ref_date)/100, _param("irrf_band_2_deduction", ref_date)
    elif base <= l3:
        rate, deduction = _param("irrf_band_3_rate", ref_date)/100, _param("irrf_band_3_deduction", ref_date)
    elif base <= l4:
        rate, deduction = _param("irrf_band_4_rate", ref_date)/100, _param("irrf_band_4_deduction", ref_date)
    else:
        rate, deduction = _param("irrf_band_5_rate", ref_date)/100, _param("irrf_band_5_deduction", ref_date)

    before_reduction = max(Decimal("0"), (base * rate - deduction).quantize(Decimal("0.01")))
    zero_limit = _param("irrf_reduction_zero_limit", ref_date)
    end_limit = _param("irrf_reduction_end_limit", ref_date)
    if taxable_income <= zero_limit:
        reduction = before_reduction
    elif taxable_income <= end_limit:
        constant = _param("irrf_reduction_formula_constant", ref_date)
        factor = _param("irrf_reduction_formula_factor", ref_date)
        reduction = max(Decimal("0"), constant - factor * taxable_income)
        reduction = min(before_reduction, reduction.quantize(Decimal("0.01")))
    else:
        reduction = Decimal("0")
    tax = max(Decimal("0"), before_reduction - reduction).quantize(Decimal("0.01"))
    return tax, {
        "taxable_income": taxable_income,
        "legal_deductions": legal_deductions.quantize(Decimal("0.01")),
        "simplified": simplified.quantize(Decimal("0.01")),
        "chosen_deduction": chosen_deduction.quantize(Decimal("0.01")),
        "deduction_method": deduction_method,
        "base": base.quantize(Decimal("0.01")),
        "rate": rate * 100,
        "table_deduction": deduction,
        "before_reduction": before_reduction,
        "reduction": reduction.quantize(Decimal("0.01")),
    }


def _month_data_warnings(emp, year, month):
    start, end = _payroll_period(year, month)
    warnings = []
    if not emp.payroll_config or Decimal(emp.payroll_config.monthly_salary or 0) <= 0:
        warnings.append("Salário-base não configurado.")
    if emp.admission_date and start <= emp.admission_date <= end:
        warnings.append("Admissão ocorreu nesta competência: confira eventual proporcionalidade antes do fechamento oficial.")
    approved_overtime = Request.query.filter(
        Request.employee_id == emp.id,
        Request.request_type == "overtime",
        Request.status == "approved",
        Request.request_date >= start,
        Request.request_date <= end,
    ).all()
    if approved_overtime:
        minutes = sum(int(x.minutes or 0) for x in approved_overtime)
        warnings.append(
            f"Há {minutes//60:02d}:{minutes%60:02d} de horas extras aprovadas no período. "
            "Elas já alimentam o banco de horas do Portal e não são pagas automaticamente; classifique como evento de folha somente se houver pagamento em dinheiro."
        )
    duties = WeekendDuty.query.filter(
        WeekendDuty.employee_id == emp.id,
        WeekendDuty.duty_date >= start,
        WeekendDuty.duty_date <= end,
    ).all()
    if duties:
        minutes = sum(int(x.minutes or 0) for x in duties)
        warnings.append(
            f"Há {minutes//60:02d}:{minutes%60:02d} de plantões registrados no banco. Não foram incluídos automaticamente como pagamento."
        )
    cert_minutes = (
        db.session.query(func.coalesce(func.sum(MedicalCertificateAllowance.minutes), 0))
        .join(MedicalCertificate, MedicalCertificate.id == MedicalCertificateAllowance.certificate_id)
        .filter(
            MedicalCertificateAllowance.employee_id == emp.id,
            MedicalCertificate.start_date <= end,
            MedicalCertificate.start_date >= start,
        ).scalar() or 0
    )
    if cert_minutes:
        warnings.append(f"Atestados no período justificam {int(cert_minutes)//60:02d}:{int(cert_minutes)%60:02d}. Nenhum desconto foi criado automaticamente.")
    return warnings


def _calculate_payroll_employee(competence, emp):
    start, end = _payroll_period(competence.year, competence.month)
    ref_date = end
    config = emp.payroll_config
    if not config or Decimal(config.monthly_salary or 0) <= 0:
        return None, ["Salário-base não configurado."]

    salary = _money(config.monthly_salary)
    hourly, divisor = _payroll_hourly_rate(emp, salary)
    existing = PayrollEmployeeCalculation.query.filter_by(competence_id=competence.id, employee_id=emp.id).first()
    if existing:
        db.session.delete(existing)
        db.session.flush()
    calc = PayrollEmployeeCalculation(
        competence_id=competence.id,
        employee_id=emp.id,
        base_salary=salary,
        hourly_rate=hourly,
    )
    db.session.add(calc); db.session.flush()

    items=[]
    def add_item(code, description, nature, amount, reference=None, source="engine", inss=False, irrf=False, fgts=False, order=100):
        amount=_money(amount)
        row=PayrollCalculationItem(
            calculation_id=calc.id, rubric_code=code, description=description, nature=nature,
            reference=reference, amount=amount, source=source,
            inss_incidence=inss, irrf_incidence=irrf, fgts_incidence=fgts, sort_order=order,
        )
        db.session.add(row); items.append(row); return row

    # Salário-base integral. Competências com admissão no mês ficam sinalizadas para conferência.
    add_item("SALARIO", "Salário-base", "earning", salary, "Mensal", "engine", True, True, True, 10)

    # Eventos variáveis confirmados pelo RH.
    events=PayrollManualEvent.query.filter_by(competence_id=competence.id, employee_id=emp.id).all()
    for ev in events:
        rub=ev.rubric
        amount=Decimal(ev.amount) if ev.amount is not None else None
        qty=Decimal(ev.reference_quantity) if ev.reference_quantity is not None else None
        ref=ev.reference_label or None
        if amount is None and qty is not None and rub.code.startswith("HE") and rub.default_percentage is not None:
            pct=Decimal(rub.default_percentage)
            amount=(hourly * qty * (Decimal("1") + pct/100)).quantize(Decimal("0.01"))
            ref=ref or f"{qty}h × R$ {hourly:.4f} × {(Decimal('1')+pct/100):.4f}"
        elif amount is None and qty is not None and rub.code == "FALTA":
            amount=(salary/Decimal("30")*qty).quantize(Decimal("0.01"))
            ref=ref or f"{qty} dia(s) × salário/30"
        elif amount is None:
            amount=Decimal("0")
        add_item(
            rub.code, rub.description, rub.nature, amount, ref, "manual",
            rub.inss_incidence, rub.irrf_incidence, rub.fgts_incidence, 40,
        )

    # Salário-família: somente dependentes marcados pelo RH como elegíveis.
    eligible=sum(1 for dep in emp.payroll_dependents if dep.active and dep.salary_family_eligible)
    income_for_family=sum(i.amount for i in items if i.nature=="earning" and i.irrf_incidence)
    sf_limit=_param("salary_family_income_limit",ref_date)
    if eligible and income_for_family <= sf_limit:
        quota=_param("salary_family_quota",ref_date)
        add_item("SALFAM","Salário-família","earning",quota*eligible,f"{eligible} cota(s)","engine",False,False,False,50)

    # Bases de INSS e IRRF vêm das incidências das rubricas.
    inss_base=sum(i.amount if i.nature=="earning" else -i.amount for i in items if i.inss_incidence)
    inss_base=max(Decimal("0"),inss_base)
    inss_amount,inss_detail=_calc_inss_2026(inss_base,ref_date)
    add_item("INSS","INSS empregado","deduction",inss_amount,"Progressivo por faixa","engine",False,False,False,80)

    irrf_taxable=sum(i.amount if i.nature=="earning" else -i.amount for i in items if i.irrf_incidence)
    irrf_taxable=max(Decimal("0"),irrf_taxable)
    dep_count=sum(1 for dep in emp.payroll_dependents if dep.active and dep.irrf_dependent)
    pension=_money(config.pension_discount_value)
    irrf_amount,irrf_detail=_calc_irrf_2026(irrf_taxable,inss_amount,dep_count,pension,ref_date)
    if irrf_amount > 0:
        add_item("IRRF","IRRF","deduction",irrf_amount,f"Base R$ {irrf_detail['base']:.2f}","engine",False,False,False,90)

    # Descontos fixos configurados.
    vt=Decimal("0")
    if config.has_transport_voucher and Decimal(config.transport_discount_percent or 0)>0:
        vt=(salary*Decimal(config.transport_discount_percent)/100).quantize(Decimal("0.01"))
        add_item("VT","Vale-transporte","deduction",vt,f"{Decimal(config.transport_discount_percent):.2f}% do salário-base","config",False,False,False,100)
    if Decimal(config.food_discount_value or 0)>0:
        add_item("VRVA","Vale-refeição / alimentação","deduction",config.food_discount_value,"Valor fixo","config",False,False,False,101)
    if Decimal(config.health_plan_discount_value or 0)>0:
        add_item("PLANO","Plano de saúde","deduction",config.health_plan_discount_value,"Valor fixo","config",False,False,False,102)
    if pension>0:
        add_item("PENSAO","Pensão alimentícia","deduction",pension,"Valor configurado","config",False,False,False,103)
    if Decimal(config.other_fixed_discount_value or 0)>0:
        add_item("OUTRO","Outro desconto fixo","deduction",config.other_fixed_discount_value,config.other_fixed_discount_description or "Valor fixo","config",False,False,False,104)

    gross=sum(i.amount for i in items if i.nature=="earning")
    deductions=sum(i.amount for i in items if i.nature=="deduction")
    net=gross-deductions
    warnings=_month_data_warnings(emp,competence.year,competence.month)
    if config.has_transport_voucher and Decimal(config.transport_discount_percent or 0)>6:
        warnings.append("Percentual de vale-transporte acima de 6%: confira a parametrização antes de usar a prévia.")

    calc.gross_amount=_money(gross)
    calc.inss_base=_money(inss_base)
    calc.inss_amount=_money(inss_amount)
    calc.irrf_taxable_income=_money(irrf_taxable)
    calc.irrf_base=_money(irrf_detail["base"])
    calc.irrf_amount=_money(irrf_amount)
    calc.deductions_amount=_money(deductions)
    calc.net_amount=_money(net)
    calc.calculation_notes=json.dumps({
        "warnings":warnings,
        "hourly_divisor":str(divisor),
        "inss":[{"band":d["band"],"base":str(d["base"]),"rate":str(d["rate"]),"amount":str(d["amount"])} for d in inss_detail],
        "irrf":{k:str(v) for k,v in irrf_detail.items()},
    },ensure_ascii=False)
    calc.calculated_at=now_local()
    return calc,warnings


def _competence_from_value(value):
    try:
        year,month=[int(x) for x in value.split("-")]
        if not 1<=month<=12: raise ValueError
        return year,month
    except Exception:
        raise ValueError("Competência inválida. Use o formato AAAA-MM.")

@bp.route("/payroll")
@login_required
@roles_required(ROLE_ADMIN)
def payroll_center():
    _ensure_payroll_2026_parameters()
    _ensure_default_payroll_rubrics()
    employees = Employee.query.order_by(Employee.is_active.desc(), Employee.full_name.asc()).all()
    configured = sum(1 for emp in employees if emp.payroll_config and Decimal(emp.payroll_config.monthly_salary or 0) > 0)
    active = sum(1 for emp in employees if emp.is_active)
    dependents = PayrollDependent.query.filter_by(active=True).count()
    rubrics = PayrollRubric.query.filter_by(active=True).order_by(PayrollRubric.code.asc()).all()
    parameters = PayrollLegalParameter.query.order_by(PayrollLegalParameter.effective_from.desc(),PayrollLegalParameter.code.asc()).all()
    recent_competences = PayrollCompetence.query.order_by(PayrollCompetence.year.desc(), PayrollCompetence.month.desc()).limit(12).all()
    return render_template("payroll_center.html", employees=employees, configured=configured, active=active,
        dependents=dependents, rubrics=rubrics, parameters=parameters, recent_competences=recent_competences,
        current_competence=today_local().strftime("%Y-%m"))


@bp.route("/employees/<int:employee_id>/payroll-config", methods=["POST"])
@login_required
@roles_required(ROLE_ADMIN)
def employee_payroll_config(employee_id):
    emp = db.get_or_404(Employee, employee_id)
    try:
        monthly_salary = parse_money(request.form.get("monthly_salary"))
        food_discount = parse_money(request.form.get("food_discount_value"))
        health_discount = parse_money(request.form.get("health_plan_discount_value"))
        pension_discount = parse_money(request.form.get("pension_discount_value"))
        other_discount = parse_money(request.form.get("other_fixed_discount_value"))
        transport_percent = Decimal((request.form.get("transport_discount_percent") or "0").replace(",", "."))
    except (ValueError, InvalidOperation):
        flash("Revise os valores informados na configuração da folha.", "danger")
        return redirect(url_for("rh.employee_detail", employee_id=employee_id) + "#folha")

    if monthly_salary < 0 or transport_percent < 0 or transport_percent > 100:
        flash("Salário e percentuais devem possuir valores válidos.", "danger")
        return redirect(url_for("rh.employee_detail", employee_id=employee_id) + "#folha")

    effective = parse_date(request.form.get("salary_effective_date")) or today_local()
    config = emp.payroll_config
    if not config:
        config = PayrollEmployeeConfig(
            employee_id=emp.id,
            monthly_salary=monthly_salary,
            salary_effective_date=effective,
            updated_by=current_user.id,
        )
        db.session.add(config)

    config.monthly_salary = monthly_salary
    config.salary_effective_date = effective
    config.salary_type = request.form.get("salary_type") or "monthly"
    config.has_transport_voucher = request.form.get("has_transport_voucher") == "1"
    config.transport_discount_percent = transport_percent
    config.food_discount_value = food_discount
    config.health_plan_discount_value = health_discount
    config.pension_discount_value = pension_discount
    config.other_fixed_discount_value = other_discount
    config.other_fixed_discount_description = (request.form.get("other_fixed_discount_description") or "").strip() or None
    config.notes = (request.form.get("notes") or "").strip() or None
    config.updated_by = current_user.id
    config.updated_at = now_local()

    log_action(
        "atualizou configuração de pré-folha",
        "payroll_employee_config",
        emp.id,
        f"{emp.full_name}; salário base R$ {monthly_salary:.2f}; vigência {effective.strftime('%d/%m/%Y')}",
    )
    db.session.commit()
    flash("Configuração remuneratória salva.", "success")
    return redirect(url_for("rh.employee_detail", employee_id=employee_id) + "#folha")


@bp.route("/employees/<int:employee_id>/payroll-dependent", methods=["POST"])
@login_required
@roles_required(ROLE_ADMIN)
def employee_payroll_dependent_add(employee_id):
    emp = db.get_or_404(Employee, employee_id)
    full_name = (request.form.get("full_name") or "").strip()
    if not full_name:
        flash("Informe o nome do dependente.", "danger")
        return redirect(url_for("rh.employee_detail", employee_id=employee_id) + "#folha")

    item = PayrollDependent(
        employee_id=emp.id,
        full_name=full_name,
        cpf=(request.form.get("cpf") or "").strip() or None,
        birth_date=parse_date(request.form.get("birth_date")),
        relationship=(request.form.get("relationship") or "").strip() or None,
        irrf_dependent=request.form.get("irrf_dependent") == "1",
        salary_family_eligible=request.form.get("salary_family_eligible") == "1",
        notes=(request.form.get("notes") or "").strip() or None,
        created_by=current_user.id,
    )
    db.session.add(item)
    db.session.flush()
    log_action(
        "cadastrou dependente para pré-folha",
        "payroll_dependent",
        item.id,
        f"{emp.full_name}; dependente {item.full_name}",
    )
    db.session.commit()
    flash("Dependente cadastrado.", "success")
    return redirect(url_for("rh.employee_detail", employee_id=employee_id) + "#folha")


@bp.route("/payroll-dependents/<int:dependent_id>/delete", methods=["POST"])
@login_required
@roles_required(ROLE_ADMIN)
def employee_payroll_dependent_delete(dependent_id):
    item = db.get_or_404(PayrollDependent, dependent_id)
    employee_id = item.employee_id
    description = f"{item.employee.full_name}; dependente {item.full_name}"
    db.session.delete(item)
    log_action("removeu dependente da pré-folha", "payroll_dependent", dependent_id, description)
    db.session.commit()
    flash("Dependente removido.", "success")
    return redirect(url_for("rh.employee_detail", employee_id=employee_id) + "#folha")


@bp.route("/payroll/rubrics", methods=["POST"])
@login_required
@roles_required(ROLE_ADMIN)
def payroll_rubric_add():
    _ensure_default_payroll_rubrics()
    code = (request.form.get("code") or "").strip().upper()
    description = (request.form.get("description") or "").strip()
    if not code or not description:
        flash("Código e descrição da rubrica são obrigatórios.", "danger")
        return redirect(url_for("rh.payroll_center") + "#rubricas")
    if PayrollRubric.query.filter_by(code=code).first():
        flash("Já existe uma rubrica com esse código.", "danger")
        return redirect(url_for("rh.payroll_center") + "#rubricas")

    pct_raw = (request.form.get("default_percentage") or "").strip().replace(",", ".")
    try:
        pct = Decimal(pct_raw) if pct_raw else None
    except InvalidOperation:
        flash("Percentual padrão inválido.", "danger")
        return redirect(url_for("rh.payroll_center") + "#rubricas")

    item = PayrollRubric(
        code=code,
        description=description,
        nature=request.form.get("nature") or "earning",
        esocial_nature=(request.form.get("esocial_nature") or "").strip() or None,
        inss_incidence=request.form.get("inss_incidence") == "1",
        fgts_incidence=request.form.get("fgts_incidence") == "1",
        irrf_incidence=request.form.get("irrf_incidence") == "1",
        default_percentage=pct,
        notes=(request.form.get("notes") or "").strip() or None,
        created_by=current_user.id,
    )
    db.session.add(item)
    db.session.flush()
    log_action("cadastrou rubrica de pré-folha", "payroll_rubric", item.id, f"{code} - {description}")
    db.session.commit()
    flash("Rubrica cadastrada.", "success")
    return redirect(url_for("rh.payroll_center") + "#rubricas")


@bp.route("/payroll/legal-parameters", methods=["POST"])
@login_required
@roles_required(ROLE_ADMIN)
def payroll_legal_parameter_add():
    code = (request.form.get("code") or "").strip()
    description = (request.form.get("description") or "").strip()
    effective_from = parse_date(request.form.get("effective_from"))
    try:
        value = Decimal((request.form.get("value") or "").strip().replace(",", "."))
    except InvalidOperation:
        value = None

    if not code or not description or not effective_from or value is None:
        flash("Código, descrição, valor e início de vigência são obrigatórios.", "danger")
        return redirect(url_for("rh.payroll_center") + "#parametros")

    if PayrollLegalParameter.query.filter_by(code=code, effective_from=effective_from).first():
        flash("Já existe esse parâmetro com a mesma data de vigência.", "danger")
        return redirect(url_for("rh.payroll_center") + "#parametros")

    item = PayrollLegalParameter(
        code=code,
        description=description,
        value=value,
        value_type=request.form.get("value_type") or "money",
        effective_from=effective_from,
        effective_to=parse_date(request.form.get("effective_to")),
        legal_reference=(request.form.get("legal_reference") or "").strip() or None,
        source_url=(request.form.get("source_url") or "").strip() or None,
        notes=(request.form.get("notes") or "").strip() or None,
        created_by=current_user.id,
    )
    db.session.add(item)
    db.session.flush()
    log_action("cadastrou parâmetro legal de pré-folha", "payroll_legal_parameter", item.id, item.code)
    db.session.commit()
    flash("Parâmetro legal cadastrado.", "success")
    return redirect(url_for("rh.payroll_center") + "#parametros")



def _payroll_closure_is_locked(comp):
    return bool(comp.closure and comp.closure.status in {"closed", "authorized"})


def _fmt_brl(value):
    number = Decimal(value or 0).quantize(Decimal("0.01"))
    raw = f"{number:,.2f}"
    return "R$ " + raw.replace(",", "X").replace(".", ",").replace("X", ".")


def _payroll_logo_path():
    path = os.path.join(current_app.root_path, "static", "img", "alecrim-dourado-logo.png")
    return path if os.path.isfile(path) else None


def _payroll_pdf_styles():
    styles = getSampleStyleSheet()
    gold = colors.HexColor("#D5A515")
    gold_soft = colors.HexColor("#FBF6E7")
    charcoal = colors.HexColor("#242422")
    muted = colors.HexColor("#70736F")
    line = colors.HexColor("#E5E7E3")
    green = colors.HexColor("#2F6B55")
    styles.add(ParagraphStyle(name="PayrollTitle", parent=styles["Heading1"], fontName="Helvetica-Bold", fontSize=16, leading=19, textColor=charcoal, spaceAfter=3))
    styles.add(ParagraphStyle(name="PayrollSub", parent=styles["BodyText"], fontSize=8.2, leading=10.5, textColor=muted))
    styles.add(ParagraphStyle(name="PayrollSmall", parent=styles["BodyText"], fontSize=7.7, leading=9.7, textColor=charcoal))
    styles.add(ParagraphStyle(name="PayrollSmallRight", parent=styles["PayrollSmall"], alignment=TA_RIGHT))
    styles.add(ParagraphStyle(name="PayrollLabel", parent=styles["PayrollSmall"], fontName="Helvetica-Bold", fontSize=7, textColor=muted, leading=8.5))
    styles.add(ParagraphStyle(name="PayrollValue", parent=styles["PayrollSmall"], fontName="Helvetica-Bold", fontSize=9.2, leading=11, textColor=charcoal))
    styles.add(ParagraphStyle(name="PayrollNet", parent=styles["PayrollSmall"], fontName="Helvetica-Bold", fontSize=14, leading=16, textColor=green))
    return styles, gold, gold_soft, charcoal, muted, line, green


def _payroll_header_story(title, subtitle, width=183*mm):
    styles, gold, gold_soft, charcoal, muted, line, green = _payroll_pdf_styles()
    logo = _payroll_logo_path()
    brand=[]
    if logo:
        brand.append(RLImage(logo, width=48*mm, height=7.2*mm))
    else:
        brand.append(Paragraph("ASSOCIACAO ALECRIM DOURADO", styles["PayrollValue"]))
    title_cell=[Paragraph(title,styles["PayrollTitle"]),Paragraph(subtitle,styles["PayrollSub"])]
    header=Table([[brand,title_cell]],colWidths=[width*.38,width*.62])
    header.setStyle(TableStyle([
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),("ALIGN",(1,0),(1,0),"RIGHT"),
        ("LEFTPADDING",(0,0),(-1,-1),0),("RIGHTPADDING",(0,0),(-1,-1),0),
        ("TOPPADDING",(0,0),(-1,-1),0),("BOTTOMPADDING",(0,0),(-1,-1),7),
        ("LINEBELOW",(0,0),(-1,-1),2,gold),
    ]))
    return [header,Spacer(1,7)]


def _payroll_status_card(comp, calculations, width):
    styles,gold,gold_soft,charcoal,muted,line,green=_payroll_pdf_styles()
    closure=comp.closure
    authorized=bool(closure and closure.status=="authorized")
    status="AUTORIZADA" if authorized else "FECHADA - AGUARDANDO AUTORIZACAO"
    total_gross=sum((Decimal(c.gross_amount or 0) for c in calculations),Decimal("0"))
    total_ded=sum((Decimal(c.deductions_amount or 0) for c in calculations),Decimal("0"))
    total_net=sum((Decimal(c.net_amount or 0) for c in calculations),Decimal("0"))
    cells=[
        [Paragraph("STATUS",styles["PayrollLabel"]),Paragraph("COLABORADORES",styles["PayrollLabel"]),Paragraph("PROVENTOS",styles["PayrollLabel"]),Paragraph("DESCONTOS",styles["PayrollLabel"]),Paragraph("LIQUIDO DA FOLHA",styles["PayrollLabel"])],
        [Paragraph(status,styles["PayrollValue"]),Paragraph(str(len(calculations)),styles["PayrollValue"]),Paragraph(_fmt_brl(total_gross),styles["PayrollValue"]),Paragraph(_fmt_brl(total_ded),styles["PayrollValue"]),Paragraph(_fmt_brl(total_net),styles["PayrollNet"])],
    ]
    t=Table(cells,colWidths=[width*.25,width*.13,width*.19,width*.19,width*.24])
    t.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,-1),gold_soft),("BOX",(0,0),(-1,-1),.6,line),
        ("INNERGRID",(0,0),(-1,-1),.4,line),("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ("TOPPADDING",(0,0),(-1,-1),6),("BOTTOMPADDING",(0,0),(-1,-1),6),
        ("LEFTPADDING",(0,0),(-1,-1),7),("RIGHTPADDING",(0,0),(-1,-1),7),
    ]))
    return t


def _payroll_consolidated_pdf(comp):
    calculations=(PayrollEmployeeCalculation.query.filter_by(competence_id=comp.id)
        .join(Employee,Employee.id==PayrollEmployeeCalculation.employee_id)
        .order_by(Employee.full_name.asc()).all())
    if not calculations: raise ValueError("A competência não possui cálculos para gerar a folha.")
    styles,gold,gold_soft,charcoal,muted,line,green=_payroll_pdf_styles()
    buff=BytesIO(); page_width=277*mm
    doc=SimpleDocTemplate(buff,pagesize=landscape(A4),rightMargin=10*mm,leftMargin=10*mm,topMargin=9*mm,bottomMargin=10*mm)
    story=_payroll_header_story("RELATORIO CONSOLIDADO DA FOLHA",f"Competência {comp.month:02d}/{comp.year} | Conferência e autorização interna",page_width)
    story.extend([_payroll_status_card(comp,calculations,page_width),Spacer(1,9)])
    data=[["COLABORADOR","CARGO / PROJETO","SALARIO-BASE","PROVENTOS","INSS","IRRF","DESCONTOS","LIQUIDO"]]
    tg=td=tn=Decimal("0")
    for i,calc in enumerate(calculations):
        emp=calc.employee; tg+=Decimal(calc.gross_amount or 0); td+=Decimal(calc.deductions_amount or 0); tn+=Decimal(calc.net_amount or 0)
        data.append([Paragraph(f"<b>{emp.full_name}</b>",styles["PayrollSmall"]),Paragraph(f"{emp.job_title}<br/><font color='#70736F'>{emp.project}</font>",styles["PayrollSmall"]),_fmt_brl(calc.base_salary),_fmt_brl(calc.gross_amount),_fmt_brl(calc.inss_amount),_fmt_brl(calc.irrf_amount),_fmt_brl(calc.deductions_amount),_fmt_brl(calc.net_amount)])
    data.append(["TOTAL DA FOLHA","","",_fmt_brl(tg),"","",_fmt_brl(td),_fmt_brl(tn)])
    table=Table(data,colWidths=[54*mm,49*mm,27*mm,28*mm,23*mm,23*mm,28*mm,31*mm],repeatRows=1)
    ts=[("BACKGROUND",(0,0),(-1,0),charcoal),("TEXTCOLOR",(0,0),(-1,0),colors.white),("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,-1),7.2),("LINEBELOW",(0,0),(-1,0),1,gold),("VALIGN",(0,0),(-1,-1),"MIDDLE"),("ALIGN",(2,1),(-1,-1),"RIGHT"),("TOPPADDING",(0,0),(-1,-1),6),("BOTTOMPADDING",(0,0),(-1,-1),6),("LEFTPADDING",(0,0),(-1,-1),6),("RIGHTPADDING",(0,0),(-1,-1),6)]
    for row in range(1,len(data)-1):
        ts.append(("BACKGROUND",(0,row),(-1,row),colors.white if row%2 else colors.HexColor("#FAFAF8")))
        ts.append(("LINEBELOW",(0,row),(-1,row),.35,line))
    ts.extend([("BACKGROUND",(0,-1),(-1,-1),gold_soft),("FONTNAME",(0,-1),(-1,-1),"Helvetica-Bold"),("LINEABOVE",(0,-1),(-1,-1),1,gold),("TEXTCOLOR",(7,-1),(7,-1),green)])
    table.setStyle(TableStyle(ts)); story.extend([table,Spacer(1,8)])
    closure=comp.closure
    footer=[]
    if closure:
        footer.append(f"Fechada em {closure.closed_at.strftime('%d/%m/%Y %H:%M')} por {closure.closer.email if closure.closer else 'RH'}")
        if closure.authorized_at: footer.append(f"Autorizada em {closure.authorized_at.strftime('%d/%m/%Y %H:%M')} - código {closure.authorization_code}")
    story.append(Paragraph(" | ".join(footer) if footer else "Relatório gerado pelo Portal RH.",styles["PayrollSub"]))
    story.append(Spacer(1,3)); story.append(Paragraph("Documento de conferência interna. Os valores devem permanecer conciliados com eSocial, DCTFWeb, FGTS Digital e demais obrigações oficiais aplicáveis.",styles["PayrollSub"]))
    doc.build(story); buff.seek(0); return buff


def _payroll_employee_payslip_pdf(calc, closure):
    comp=calc.competence; emp=calc.employee
    styles,gold,gold_soft,charcoal,muted,line,green=_payroll_pdf_styles()
    buff=BytesIO(); width=183*mm
    doc=SimpleDocTemplate(buff,pagesize=A4,rightMargin=14*mm,leftMargin=14*mm,topMargin=10*mm,bottomMargin=10*mm)
    story=_payroll_header_story("DEMONSTRATIVO DE PAGAMENTO",f"Competência {comp.month:02d}/{comp.year}",width)
    # employee identification card
    info=Table([
        [Paragraph("COLABORADOR",styles["PayrollLabel"]),Paragraph("CPF",styles["PayrollLabel"]),Paragraph("MATRICULA",styles["PayrollLabel"]),Paragraph("ADMISSAO",styles["PayrollLabel"])],
        [Paragraph(emp.full_name,styles["PayrollValue"]),Paragraph(emp.cpf or "-",styles["PayrollSmall"]),Paragraph(emp.registration or "-",styles["PayrollSmall"]),Paragraph(emp.admission_date.strftime('%d/%m/%Y') if emp.admission_date else "-",styles["PayrollSmall"])],
        [Paragraph("CARGO",styles["PayrollLabel"]),Paragraph("PROJETO / SERVICO",styles["PayrollLabel"]),"",Paragraph("COMPETENCIA",styles["PayrollLabel"])],
        [Paragraph(emp.job_title or "-",styles["PayrollSmall"]),Paragraph(emp.project or "-",styles["PayrollSmall"]),"",Paragraph(f"{comp.month:02d}/{comp.year}",styles["PayrollValue"])],
    ],colWidths=[76*mm,42*mm,30*mm,35*mm])
    info.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),colors.HexColor("#FAFAF8")),("SPAN",(1,2),(2,2)),("SPAN",(1,3),(2,3)),("BOX",(0,0),(-1,-1),.6,line),("INNERGRID",(0,0),(-1,-1),.35,line),("VALIGN",(0,0),(-1,-1),"MIDDLE"),("TOPPADDING",(0,0),(-1,-1),5),("BOTTOMPADDING",(0,0),(-1,-1),5),("LEFTPADDING",(0,0),(-1,-1),7),("RIGHTPADDING",(0,0),(-1,-1),7)]))
    story.extend([info,Spacer(1,9)])
    items=PayrollCalculationItem.query.filter_by(calculation_id=calc.id).order_by(PayrollCalculationItem.sort_order,PayrollCalculationItem.id).all()
    data=[["COD.","DESCRICAO","REFERENCIA","PROVENTOS","DESCONTOS"]]
    for item in items:
        data.append([item.rubric_code,Paragraph(item.description,styles["PayrollSmall"]),item.reference or "-",_fmt_brl(item.amount) if item.nature=="earning" else "",_fmt_brl(item.amount) if item.nature=="deduction" else ""])
    data.append(["","TOTAIS","",_fmt_brl(calc.gross_amount),_fmt_brl(calc.deductions_amount)])
    tab=Table(data,colWidths=[17*mm,69*mm,39*mm,29*mm,29*mm],repeatRows=1)
    ts=[("BACKGROUND",(0,0),(-1,0),charcoal),("TEXTCOLOR",(0,0),(-1,0),colors.white),("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("LINEBELOW",(0,0),(-1,0),1,gold),("FONTSIZE",(0,0),(-1,-1),7.6),("VALIGN",(0,0),(-1,-1),"MIDDLE"),("ALIGN",(3,1),(-1,-1),"RIGHT"),("TOPPADDING",(0,0),(-1,-1),6),("BOTTOMPADDING",(0,0),(-1,-1),6),("LEFTPADDING",(0,0),(-1,-1),6),("RIGHTPADDING",(0,0),(-1,-1),6)]
    for row in range(1,len(data)-1):
        ts.append(("BACKGROUND",(0,row),(-1,row),colors.white if row%2 else colors.HexColor("#FAFAF8"))); ts.append(("LINEBELOW",(0,row),(-1,row),.35,line))
    ts.extend([("BACKGROUND",(0,-1),(-1,-1),gold_soft),("FONTNAME",(0,-1),(-1,-1),"Helvetica-Bold"),("LINEABOVE",(0,-1),(-1,-1),1,gold)])
    tab.setStyle(TableStyle(ts)); story.extend([tab,Spacer(1,9)])
    # prominent net pay + bases
    net=Table([[Paragraph("LIQUIDO A RECEBER",styles["PayrollLabel"]),Paragraph(_fmt_brl(calc.net_amount),styles["PayrollNet"])],[Paragraph("Total de proventos",styles["PayrollSmall"]),Paragraph(_fmt_brl(calc.gross_amount),styles["PayrollSmallRight"])],[Paragraph("Total de descontos",styles["PayrollSmall"]),Paragraph(_fmt_brl(calc.deductions_amount),styles["PayrollSmallRight"])]],colWidths=[95*mm,88*mm])
    net.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,0),gold_soft),("BOX",(0,0),(-1,-1),.7,line),("LINEBELOW",(0,0),(-1,-2),.35,line),("VALIGN",(0,0),(-1,-1),"MIDDLE"),("ALIGN",(1,0),(-1,-1),"RIGHT"),("TOPPADDING",(0,0),(-1,-1),7),("BOTTOMPADDING",(0,0),(-1,-1),7),("LEFTPADDING",(0,0),(-1,-1),8),("RIGHTPADDING",(0,0),(-1,-1),8)]))
    story.extend([net,Spacer(1,8)])
    bases=Table([[Paragraph("SALARIO-BASE",styles["PayrollLabel"]),Paragraph("BASE INSS",styles["PayrollLabel"]),Paragraph("INSS",styles["PayrollLabel"]),Paragraph("BASE IRRF",styles["PayrollLabel"]),Paragraph("IRRF",styles["PayrollLabel"])],[_fmt_brl(calc.base_salary),_fmt_brl(calc.inss_base),_fmt_brl(calc.inss_amount),_fmt_brl(calc.irrf_base),_fmt_brl(calc.irrf_amount)]],colWidths=[36.6*mm]*5)
    bases.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,0),colors.HexColor("#F4F5F2")),("BOX",(0,0),(-1,-1),.5,line),("INNERGRID",(0,0),(-1,-1),.35,line),("ALIGN",(0,0),(-1,-1),"CENTER"),("FONTSIZE",(0,1),(-1,1),7.5),("TOPPADDING",(0,0),(-1,-1),5),("BOTTOMPADDING",(0,0),(-1,-1),5)]))
    story.extend([bases,Spacer(1,11)])
    story.append(Paragraph(f"Folha autorizada pelo RH em {closure.authorized_at.strftime('%d/%m/%Y %H:%M')}. Código de autorização: <b>{closure.authorization_code}</b>.",styles["PayrollSmall"]))
    story.append(Spacer(1,3)); story.append(Paragraph("Documento individual e confidencial disponibilizado no Portal RH da Associação Alecrim Dourado.",styles["PayrollSub"]))
    doc.build(story); buff.seek(0); return buff

def _save_generated_payslip(calc, closure):
    emp=calc.employee; comp=calc.competence
    pdf=_payroll_employee_payslip_pdf(calc,closure)
    stored=f"holerite_portal_{comp.year}_{comp.month:02d}_{emp.id}_{uuid.uuid4().hex}.pdf"
    path=os.path.join(current_app.config['UPLOAD_FOLDER'],stored)
    os.makedirs(current_app.config['UPLOAD_FOLDER'],exist_ok=True)
    with open(path,'wb') as handle: handle.write(pdf.getvalue())
    _chmod_private(path)
    original=f"Holerite_{comp.month:02d}-{comp.year}_{secure_filename(emp.full_name) or emp.id}.pdf"
    existing=Payslip.query.filter_by(employee_id=emp.id,year=comp.year,month=comp.month).first()
    old=None
    if existing:
        old=existing.stored_name; existing.original_name=original; existing.stored_name=stored; existing.matched_by='portal_payroll'; existing.uploaded_at=now_local(); existing.uploaded_by=current_user.id; existing.employee_viewed_at=None; item=existing
    else:
        item=Payslip(employee_id=emp.id,year=comp.year,month=comp.month,original_name=original,stored_name=stored,matched_by='portal_payroll',uploaded_by=current_user.id)
        db.session.add(item)
    db.session.flush()
    if old and old!=stored:
        try:
            old_path=os.path.join(current_app.config['UPLOAD_FOLDER'],old)
            if os.path.isfile(old_path): os.remove(old_path)
        except OSError: pass
    return item


@bp.route("/payroll/competence", methods=["POST"])
@login_required
@roles_required(ROLE_ADMIN)
def payroll_competence_open():
    try:
        year,month=_competence_from_value(request.form.get("competence") or "")
    except ValueError as exc:
        flash(str(exc),"danger"); return redirect(url_for("rh.payroll_center"))
    comp=PayrollCompetence.query.filter_by(year=year,month=month).first()
    if not comp:
        comp=PayrollCompetence(year=year,month=month,created_by=current_user.id)
        db.session.add(comp); db.session.flush()
        log_action("abriu competência de pré-folha","payroll_competence",comp.id,f"{month:02d}/{year}")
        db.session.commit()
    return redirect(url_for("rh.payroll_competence_detail",competence_id=comp.id))


@bp.route("/payroll/competence/<int:competence_id>")
@login_required
@roles_required(ROLE_ADMIN)
def payroll_competence_detail(competence_id):
    _ensure_payroll_2026_parameters(); _ensure_default_payroll_rubrics()
    comp=db.get_or_404(PayrollCompetence,competence_id)
    employees=Employee.query.filter_by(is_active=True).order_by(Employee.full_name.asc()).all()
    calculations={c.employee_id:c for c in PayrollEmployeeCalculation.query.filter_by(competence_id=comp.id).all()}
    warnings={emp.id:_month_data_warnings(emp,comp.year,comp.month) for emp in employees}
    rubrics=PayrollRubric.query.filter_by(active=True).order_by(PayrollRubric.nature.asc(),PayrollRubric.code.asc()).all()
    events=PayrollManualEvent.query.filter_by(competence_id=comp.id).order_by(PayrollManualEvent.employee_id,PayrollManualEvent.id).all()
    events_by_employee={}
    for ev in events: events_by_employee.setdefault(ev.employee_id,[]).append(ev)
    totals={
        "gross":sum(Decimal(c.gross_amount or 0) for c in calculations.values()),
        "deductions":sum(Decimal(c.deductions_amount or 0) for c in calculations.values()),
        "net":sum(Decimal(c.net_amount or 0) for c in calculations.values()),
    }
    return render_template("payroll_competence.html",comp=comp,employees=employees,calculations=calculations,
        warnings=warnings,rubrics=rubrics,events_by_employee=events_by_employee,totals=totals,closure=comp.closure)


@bp.route("/payroll/competence/<int:competence_id>/events",methods=["POST"])
@login_required
@roles_required(ROLE_ADMIN)
def payroll_event_add(competence_id):
    comp=db.get_or_404(PayrollCompetence,competence_id)
    if _payroll_closure_is_locked(comp):
        flash("A competência está fechada. Reabra antes de alterar eventos.","danger")
        return redirect(url_for("rh.payroll_competence_detail",competence_id=comp.id))
    emp=db.get_or_404(Employee,request.form.get("employee_id",type=int))
    rub=db.get_or_404(PayrollRubric,request.form.get("rubric_id",type=int))
    qty_raw=(request.form.get("reference_quantity") or "").strip().replace(",",".")
    amount_raw=(request.form.get("amount") or "").strip()
    try:
        qty=Decimal(qty_raw) if qty_raw else None
        amount=parse_money(amount_raw) if amount_raw else None
    except (InvalidOperation,ValueError):
        flash("Quantidade ou valor inválido.","danger"); return redirect(url_for("rh.payroll_competence_detail",competence_id=comp.id)+f"#emp-{emp.id}")
    if amount is None and qty is None:
        flash("Informe uma quantidade/referência ou um valor.","danger"); return redirect(url_for("rh.payroll_competence_detail",competence_id=comp.id)+f"#emp-{emp.id}")
    if amount is None and not (rub.code.startswith("HE") or rub.code=="FALTA"):
        flash("Para esta rubrica, informe o valor monetário do evento. O Portal não presume base de cálculo para adicionais que dependem de laudo, CCT ou enquadramento.","danger")
        return redirect(url_for("rh.payroll_competence_detail",competence_id=comp.id)+f"#emp-{emp.id}")
    ev=PayrollManualEvent(competence_id=comp.id,employee_id=emp.id,rubric_id=rub.id,reference_quantity=qty,
        reference_label=(request.form.get("reference_label") or "").strip() or None,amount=amount,
        notes=(request.form.get("notes") or "").strip() or None,created_by=current_user.id)
    db.session.add(ev); db.session.flush()
    log_action("incluiu evento na pré-folha","payroll_event",ev.id,f"{emp.full_name}; {rub.code}; {comp.month:02d}/{comp.year}")
    comp.status="open"; db.session.commit()
    flash("Evento incluído. Recalcule a competência para atualizar os valores.","success")
    return redirect(url_for("rh.payroll_competence_detail",competence_id=comp.id)+f"#emp-{emp.id}")


@bp.route("/payroll/events/<int:event_id>/delete",methods=["POST"])
@login_required
@roles_required(ROLE_ADMIN)
def payroll_event_delete(event_id):
    ev=db.get_or_404(PayrollManualEvent,event_id)
    if _payroll_closure_is_locked(ev.competence):
        flash("A competência está fechada. Reabra antes de remover eventos.","danger")
        return redirect(url_for("rh.payroll_competence_detail",competence_id=ev.competence_id)); cid=ev.competence_id; eid=ev.employee_id
    log_action("removeu evento da pré-folha","payroll_event",ev.id,f"{ev.employee.full_name}; {ev.rubric.code}")
    ev.competence.status="open"; db.session.delete(ev); db.session.commit()
    flash("Evento removido. Recalcule a competência.","success")
    return redirect(url_for("rh.payroll_competence_detail",competence_id=cid)+f"#emp-{eid}")


@bp.route("/payroll/competence/<int:competence_id>/calculate",methods=["POST"])
@login_required
@roles_required(ROLE_ADMIN)
def payroll_competence_calculate(competence_id):
    _ensure_payroll_2026_parameters(); _ensure_default_payroll_rubrics()
    comp=db.get_or_404(PayrollCompetence,competence_id)
    if _payroll_closure_is_locked(comp):
        flash("A competência está fechada. Reabra antes de recalcular.","danger")
        return redirect(url_for("rh.payroll_competence_detail",competence_id=comp.id))
    employees=Employee.query.filter_by(is_active=True).order_by(Employee.full_name.asc()).all()
    calculated=0; skipped=[]
    try:
        for emp in employees:
            calc,warns=_calculate_payroll_employee(comp,emp)
            if calc: calculated+=1
            else: skipped.append(emp.full_name)
        comp.status="calculated"; comp.calculated_at=now_local(); comp.calculated_by=current_user.id
        log_action("calculou pré-folha","payroll_competence",comp.id,f"{comp.month:02d}/{comp.year}; {calculated} colaboradores")
        db.session.commit()
    except ValueError as exc:
        db.session.rollback(); flash(f"Não foi possível calcular: {exc}","danger")
        return redirect(url_for("rh.payroll_competence_detail",competence_id=comp.id))
    msg=f"Pré-folha calculada para {calculated} colaborador(es)."
    if skipped: msg+=f" {len(skipped)} sem salário configurado foram ignorados."
    flash(msg,"success")
    return redirect(url_for("rh.payroll_competence_detail",competence_id=comp.id))


@bp.route("/payroll/competence/<int:competence_id>/close",methods=["POST"])
@login_required
@roles_required(ROLE_ADMIN)
def payroll_competence_close(competence_id):
    comp=db.get_or_404(PayrollCompetence,competence_id)
    if comp.status!="calculated":
        flash("Calcule a competência antes de fechá-la.","danger")
        return redirect(url_for("rh.payroll_competence_detail",competence_id=comp.id))
    calculations=PayrollEmployeeCalculation.query.filter_by(competence_id=comp.id).all()
    if not calculations:
        flash("Não há valores calculados nesta competência.","danger")
        return redirect(url_for("rh.payroll_competence_detail",competence_id=comp.id))
    # Todos os colaboradores ativos com salário configurado devem possuir cálculo antes do fechamento.
    configured=(Employee.query.join(PayrollEmployeeConfig).filter(Employee.is_active==True,PayrollEmployeeConfig.monthly_salary>0).all())
    missing=[emp.full_name for emp in configured if not any(c.employee_id==emp.id for c in calculations)]
    if missing:
        flash(f"Não foi possível fechar: {len(missing)} colaborador(es) com remuneração configurada estão sem cálculo.","danger")
        return redirect(url_for("rh.payroll_competence_detail",competence_id=comp.id))
    closure=comp.closure
    if closure and closure.status=="authorized":
        flash("Esta folha já foi autorizada e não pode ser fechada novamente.","danger")
        return redirect(url_for("rh.payroll_competence_detail",competence_id=comp.id))
    if not closure:
        closure=PayrollClosure(competence_id=comp.id,closed_by=current_user.id)
        db.session.add(closure)
    closure.status="closed"; closure.closed_at=now_local(); closure.closed_by=current_user.id
    closure.close_note=(request.form.get("note") or "").strip() or None
    closure.reopened_at=None; closure.reopened_by=None; closure.reopen_reason=None
    log_action("fechou pré-folha de pagamento","payroll_closure",comp.id,f"{comp.month:02d}/{comp.year}; {len(calculations)} colaboradores")
    log_security_event("payroll_closed",severity="info",user=current_user,details=f"Pré-folha {comp.month:02d}/{comp.year} fechada para autorização.")
    db.session.commit()
    flash("Competência fechada. O PDF consolidado está disponível para conferência e a folha aguarda autorização.","success")
    return redirect(url_for("rh.payroll_competence_detail",competence_id=comp.id))


@bp.route("/payroll/competence/<int:competence_id>/sheet.pdf")
@login_required
@roles_required(ROLE_ADMIN)
def payroll_competence_sheet_pdf(competence_id):
    comp=db.get_or_404(PayrollCompetence,competence_id)
    if not comp.closure or comp.closure.status not in {"closed","authorized"}:
        flash("Feche a competência antes de gerar a folha consolidada.","danger")
        return redirect(url_for("rh.payroll_competence_detail",competence_id=comp.id))
    try: buff=_payroll_consolidated_pdf(comp)
    except ValueError as exc:
        flash(str(exc),"danger"); return redirect(url_for("rh.payroll_competence_detail",competence_id=comp.id))
    log_action("gerou PDF consolidado da folha","payroll_competence",comp.id,f"{comp.month:02d}/{comp.year}")
    db.session.commit()
    return send_file(buff,mimetype="application/pdf",as_attachment=True,download_name=f"Folha_Pagamento_{comp.month:02d}-{comp.year}.pdf")


@bp.route("/payroll/competence/<int:competence_id>/authorize",methods=["POST"])
@login_required
@roles_required(ROLE_ADMIN)
def payroll_competence_authorize(competence_id):
    comp=db.get_or_404(PayrollCompetence,competence_id); closure=comp.closure
    if not closure or closure.status!="closed":
        flash("A competência precisa estar fechada antes da autorização.","danger")
        return redirect(url_for("rh.payroll_competence_detail",competence_id=comp.id))
    if request.form.get("confirm_authorization")!="1":
        flash("Confirme a autorização da folha.","danger"); return redirect(url_for("rh.payroll_competence_detail",competence_id=comp.id))
    password=request.form.get("password") or ""
    if not current_user.check_password(password):
        log_security_event("payroll_authorization_failed",severity="warning",user=current_user,details=f"Senha inválida ao tentar autorizar a folha {comp.month:02d}/{comp.year}.")
        db.session.commit(); flash("Senha do administrador inválida.","danger")
        return redirect(url_for("rh.payroll_competence_detail",competence_id=comp.id))
    calculations=PayrollEmployeeCalculation.query.filter_by(competence_id=comp.id).all()
    stamp=now_local(); raw=f"payroll|{comp.id}|{current_user.id}|{stamp.isoformat()}|{uuid.uuid4().hex}"
    closure.status="authorized"; closure.authorized_at=stamp; closure.authorized_by=current_user.id
    closure.authorization_ip=client_ip(); closure.authorization_code=hashlib.sha256(raw.encode('utf-8')).hexdigest()[:32].upper()
    closure.authorization_note=(request.form.get("note") or "").strip() or None
    generated=0
    for calc in calculations:
        if Decimal(calc.net_amount or 0)!=0 or Decimal(calc.gross_amount or 0)!=0:
            _save_generated_payslip(calc,closure); generated+=1
    log_action("autorizou folha de pagamento","payroll_closure",comp.id,f"{comp.month:02d}/{comp.year}; código {closure.authorization_code}; {generated} holerites liberados")
    log_security_event("payroll_authorized",severity="info",user=current_user,details=f"Folha {comp.month:02d}/{comp.year} autorizada; {generated} holerites disponibilizados.")
    db.session.commit()
    flash(f"Folha autorizada. {generated} holerite(s) personalizados foram gerados e disponibilizados aos colaboradores.","success")
    return redirect(url_for("rh.payroll_competence_detail",competence_id=comp.id))


@bp.route("/payroll/competence/<int:competence_id>/reopen",methods=["POST"])
@login_required
@roles_required(ROLE_ADMIN)
def payroll_competence_reopen(competence_id):
    comp=db.get_or_404(PayrollCompetence,competence_id); closure=comp.closure
    if not closure or closure.status!="closed":
        flash("Somente uma competência fechada e ainda não autorizada pode ser reaberta.","danger")
        return redirect(url_for("rh.payroll_competence_detail",competence_id=comp.id))
    reason=(request.form.get("reason") or "").strip()
    if len(reason)<5:
        flash("Informe o motivo da reabertura.","danger"); return redirect(url_for("rh.payroll_competence_detail",competence_id=comp.id))
    closure.status="reopened"; closure.reopened_at=now_local(); closure.reopened_by=current_user.id; closure.reopen_reason=reason
    log_action("reabriu pré-folha","payroll_closure",comp.id,f"{comp.month:02d}/{comp.year}; {reason}")
    db.session.commit(); flash("Competência reaberta. Recalcule e feche novamente após os ajustes.","success")
    return redirect(url_for("rh.payroll_competence_detail",competence_id=comp.id))


@bp.route("/payroll/calculation/<int:calculation_id>")
@login_required
@roles_required(ROLE_ADMIN)
def payroll_calculation_detail(calculation_id):
    calc=db.get_or_404(PayrollEmployeeCalculation,calculation_id)
    items=PayrollCalculationItem.query.filter_by(calculation_id=calc.id).order_by(PayrollCalculationItem.sort_order,PayrollCalculationItem.id).all()
    try: detail=json.loads(calc.calculation_notes or "{}")
    except Exception: detail={}
    return render_template("payroll_calculation_detail.html",calc=calc,items=items,detail=detail)

@bp.route("/employees")
@login_required
def employees():
    return render_template("employees.html", employees=visible_employees())

@bp.route("/employees/new", methods=["GET", "POST"])
@login_required
@roles_required(ROLE_ADMIN)
def employee_new():
    managers = Employee.query.join(User).filter(User.role.in_([ROLE_ADMIN, ROLE_MANAGER])).order_by(Employee.full_name).all()
    if request.method == "POST":
        email = request.form["email"].strip().lower()
        cpf = request.form["cpf"].strip()
        if User.query.filter_by(email=email).first() or Employee.query.filter_by(cpf=cpf).first():
            flash("E-mail ou CPF já cadastrado.", "danger")
            return render_template("employee_form.html", managers=managers)
        initial_password = (request.form.get("password") or "").strip()
        point_pin = (request.form.get("point_pin") or "").strip()
        if len(initial_password) < 10:
            flash("Defina uma senha provisória de pelo menos 10 caracteres para o primeiro acesso.", "danger")
            return render_template("employee_form.html", managers=managers)
        if not (len(point_pin) == 6 and point_pin.isdigit()):
            flash("A senha de ponto deve conter exatamente 6 dígitos numéricos.", "danger")
            return render_template("employee_form.html", managers=managers)
        user = User(
            email=email,
            role=request.form.get("role", ROLE_EMPLOYEE),
            must_change_password=True
        )
        user.set_password(initial_password)
        db.session.add(user); db.session.flush()
        emp = Employee(
            user_id=user.id, full_name=request.form["full_name"], cpf=cpf,
            rg=request.form.get("rg"), birth_date=parse_date(request.form.get("birth_date")),
            phone=request.form.get("phone"), address=request.form.get("address"),
            registration=request.form.get("registration") or None, job_title=request.form["job_title"],
            department=request.form["department"], project=request.form.get("project") or "Administração",
            admission_date=parse_date(request.form["admission_date"]), contract_type=request.form.get("contract_type") or "CLT",
            weekly_hours=float(request.form.get("weekly_hours") or 44),
            standard_start=parse_time(request.form.get("standard_start")) or datetime.strptime("08:00", "%H:%M").time(),
            standard_end=parse_time(request.form.get("standard_end")) or datetime.strptime("17:00", "%H:%M").time(),
            manager_id=int(request.form["manager_id"]) if request.form.get("manager_id") else None,
        )
        emp.set_point_pin(point_pin)
        db.session.add(emp); db.session.flush(); log_action("criou colaborador", "employee", emp.id, emp.full_name); db.session.commit()
        flash("Colaborador cadastrado. No primeiro acesso, ele deverá substituir a senha provisória pela senha pessoal.", "success")
        return redirect(url_for("rh.employee_detail", employee_id=emp.id))
    return render_template("employee_form.html", managers=managers)


@bp.route("/employees/<int:employee_id>/edit", methods=["GET", "POST"])
@login_required
@roles_required(ROLE_ADMIN)
def employee_edit(employee_id):
    emp = db.get_or_404(Employee, employee_id)
    managers = (Employee.query.join(User)
                .filter(User.role.in_([ROLE_ADMIN, ROLE_MANAGER]), Employee.id != emp.id)
                .order_by(Employee.full_name).all())

    if request.method == "POST":
        email = (request.form.get("email") or "").strip().lower()
        cpf = (request.form.get("cpf") or "").strip()
        registration = (request.form.get("registration") or "").strip() or None

        duplicate_email = User.query.filter(User.email == email, User.id != emp.user_id).first()
        duplicate_cpf = Employee.query.filter(Employee.cpf == cpf, Employee.id != emp.id).first()
        duplicate_registration = None
        if registration:
            duplicate_registration = Employee.query.filter(
                Employee.registration == registration,
                Employee.id != emp.id
            ).first()

        if duplicate_email:
            flash("Já existe outro usuário cadastrado com esse e-mail.", "danger")
            return render_template("employee_edit.html", emp=emp, managers=managers)
        if duplicate_cpf:
            flash("Já existe outro colaborador cadastrado com esse CPF.", "danger")
            return render_template("employee_edit.html", emp=emp, managers=managers)
        if duplicate_registration:
            flash("Já existe outro colaborador cadastrado com essa matrícula.", "danger")
            return render_template("employee_edit.html", emp=emp, managers=managers)

        try:
            weekly_hours = float(request.form.get("weekly_hours") or emp.weekly_hours or 44)
            birth_date = parse_date(request.form.get("birth_date"))
            admission_date = parse_date(request.form.get("admission_date"))
            standard_start = parse_time(request.form.get("standard_start"))
            standard_end = parse_time(request.form.get("standard_end"))
        except (ValueError, TypeError) as exc:
            flash(f"Revise os dados de data, horário ou carga horária: {exc}", "danger")
            return render_template("employee_edit.html", emp=emp, managers=managers)

        old_summary = (
            f"nome={emp.full_name}; email={emp.user.email}; cpf={emp.cpf}; "
            f"cargo={emp.job_title}; setor={emp.department}; projeto={emp.project}"
        )

        emp.full_name = (request.form.get("full_name") or "").strip()
        emp.cpf = cpf
        emp.rg = (request.form.get("rg") or "").strip() or None
        emp.birth_date = birth_date
        emp.phone = (request.form.get("phone") or "").strip() or None
        emp.address = (request.form.get("address") or "").strip() or None
        emp.registration = registration
        emp.job_title = (request.form.get("job_title") or "").strip()
        emp.department = (request.form.get("department") or "").strip()
        emp.project = (request.form.get("project") or "").strip() or "Administração"
        emp.admission_date = admission_date
        emp.contract_type = (request.form.get("contract_type") or "").strip() or "CLT"
        emp.weekly_hours = weekly_hours
        emp.standard_start = standard_start
        emp.standard_end = standard_end
        emp.manager_id = int(request.form["manager_id"]) if request.form.get("manager_id") else None
        emp.is_active = request.form.get("is_active") == "1"

        emp.user.email = email
        role = request.form.get("role") or ROLE_EMPLOYEE
        if role not in {ROLE_EMPLOYEE, ROLE_MANAGER, ROLE_ADMIN}:
            role = ROLE_EMPLOYEE
        emp.user.role = role
        emp.user.active = emp.is_active

        photo = request.files.get("profile_photo")
        if photo and photo.filename:
            try:
                new_photo = save_profile_photo(photo)
            except ValueError as exc:
                flash(str(exc), "danger")
                return render_template("employee_edit.html", emp=emp, managers=managers)
            old_photo = emp.profile_photo
            emp.profile_photo = new_photo
            if old_photo:
                old_path = os.path.join(current_app.config["UPLOAD_FOLDER"], old_photo)
                try:
                    if os.path.isfile(old_path):
                        os.remove(old_path)
                except OSError:
                    pass

        new_summary = (
            f"nome={emp.full_name}; email={emp.user.email}; cpf={emp.cpf}; "
            f"cargo={emp.job_title}; setor={emp.department}; projeto={emp.project}"
        )
        log_action("alterou cadastro do colaborador", "employee", emp.id,
                   f"Antes: {old_summary}. Depois: {new_summary}.")
        db.session.commit()
        flash("Cadastro do colaborador atualizado com sucesso.", "success")
        return redirect(url_for("rh.employee_detail", employee_id=emp.id))

    return render_template("employee_edit.html", emp=emp, managers=managers)


@bp.route("/employees/<int:employee_id>/photo")
@login_required
def employee_photo(employee_id):
    emp = db.get_or_404(Employee, employee_id)
    if not can_manage_employee(emp):
        abort(403)
    if not emp.profile_photo:
        abort(404)

    photo_path = os.path.join(current_app.config["UPLOAD_FOLDER"], emp.profile_photo)
    if not os.path.isfile(photo_path):
        # O banco pode manter o nome de uma foto antiga que não existe mais no disco.
        # Nesse caso, o front-end exibe as iniciais até que uma nova foto seja enviada.
        abort(404)

    response = send_from_directory(
        current_app.config["UPLOAD_FOLDER"],
        emp.profile_photo,
        as_attachment=False,
        conditional=False,
    )
    response.headers["Cache-Control"] = "no-store, private, max-age=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["X-Content-Type-Options"] = "nosniff"
    return response


@bp.route("/employees/<int:employee_id>")
@login_required
def employee_detail(employee_id):
    emp = db.get_or_404(Employee, employee_id)
    if not can_manage_employee(emp): abort(403)
    certs = MedicalCertificate.query.filter_by(employee_id=emp.id).order_by(MedicalCertificate.uploaded_at.desc()).all()
    reqs = Request.query.filter_by(employee_id=emp.id).order_by(Request.requested_at.desc()).limit(50).all()
    docs = Document.query.filter_by(employee_id=emp.id).order_by(Document.uploaded_at.desc()).all()
    payslips = Payslip.query.filter_by(employee_id=emp.id).order_by(Payslip.year.desc(), Payslip.month.desc()).all()
    clocks = TimeClock.query.filter_by(employee_id=emp.id).order_by(TimeClock.punched_at.desc()).limit(120).all()
    bank_adjustments = BankHourAdjustment.query.filter_by(employee_id=emp.id).order_by(BankHourAdjustment.created_at.desc()).limit(100).all()
    weekend_duties = WeekendDuty.query.filter_by(employee_id=emp.id).order_by(WeekendDuty.duty_date.desc()).limit(100).all()
    work_schedule = emp.work_schedule
    bank_summary = _bank_summary(emp)
    vacations = Vacation.query.filter_by(employee_id=emp.id).order_by(Vacation.start_date.desc()).all()
    vacation_schedules = VacationSchedule.query.filter_by(employee_id=emp.id).order_by(VacationSchedule.planned_start.asc()).all()
    vacation_summary = _vacation_entitlement(emp)
    return render_template("employee_detail.html", emp=emp, certs=certs, reqs=reqs, docs=docs, payslips=payslips, clocks=clocks, bank_adjustments=bank_adjustments, weekend_duties=weekend_duties, work_schedule=work_schedule, bank_summary=bank_summary, vacations=vacations, vacation_schedules=vacation_schedules, vacation_summary=vacation_summary, current_month=today_local().strftime("%Y-%m"), today_iso=today_local().isoformat())


@bp.route("/employees/<int:employee_id>/reset-password", methods=["POST"])
@login_required
@roles_required(ROLE_ADMIN)
def employee_reset_password(employee_id):
    emp = db.get_or_404(Employee, employee_id)
    temporary_password = (request.form.get("temporary_password") or "").strip()
    if len(temporary_password) < 10:
        flash("A senha provisória deve possuir pelo menos 10 caracteres.", "danger")
        return redirect(url_for("rh.employee_detail", employee_id=employee_id) + "#acesso")
    emp.user.set_password(temporary_password)
    emp.user.must_change_password = True
    emp.user.password_changed_at = None
    log_action("redefiniu senha provisória", "user", emp.user.id,
               f"Nova senha provisória criada para {emp.full_name}; troca obrigatória no próximo acesso.")
    log_security_event(
        "password_reset_by_rh",
        severity="warning",
        user=current_user,
        employee=emp,
        details=f"RH redefiniu a senha provisória de {emp.full_name}.",
    )
    db.session.commit()
    flash("Senha provisória redefinida. O colaborador deverá criar uma nova senha no próximo acesso.", "success")
    return redirect(url_for("rh.employee_detail", employee_id=employee_id) + "#acesso")



@bp.route("/employees/<int:employee_id>/reset-point-pin", methods=["POST"])
@login_required
@roles_required(ROLE_ADMIN)
def employee_reset_point_pin(employee_id):
    emp = db.get_or_404(Employee, employee_id)
    pin = (request.form.get("point_pin") or "").strip()
    if not (len(pin) == 6 and pin.isdigit()):
        flash("A senha de ponto deve conter exatamente 6 dígitos numéricos.", "danger")
        return redirect(url_for("rh.employee_detail", employee_id=employee_id) + "#acesso")
    emp.set_point_pin(pin)
    log_action("redefiniu senha de ponto", "employee", emp.id,
               f"Senha de ponto de 6 dígitos redefinida para {emp.full_name}.")
    log_security_event(
        "point_pin_reset_by_rh",
        severity="warning",
        user=current_user,
        employee=emp,
        details=f"RH redefiniu a senha de ponto de {emp.full_name}.",
    )
    db.session.commit()
    flash("Senha de ponto redefinida com sucesso.", "success")
    return redirect(url_for("rh.employee_detail", employee_id=employee_id) + "#acesso")



@bp.route("/employees/<int:employee_id>/work-schedule", methods=["POST"])
@login_required
@roles_required(ROLE_ADMIN)
def employee_work_schedule(employee_id):
    emp = db.get_or_404(Employee, employee_id)
    interval_start = parse_time(request.form.get("interval_start"))
    interval_end = parse_time(request.form.get("interval_end"))

    if (interval_start and not interval_end) or (interval_end and not interval_start):
        flash("Informe o início e o fim do intervalo.", "danger")
        return redirect(url_for("rh.employee_detail", employee_id=employee_id) + "#jornada-individual")

    if interval_start and interval_end and interval_end <= interval_start:
        flash("O fim do intervalo deve ser posterior ao início.", "danger")
        return redirect(url_for("rh.employee_detail", employee_id=employee_id) + "#jornada-individual")

    item = emp.work_schedule
    if not item:
        item = EmployeeWorkSchedule(
            employee_id=emp.id,
            interval_start=interval_start,
            interval_end=interval_end,
            updated_by=current_user.id,
        )
        db.session.add(item)
        db.session.flush()
    else:
        item.interval_start = interval_start
        item.interval_end = interval_end
        item.updated_by = current_user.id
        item.updated_at = now_local()

    interval_text = (
        f"{interval_start.strftime('%H:%M')} às {interval_end.strftime('%H:%M')}"
        if interval_start and interval_end else "sem intervalo individual definido"
    )
    log_action(
        "alterou horário de intervalo do colaborador",
        "employee_work_schedule",
        item.id,
        f"{emp.full_name}; {interval_text}",
    )
    db.session.commit()
    flash("Horário de intervalo atualizado.", "success")
    return redirect(url_for("rh.employee_detail", employee_id=employee_id) + "#jornada-individual")


@bp.route("/employees/<int:employee_id>/weekend-duty", methods=["POST"])
@login_required
@roles_required(ROLE_ADMIN)
def employee_weekend_duty(employee_id):
    emp = db.get_or_404(Employee, employee_id)
    duty_date = parse_date(request.form.get("duty_date"))

    try:
        hours = max(int(request.form.get("hours") or 14), 0)
        minutes_part = max(min(int(request.form.get("minutes") or 0), 59), 0)
    except (TypeError, ValueError):
        flash("Informe uma duração válida para o plantão.", "danger")
        return redirect(url_for("rh.employee_detail", employee_id=employee_id) + "#plantoes")

    total = hours * 60 + minutes_part
    note = (request.form.get("note") or "").strip() or None

    if not duty_date:
        flash("Informe a data do plantão.", "danger")
        return redirect(url_for("rh.employee_detail", employee_id=employee_id) + "#plantoes")

    if duty_date.weekday() not in (5, 6):
        flash("A função Plantão executado aceita datas de sábado ou domingo.", "danger")
        return redirect(url_for("rh.employee_detail", employee_id=employee_id) + "#plantoes")

    if total <= 0 or total > 24 * 60:
        flash("Informe uma duração de plantão entre 00:01 e 24:00.", "danger")
        return redirect(url_for("rh.employee_detail", employee_id=employee_id) + "#plantoes")

    existing = WeekendDuty.query.filter_by(employee_id=emp.id, duty_date=duty_date).first()
    if existing:
        flash("Já existe um plantão registrado para este colaborador nessa data.", "danger")
        return redirect(url_for("rh.employee_detail", employee_id=employee_id) + "#plantoes")

    old_balance = int(emp.bank_minutes or 0)
    emp.bank_minutes = old_balance + total

    duty = WeekendDuty(
        employee_id=emp.id,
        duty_date=duty_date,
        minutes=total,
        note=note,
        created_by=current_user.id,
    )
    db.session.add(duty)
    db.session.flush()

    log_action(
        "registrou plantão executado",
        "weekend_duty",
        duty.id,
        (
            f"{emp.full_name}; {duty_date.strftime('%d/%m/%Y')}; "
            f"{total // 60:02d}:{total % 60:02d}; "
            f"banco {old_balance} min -> {emp.bank_minutes} min; "
            f"observação: {note or '-'}"
        ),
    )
    db.session.commit()

    flash(
        f"Plantão executado registrado. Crédito de {total // 60:02d}:{total % 60:02d} "
        "adicionado ao banco de horas.",
        "success",
    )
    return redirect(url_for("rh.employee_detail", employee_id=employee_id) + "#plantoes")


@bp.route("/weekend-duty/<int:duty_id>/delete", methods=["POST"])
@login_required
@roles_required(ROLE_ADMIN)
def weekend_duty_delete(duty_id):
    duty = db.get_or_404(WeekendDuty, duty_id)
    emp = duty.employee
    reason = (request.form.get("reason") or "").strip()

    if not reason:
        flash("Informe o motivo para remover o plantão.", "danger")
        return redirect(url_for("rh.employee_detail", employee_id=emp.id) + "#plantoes")

    old_balance = int(emp.bank_minutes or 0)
    emp.bank_minutes = old_balance - int(duty.minutes or 0)

    log_action(
        "removeu plantão executado",
        "weekend_duty",
        duty.id,
        (
            f"{emp.full_name}; {duty.duty_date.strftime('%d/%m/%Y')}; "
            f"reversão {-int(duty.minutes or 0)} min; "
            f"banco {old_balance} min -> {emp.bank_minutes} min; motivo: {reason}"
        ),
    )
    db.session.delete(duty)
    db.session.commit()

    flash("Plantão removido e crédito revertido do banco de horas.", "success")
    return redirect(url_for("rh.employee_detail", employee_id=emp.id) + "#plantoes")


@bp.route("/employees/<int:employee_id>/bank-adjustment", methods=["POST"])
@login_required
@roles_required(ROLE_ADMIN)
def employee_bank_adjustment(employee_id):
    emp = db.get_or_404(Employee, employee_id)
    operation = request.form.get("operation", "credit")
    hours = max(int(request.form.get("hours") or 0), 0)
    minutes_part = max(min(int(request.form.get("minutes") or 0), 59), 0)
    total = hours * 60 + minutes_part
    reason = (request.form.get("reason") or "").strip()
    if total <= 0 or not reason:
        flash("Informe a quantidade de horas/minutos e a justificativa do ajuste.", "danger")
        return redirect(url_for("rh.employee_detail", employee_id=employee_id) + "#banco-horas")
    signed = total if operation == "credit" else -total
    old_balance = emp.bank_minutes
    emp.bank_minutes += signed
    movement = BankHourAdjustment(employee_id=emp.id, minutes=signed, reason=reason, created_by=current_user.id)
    db.session.add(movement); db.session.flush()
    log_action("ajustou banco de horas", "bank_hour_adjustment", movement.id,
               f"{emp.full_name}: {old_balance} min -> {emp.bank_minutes} min; ajuste {signed:+d} min; motivo: {reason}")
    db.session.commit()
    flash("Banco de horas ajustado com sucesso.", "success")
    return redirect(url_for("rh.employee_detail", employee_id=employee_id) + "#banco-horas")


@bp.route("/employees/<int:employee_id>/time-clock/add", methods=["POST"])
@login_required
@roles_required(ROLE_ADMIN)
def employee_time_clock_add(employee_id):
    emp = db.get_or_404(Employee, employee_id)
    punch_date = parse_date(request.form.get("date"))
    punch_time = parse_time(request.form.get("time"))
    kind = request.form.get("kind")
    reason = (request.form.get("reason") or "").strip()
    allowed_kinds = {"entrada", "saida_intervalo", "retorno", "saida"}
    if not punch_date or not punch_time or kind not in allowed_kinds or not reason:
        flash("Preencha data, horário, tipo e justificativa do ajuste de ponto.", "danger")
        return redirect(url_for("rh.employee_detail", employee_id=employee_id) + "#controle-ponto")
    dt = datetime.combine(punch_date, punch_time)
    row = TimeClock(employee_id=emp.id, punched_at=dt, kind=kind, source="ajuste_rh",
                    ip_address=client_ip())
    db.session.add(row); db.session.flush()
    log_action("incluiu marcação de ponto pelo RH", "time_clock", row.id,
               f"{emp.full_name}; {dt.strftime('%d/%m/%Y %H:%M:%S')}; {kind}; motivo: {reason}")
    db.session.commit()
    flash("Marcação de ponto incluída com sucesso.", "success")
    return redirect(url_for("rh.employee_detail", employee_id=employee_id) + "#controle-ponto")


@bp.route("/time-clock/<int:clock_id>/edit", methods=["POST"])
@login_required
@roles_required(ROLE_ADMIN)
def time_clock_edit(clock_id):
    row = db.get_or_404(TimeClock, clock_id)
    punch_date = parse_date(request.form.get("date"))
    punch_time = parse_time(request.form.get("time"))
    kind = request.form.get("kind")
    reason = (request.form.get("reason") or "").strip()
    allowed_kinds = {"entrada", "saida_intervalo", "retorno", "saida"}
    if not punch_date or not punch_time or kind not in allowed_kinds or not reason:
        flash("Preencha data, horário, tipo e justificativa para alterar a marcação.", "danger")
        return redirect(url_for("rh.employee_detail", employee_id=row.employee_id) + "#controle-ponto")
    old = f"{row.punched_at.strftime('%d/%m/%Y %H:%M:%S')} {row.kind} origem={row.source}"
    row.punched_at = datetime.combine(punch_date, punch_time)
    row.kind = kind
    row.source = "ajuste_rh"
    new = f"{row.punched_at.strftime('%d/%m/%Y %H:%M:%S')} {row.kind}"
    log_action("alterou marcação de ponto pelo RH", "time_clock", row.id, f"antes: {old}; depois: {new}; motivo: {reason}")
    db.session.commit()
    flash("Marcação alterada com sucesso.", "success")
    return redirect(url_for("rh.employee_detail", employee_id=row.employee_id) + "#controle-ponto")


@bp.route("/time-clock/<int:clock_id>/delete", methods=["POST"])
@login_required
@roles_required(ROLE_ADMIN)
def time_clock_delete(clock_id):
    row = db.get_or_404(TimeClock, clock_id)
    employee_id = row.employee_id
    reason = (request.form.get("reason") or "").strip()
    if not reason:
        flash("Informe a justificativa para excluir a marcação duplicada/incorreta.", "danger")
        return redirect(url_for("rh.employee_detail", employee_id=employee_id) + "#controle-ponto")
    details = f"{row.employee.full_name}; registro #{row.id}; {row.punched_at.strftime('%d/%m/%Y %H:%M:%S')}; {row.kind}; origem={row.source}; motivo: {reason}"
    log_action("excluiu marcação de ponto pelo RH", "time_clock", row.id, details)
    db.session.delete(row)
    db.session.commit()
    flash("Marcação removida. A exclusão permanece registrada na Auditoria.", "success")
    return redirect(url_for("rh.employee_detail", employee_id=employee_id) + "#controle-ponto")


@bp.route("/clock", methods=["GET", "POST"])
@login_required
def clock():
    emp = current_user.employee
    if not emp: abort(403)
    if request.method == "POST":
        point_pin = (request.form.get("point_pin") or "").strip()
        if not emp.point_pin_hash:
            flash("Sua senha de ponto ainda não foi cadastrada. Procure o RH.", "danger")
            return redirect(url_for("rh.clock"))

        if _pin_is_blocked(emp):
            log_security_event(
                "pin_attempt_while_blocked",
                severity="critical",
                user=current_user,
                employee=emp,
                details="Tentativa de registro de ponto recebida durante bloqueio de PIN.",
            )
            db.session.commit()
            flash("Muitas tentativas de PIN. Aguarde 15 minutos antes de tentar novamente.", "danger")
            return redirect(url_for("rh.clock"))

        if not (len(point_pin) == 6 and point_pin.isdigit()) or not emp.check_point_pin(point_pin):
            _pin_failure(emp)
            log_action("tentativa inválida de registro de ponto", "employee", emp.id, "PIN de ponto inválido.")
            db.session.commit()
            flash("Senha de ponto inválida. O registro não foi realizado.", "danger")
            return redirect(url_for("rh.clock"))

        _pin_success(emp)
        kinds = ["entrada", "saida_intervalo", "retorno", "saida"]
        today_rows = TimeClock.query.filter(TimeClock.employee_id==emp.id, db.func.date(TimeClock.punched_at) == today_local()).order_by(TimeClock.punched_at).all()
        kind = kinds[min(len(today_rows), 3)]
        row = TimeClock(employee_id=emp.id, kind=kind, ip_address=client_ip())
        db.session.add(row); db.session.flush(); log_action("registrou ponto", "time_clock", row.id, kind); db.session.commit()
        flash(f"Ponto registrado: {kind.replace('_',' ')} às {row.punched_at.strftime('%H:%M:%S')}.", "success")
        return redirect(url_for("rh.clock"))
    rows = TimeClock.query.filter(TimeClock.employee_id==emp.id, db.func.date(TimeClock.punched_at) == today_local()).order_by(TimeClock.punched_at).all()
    return render_template("clock.html", rows=rows)

@bp.route("/certificates", methods=["GET", "POST"])
@login_required
def certificates():
    emp = current_user.employee
    if not emp: abort(403)
    if request.method == "POST":
        try: original, stored = save_upload(request.files.get("file"))
        except ValueError as e:
            flash(str(e), "danger"); return redirect(url_for("rh.certificates"))
        cert = MedicalCertificate(employee_id=emp.id, start_date=parse_date(request.form["start_date"]), days=int(request.form.get("days") or 1), note=request.form.get("note"), original_name=original, stored_name=stored)
        db.session.add(cert); db.session.flush(); log_action("enviou atestado", "medical_certificate", cert.id, original); db.session.commit()
        flash("Atestado enviado ao RH.", "success"); return redirect(url_for("rh.certificates"))
    rows = MedicalCertificate.query.filter_by(employee_id=emp.id).order_by(MedicalCertificate.uploaded_at.desc()).all()
    return render_template("certificates.html", rows=rows)


def _certificate_cover_pdf(cert):
    """Gera uma folha de identificação antes de cada atestado no PDF em lote."""
    output = BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=A4,
        leftMargin=18 * mm,
        rightMargin=18 * mm,
        topMargin=18 * mm,
        bottomMargin=18 * mm,
        title=f"Atestado - {cert.employee.full_name}",
        author="Portal RH",
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "CertificateBatchTitle",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=17,
        leading=21,
        alignment=TA_CENTER,
        spaceAfter=8 * mm,
    )
    normal = ParagraphStyle(
        "CertificateBatchNormal",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=10,
        leading=14,
    )
    data = [
        [Paragraph("<b>Colaborador</b>", normal), Paragraph(_pdf_text(cert.employee.full_name), normal)],
        [Paragraph("<b>Cargo</b>", normal), Paragraph(_pdf_text(cert.employee.job_title), normal)],
        [Paragraph("<b>Projeto/Unidade</b>", normal), Paragraph(_pdf_text(cert.employee.project), normal)],
        [Paragraph("<b>Data inicial do atestado</b>", normal), Paragraph(cert.start_date.strftime("%d/%m/%Y"), normal)],
        [Paragraph("<b>Dias</b>", normal), Paragraph(str(cert.days), normal)],
        [Paragraph("<b>Data de envio ao Portal</b>", normal), Paragraph(cert.uploaded_at.strftime("%d/%m/%Y %H:%M"), normal)],
        [Paragraph("<b>Status</b>", normal), Paragraph(_pdf_text(cert.status.title()), normal)],
        [Paragraph("<b>Horas abonadas pelo RH</b>", normal),
         Paragraph(
             _format_minutes(cert.allowance.minutes) if cert.allowance else "Aguardando definição do RH",
             normal
         )],
        [Paragraph("<b>Arquivo original</b>", normal), Paragraph(_pdf_text(cert.original_name), normal)],
        [Paragraph("<b>Observação</b>", normal), Paragraph(_pdf_text(cert.note or "—"), normal)],
    ]
    table = Table(data, colWidths=[52 * mm, 118 * mm], hAlign="LEFT")
    table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D9E2E8")),
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#F2F5F7")),
        ("LEFTPADDING", (0, 0), (-1, -1), 7),
        ("RIGHTPADDING", (0, 0), (-1, -1), 7),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]))
    story = [
        Paragraph("Atestado Médico — Documento Anexado", title_style),
        Paragraph(
            "Folha de identificação gerada automaticamente pelo Portal RH. "
            "O documento original anexado pelo colaborador aparece nas páginas seguintes.",
            normal,
        ),
        Spacer(1, 7 * mm),
        table,
    ]
    doc.build(story)
    output.seek(0)
    return output


def _image_file_to_pdf(path):
    """Converte JPG/PNG/WEBP em PDF para permitir impressão em lote."""
    output = BytesIO()
    with PILImage.open(path) as image:
        if getattr(image, "is_animated", False):
            image.seek(0)
        if image.mode not in ("RGB", "L"):
            background = PILImage.new("RGB", image.size, "white")
            if "A" in image.getbands():
                background.paste(image, mask=image.getchannel("A"))
            else:
                background.paste(image)
            image = background
        elif image.mode == "L":
            image = image.convert("RGB")
        else:
            image = image.copy()
        image.save(output, format="PDF", resolution=150.0)
    output.seek(0)
    return output


@bp.route("/certificates/print-batch.pdf")
@login_required
@roles_required(ROLE_ADMIN)
def certificates_print_batch():
    start_raw = (request.args.get("start_date") or "").strip()
    end_raw = (request.args.get("end_date") or "").strip()

    try:
        start = parse_date(start_raw)
        end = parse_date(end_raw)
    except (ValueError, TypeError):
        start = end = None

    if not start or not end:
        flash("Informe a data inicial e final para imprimir os atestados.", "danger")
        return redirect(url_for("rh.certificates_manage"))

    if end < start:
        flash("A data final não pode ser anterior à data inicial.", "danger")
        return redirect(url_for("rh.certificates_manage", start_date=start_raw, end_date=end_raw))

    if (end - start).days > 29:
        flash("O intervalo para impressão deve ter no máximo 30 dias corridos.", "danger")
        return redirect(url_for("rh.certificates_manage", start_date=start_raw, end_date=end_raw))

    # O pedido é pelos documentos ANEXADOS no período; portanto usamos uploaded_at.
    rows = (MedicalCertificate.query
            .filter(
                func.date(MedicalCertificate.uploaded_at) >= start,
                func.date(MedicalCertificate.uploaded_at) <= end,
            )
            .order_by(MedicalCertificate.uploaded_at.asc(), MedicalCertificate.id.asc())
            .all())

    if not rows:
        flash("Nenhum atestado foi anexado ao Portal nesse intervalo.", "danger")
        return redirect(url_for("rh.certificates_manage", start_date=start_raw, end_date=end_raw))

    writer = PdfWriter()
    included = 0
    skipped = []

    for cert in rows:
        # Sempre inclui uma capa de identificação do documento.
        cover = _certificate_cover_pdf(cert)
        cover_reader = PdfReader(cover)
        for page in cover_reader.pages:
            writer.add_page(page)

        file_path = os.path.join(current_app.config["UPLOAD_FOLDER"], cert.stored_name)
        if not os.path.isfile(file_path):
            skipped.append(f"{cert.employee.full_name}: arquivo não localizado")
            continue

        ext = cert.stored_name.rsplit(".", 1)[-1].lower() if "." in cert.stored_name else ""
        try:
            if ext == "pdf":
                source = PdfReader(file_path)
                for page in source.pages:
                    writer.add_page(page)
            elif ext in {"jpg", "jpeg", "png", "webp"}:
                image_pdf = _image_file_to_pdf(file_path)
                source = PdfReader(image_pdf)
                for page in source.pages:
                    writer.add_page(page)
            else:
                skipped.append(f"{cert.employee.full_name}: formato não suportado")
                continue
            included += 1
        except Exception:
            skipped.append(f"{cert.employee.full_name}: não foi possível processar o arquivo")

    if included == 0:
        flash("Os atestados foram encontrados, mas nenhum arquivo pôde ser processado para impressão.", "danger")
        return redirect(url_for("rh.certificates_manage", start_date=start_raw, end_date=end_raw))

    output = BytesIO()
    writer.write(output)
    output.seek(0)

    details = (
        f"Período de envio {start.strftime('%d/%m/%Y')} a {end.strftime('%d/%m/%Y')}; "
        f"{included} arquivo(s) incluído(s)"
    )
    if skipped:
        details += f"; {len(skipped)} arquivo(s) com falha"
    log_action("gerou impressão em lote de atestados", "medical_certificate", None, details)
    db.session.commit()

    filename = f"atestados_{start.strftime('%Y-%m-%d')}_a_{end.strftime('%Y-%m-%d')}.pdf"
    return send_file(
        output,
        mimetype="application/pdf",
        as_attachment=False,
        download_name=filename,
    )


@bp.route("/certificates/manage")
@login_required
@roles_required(ROLE_ADMIN)
def certificates_manage():
    q = MedicalCertificate.query
    employee_id = request.args.get("employee_id", type=int)
    status = request.args.get("status", "").strip()
    start_date = request.args.get("start_date", "").strip()
    end_date = request.args.get("end_date", "").strip()

    if employee_id:
        q = q.filter(MedicalCertificate.employee_id == employee_id)
    if status:
        q = q.filter(MedicalCertificate.status == status)
    if start_date:
        q = q.filter(MedicalCertificate.start_date >= parse_date(start_date))
    if end_date:
        q = q.filter(MedicalCertificate.start_date <= parse_date(end_date))

    rows = q.order_by(MedicalCertificate.uploaded_at.desc()).all()
    employees = Employee.query.order_by(Employee.full_name).all()
    local_today = today_local()
    batch_end_date = local_today.isoformat()
    batch_start_date = (local_today - timedelta(days=29)).isoformat()
    return render_template(
        "certificates_manage.html",
        rows=rows,
        employees=employees,
        selected_employee=employee_id,
        selected_status=status,
        start_date=start_date,
        end_date=end_date,
        batch_start_date=batch_start_date,
        batch_end_date=batch_end_date,
    )



@bp.route("/certificates/<int:certificate_id>/allowance", methods=["POST"])
@login_required
@roles_required(ROLE_ADMIN)
def certificate_allowance(certificate_id):
    cert = db.get_or_404(MedicalCertificate, certificate_id)
    emp = cert.employee

    try:
        hours = max(int(request.form.get("hours") or 0), 0)
        minutes_part = max(min(int(request.form.get("minutes") or 0), 59), 0)
    except (TypeError, ValueError):
        flash("Informe uma quantidade válida de horas e minutos.", "danger")
        return redirect(request.referrer or url_for("rh.certificates_manage"))

    total = hours * 60 + minutes_part
    note = (request.form.get("allowance_note") or "").strip() or None

    # Máximo operacional: jornada prevista multiplicada pelos dias úteis abrangidos.
    expected_daily = _expected_daily_minutes(emp)
    cert_end = cert.start_date + timedelta(days=max(int(cert.days or 1) - 1, 0))
    workdays = 0
    cursor = cert.start_date
    while cursor <= cert_end:
        if cursor.weekday() < 5 and cursor >= emp.admission_date:
            workdays += 1
        cursor += timedelta(days=1)
    max_minutes = expected_daily * workdays

    if total > max_minutes:
        flash(
            f"O abono informado excede a jornada prevista para os dias úteis do atestado "
            f"({max_minutes // 60:02d}:{max_minutes % 60:02d}).",
            "danger",
        )
        return redirect(request.referrer or url_for("rh.certificates_manage"))

    existing = MedicalCertificateAllowance.query.filter_by(certificate_id=cert.id).first()

    if total == 0:
        if existing:
            db.session.delete(existing)
            log_action(
                "removeu abono por atestado",
                "medical_certificate_allowance",
                existing.id,
                f"{emp.full_name}; atestado {cert.id}; início {cert.start_date.strftime('%d/%m/%Y')}",
            )
            db.session.commit()
            flash("Abono por atestado removido.", "success")
        else:
            flash("Esse atestado não possuía horas abonadas.", "warning")
        return redirect(request.referrer or url_for("rh.certificates_manage"))

    if existing:
        old_minutes = int(existing.minutes or 0)
        existing.minutes = total
        existing.note = note
        existing.approved_by = current_user.id
        existing.updated_at = now_local()
        item = existing
        action = "alterou abono por atestado"
        detail_change = f"{old_minutes} min -> {total} min"
    else:
        item = MedicalCertificateAllowance(
            certificate_id=cert.id,
            employee_id=emp.id,
            minutes=total,
            note=note,
            approved_by=current_user.id,
        )
        db.session.add(item)
        db.session.flush()
        action = "registrou abono por atestado"
        detail_change = f"{total} min"

    log_action(
        action,
        "medical_certificate_allowance",
        item.id,
        (
            f"{emp.full_name}; atestado {cert.id}; início {cert.start_date.strftime('%d/%m/%Y')}; "
            f"abono {detail_change}; observação: {note or '-'}"
        ),
    )
    db.session.commit()
    flash(
        f"Abono de {total // 60:02d}:{total % 60:02d} registrado. "
        "As horas passarão a compor a jornada justificada no fechamento mensal.",
        "success",
    )
    return redirect(request.referrer or url_for("rh.certificates_manage"))


@bp.route("/certificates/<int:certificate_id>/status", methods=["POST"])
@login_required
@roles_required(ROLE_ADMIN)
def certificate_status(certificate_id):
    cert = db.get_or_404(MedicalCertificate, certificate_id)
    new_status = request.form.get("status", "")
    allowed = {"recebido", "conferido", "arquivado", "devolvido"}
    if new_status not in allowed:
        abort(400)
    old_status = cert.status
    cert.status = new_status
    log_action(
        "alterou status do atestado",
        "medical_certificate",
        cert.id,
        f"{old_status} -> {new_status}",
    )
    db.session.commit()
    flash("Status do atestado atualizado.", "success")
    return redirect(request.referrer or url_for("rh.certificates_manage"))


@bp.route("/requests", methods=["GET", "POST"])
@login_required
def requests_page():
    emp = current_user.employee
    if not emp: abort(403)
    if request.method == "POST":
        rtype = request.form["request_type"]
        start = parse_time(request.form.get("start_time")); end = parse_time(request.form.get("end_time"))
        minutes = int(request.form.get("minutes") or 0)
        if not minutes and start and end:
            a = datetime.combine(today_local(), start); b = datetime.combine(today_local(), end); minutes = int((b-a).total_seconds()/60)
        item = Request(employee_id=emp.id, request_type=rtype, request_date=parse_date(request.form["request_date"]), start_time=start, end_time=end, minutes=max(minutes,0), reason=request.form["reason"], target_clock_kind=request.form.get("target_clock_kind") or None)
        db.session.add(item); db.session.flush(); log_action("criou solicitação", "request", item.id, rtype); db.session.commit()
        flash("Solicitação enviada para aprovação.", "success"); return redirect(url_for("rh.requests_page"))
    rows = Request.query.filter_by(employee_id=emp.id).order_by(Request.requested_at.desc()).all()
    bank_summary = _bank_summary(emp)
    return render_template("requests.html", rows=rows, bank_summary=bank_summary)

@bp.route("/approvals")
@login_required
@roles_required(ROLE_ADMIN, ROLE_MANAGER)
def approvals():
    q = Request.query.filter_by(status="pending")
    if current_user.role == ROLE_MANAGER:
        ids = [e.id for e in current_user.employee.team]
        q = q.filter(Request.employee_id.in_(ids))
    return render_template("approvals.html", rows=q.order_by(Request.requested_at).all())

@bp.route("/requests/<int:request_id>/decision", methods=["POST"])
@login_required
@roles_required(ROLE_ADMIN, ROLE_MANAGER)
def request_decision(request_id):
    item = db.get_or_404(Request, request_id)
    if current_user.role == ROLE_MANAGER and item.employee.manager_id != current_user.employee.id: abort(403)
    decision = request.form["decision"]
    if decision not in ["approved", "rejected"]: abort(400)
    item.status=decision; item.decided_at=now_local(); item.decided_by=current_user.id; item.decision_note=request.form.get("decision_note")
    if decision == "approved":
        if item.request_type == "bank_use":
            if item.request_date <= today_local():
                item.employee.bank_minutes -= int(item.minutes or 0)
                item.bank_effect_applied = True
                item.bank_effect_applied_at = now_local()
            else:
                # Solicitação futura: aprovada e reservada, sem reduzir o saldo realizado ainda.
                item.bank_effect_applied = False
        elif item.request_type == "overtime": item.employee.bank_minutes += item.minutes
        elif item.request_type == "clock_adjustment" and item.start_time:
            dt = datetime.combine(item.request_date, item.start_time)
            tc = TimeClock(employee_id=item.employee_id, punched_at=dt, kind=item.target_clock_kind or "ajuste", source="ajuste_aprovado")
            db.session.add(tc)
    log_action(f"{decision} solicitação", "request", item.id, item.request_type); db.session.commit()
    flash("Decisão registrada.", "success"); return redirect(url_for("rh.approvals"))


@bp.route("/payslips")
@login_required
def payslips():
    if current_user.role == ROLE_ADMIN:
        return redirect(url_for("rh.payslips_manage"))

    emp = current_user.employee
    if not emp:
        abort(403)

    rows = (Payslip.query
            .filter(Payslip.employee_id == emp.id)
            .order_by(Payslip.year.desc(), Payslip.month.desc(), Payslip.uploaded_at.desc())
            .all())

    return render_template(
        "payslips.html",
        rows=rows,
        emp=emp,
    )


@bp.route("/payslips/manage")
@login_required
@roles_required(ROLE_ADMIN)
def payslips_manage():
    employees = Employee.query.filter_by(is_active=True).order_by(Employee.full_name).all()
    rows = (Payslip.query
            .order_by(Payslip.year.desc(), Payslip.month.desc(), Payslip.uploaded_at.desc())
            .limit(300).all())
    return render_template(
        "payslips_manage.html",
        employees=employees,
        rows=rows,
        current_competence=today_local().strftime("%Y-%m"),
    )


@bp.route("/payslips/upload-batch", methods=["POST"])
@login_required
@roles_required(ROLE_ADMIN)
def payslips_upload_batch():
    try:
        year, month = _parse_competence(request.form.get("competence"))
    except ValueError as exc:
        flash(str(exc), "danger")
        return redirect(url_for("rh.payslips_manage"))

    file = request.files.get("file")
    if not file or not file.filename:
        flash("Selecione o PDF consolidado dos holerites.", "danger")
        return redirect(url_for("rh.payslips_manage"))

    if "." not in file.filename or file.filename.rsplit(".", 1)[1].lower() != "pdf":
        flash("O arquivo consolidado precisa estar em PDF.", "danger")
        return redirect(url_for("rh.payslips_manage"))

    try:
        _validate_pdf_stream(file)
    except ValueError as exc:
        flash(str(exc), "danger")
        return redirect(url_for("rh.payslips_manage"))

    employees = Employee.query.filter_by(is_active=True).order_by(Employee.full_name).all()

    try:
        groups, unmatched = _split_payslip_pdf_by_employee(file, employees)
    except ValueError as exc:
        flash(str(exc), "danger")
        return redirect(url_for("rh.payslips_manage"))

    created = []
    skipped_without_access = []
    for group in groups.values():
        emp = group["employee"]

        # O holerite precisa estar associado a um cadastro com acesso ao Portal.
        if not emp.user:
            skipped_without_access.append(emp.full_name)
            continue

        item = _save_employee_payslip_pages(
            emp=emp,
            year=year,
            month=month,
            pages=group["pages"],
            original_name=file.filename,
            matched_by="automatic_pdf_page",
        )
        created.append((emp, item, group["page_numbers"]))
        log_action(
            "separou holerite de PDF consolidado",
            "payslip",
            item.id,
            (
                f"{emp.full_name}; competência {month:02d}/{year}; "
                f"página(s) {', '.join(map(str, group['page_numbers']))}"
            ),
        )

    db.session.commit()

    if created:
        flash(
            f"{len(created)} colaborador(es) receberam holerite da competência {month:02d}/{year}.",
            "success",
        )
    else:
        flash(
            "Nenhuma página pôde ser associada automaticamente a um colaborador.",
            "danger",
        )

    if skipped_without_access:
        flash(
            f"{len(skipped_without_access)} colaborador(es) foram identificados no PDF, "
            "mas não possuem usuário de acesso vinculado ao Portal RH: "
            + ", ".join(skipped_without_access[:8]),
            "warning",
        )

    if unmatched:
        reason_counts = {}
        for row in unmatched:
            reason_counts[row["reason"]] = reason_counts.get(row["reason"], 0) + 1

        details = []
        if reason_counts.get("sem_texto"):
            details.append(f"{reason_counts['sem_texto']} sem texto pesquisável")
        if reason_counts.get("nao_encontrado"):
            details.append(f"{reason_counts['nao_encontrado']} sem nome cadastrado identificado")
        if reason_counts.get("ambiguo"):
            details.append(f"{reason_counts['ambiguo']} com identificação ambígua")

        page_list = ", ".join(str(row["page"]) for row in unmatched[:20])
        if len(unmatched) > 20:
            page_list += "..."

        flash(
            f"{len(unmatched)} página(s) não foram distribuídas automaticamente "
            f"(páginas: {page_list}). " + "; ".join(details) + ".",
            "warning",
        )

    return redirect(url_for("rh.payslips_manage"))


@bp.route("/payslips/upload-manual", methods=["POST"])
@login_required
@roles_required(ROLE_ADMIN)
def payslips_upload_manual():
    emp = db.get_or_404(Employee, request.form.get("employee_id", type=int))
    try:
        year, month = _parse_competence(request.form.get("competence"))
        item = _upsert_payslip(emp, year, month, request.files.get("file"), "manual")
    except ValueError as exc:
        flash(str(exc), "danger")
        return redirect(url_for("rh.payslips_manage"))

    log_action(
        "anexou holerite manualmente",
        "payslip",
        item.id,
        f"{emp.full_name}; competência {month:02d}/{year}; arquivo {item.original_name}"
    )
    db.session.commit()
    flash(f"Holerite de {emp.full_name} — {month:02d}/{year} — salvo com sucesso.", "success")
    return redirect(url_for("rh.payslips_manage"))


@bp.route("/payslips/<int:payslip_id>/delete", methods=["POST"])
@login_required
@roles_required(ROLE_ADMIN)
def payslip_delete(payslip_id):
    item = db.get_or_404(Payslip, payslip_id)
    stored = item.stored_name
    description = f"{item.employee.full_name}; competência {item.month:02d}/{item.year}"
    db.session.delete(item)
    log_action("removeu holerite", "payslip", payslip_id, description)
    db.session.commit()
    path = os.path.join(current_app.config["UPLOAD_FOLDER"], stored)
    try:
        if os.path.isfile(path):
            os.remove(path)
    except OSError:
        pass
    flash("Holerite removido.", "success")
    return redirect(url_for("rh.payslips_manage"))


@bp.route("/documents/<int:employee_id>", methods=["POST"])
@login_required
@roles_required(ROLE_ADMIN)
def document_upload(employee_id):
    emp = db.get_or_404(Employee, employee_id)
    require_signature = request.form.get("require_signature") == "1"
    try:
        original, stored = save_upload(request.files.get("file"))
    except ValueError as e:
        flash(str(e), "danger")
        return redirect(url_for("rh.employee_detail", employee_id=employee_id) + "#documentos")

    if require_signature and not original.lower().endswith(".pdf"):
        path = os.path.join(current_app.config["UPLOAD_FOLDER"], stored)
        try:
            if os.path.isfile(path):
                os.remove(path)
        except OSError:
            pass
        flash("Documentos destinados à assinatura eletrônica devem ser enviados em PDF.", "danger")
        return redirect(url_for("rh.employee_detail", employee_id=employee_id) + "#documentos")

    doc = Document(
        employee_id=emp.id,
        category=request.form.get("category") or "Outros",
        title=request.form.get("title") or original,
        original_name=original,
        stored_name=stored,
    )
    db.session.add(doc)
    db.session.flush()

    if require_signature:
        flow = DocumentSignatureFlow(
            document_id=doc.id,
            requested_by=current_user.id,
        )
        db.session.add(flow)
        log_action(
            "solicitou assinatura eletrônica de documento",
            "document",
            doc.id,
            f"{emp.full_name}; {doc.title}; arquivo {original}",
        )
        log_security_event(
            "document_signature_requested",
            severity="info",
            user=current_user,
            employee=emp,
            details=f"Assinatura solicitada para documento #{doc.id}: {doc.title}.",
        )
    else:
        log_action("anexou documento", "document", doc.id, original)

    db.session.commit()
    flash(
        "Documento enviado para assinatura do colaborador."
        if require_signature else
        "Documento anexado.",
        "success",
    )
    return redirect(url_for("rh.employee_detail", employee_id=employee_id) + "#documentos")


@bp.route("/documents/signatures")
@login_required
@roles_required(ROLE_ADMIN)
def document_signatures_manage():
    flows = (
        DocumentSignatureFlow.query
        .join(Document)
        .order_by(DocumentSignatureFlow.requested_at.desc())
        .all()
    )
    counts = {
        "total": len(flows),
        "employee_pending": sum(1 for flow in flows if flow.status in {"pending", "viewed"}),
        "rh_pending": sum(1 for flow in flows if flow.status == "awaiting_rh"),
        "finalized": sum(1 for flow in flows if flow.status == "finalized"),
    }
    return render_template("document_signatures.html", flows=flows, counts=counts)


@bp.route("/documents/<int:document_id>/sign", methods=["POST"])
@login_required
def document_sign(document_id):
    doc = db.get_or_404(Document, document_id)
    flow = doc.signature_flow
    emp = doc.employee

    if not current_user.employee or current_user.employee.id != emp.id:
        abort(403)
    if not flow or flow.cancelled_at or flow.finalized_at:
        flash("Este documento não está disponível para assinatura.", "danger")
        return redirect(url_for("rh.employee_detail", employee_id=emp.id) + "#documentos")
    if flow.signed_at:
        flash("Este documento já foi assinado.", "warning")
        return redirect(url_for("rh.employee_detail", employee_id=emp.id) + "#documentos")
    if not flow.employee_viewed_at:
        flash("Abra e confira o documento antes de assinar.", "danger")
        return redirect(url_for("rh.employee_detail", employee_id=emp.id) + "#documentos")
    if request.form.get("confirm_ack") != "1":
        flash("Confirme que leu e conferiu o documento.", "danger")
        return redirect(url_for("rh.employee_detail", employee_id=emp.id) + "#documentos")

    if _pin_is_blocked(emp):
        log_security_event(
            "pin_attempt_while_blocked",
            severity="critical",
            user=current_user,
            employee=emp,
            details=f"Tentativa de assinatura do documento #{doc.id} durante bloqueio de PIN.",
        )
        db.session.commit()
        flash("Senha de ponto temporariamente bloqueada por tentativas inválidas.", "danger")
        return redirect(url_for("rh.employee_detail", employee_id=emp.id) + "#documentos")

    pin = (request.form.get("point_pin") or "").strip()
    if not emp.check_point_pin(pin):
        _pin_failure(emp)
        flash("Senha de ponto inválida.", "danger")
        return redirect(url_for("rh.employee_detail", employee_id=emp.id) + "#documentos")

    _pin_success(emp)
    stamp = now_local()
    raw_code = f"doc|{doc.id}|{emp.id}|{stamp.isoformat()}|{uuid.uuid4().hex}"
    flow.signed_at = stamp
    flow.signature_code = hashlib.sha256(raw_code.encode("utf-8")).hexdigest()[:32].upper()
    flow.signer_ip = client_ip()

    log_action(
        "assinou eletronicamente documento",
        "document",
        doc.id,
        f"{emp.full_name}; {doc.title}; código {flow.signature_code}",
    )
    log_security_event(
        "electronic_document_signature",
        severity="info",
        user=current_user,
        employee=emp,
        details=f"Documento #{doc.id} assinado eletronicamente; código {flow.signature_code}.",
    )
    db.session.commit()
    flash("Documento assinado eletronicamente. Agora aguarda validação final do RH.", "success")
    return redirect(url_for("rh.employee_detail", employee_id=emp.id) + "#documentos")


@bp.route("/documents/<int:document_id>/finalize", methods=["POST"])
@login_required
@roles_required(ROLE_ADMIN)
def document_signature_finalize(document_id):
    doc = db.get_or_404(Document, document_id)
    flow = doc.signature_flow
    if not flow or flow.cancelled_at or not flow.signed_at:
        flash("O documento ainda não possui assinatura válida para finalizar.", "danger")
        return redirect(url_for("rh.document_signatures_manage"))

    if not flow.finalized_at:
        flow.finalized_at = now_local()
        flow.finalized_by = current_user.id
        flow.final_note = (request.form.get("note") or "").strip() or None
        log_action(
            "validou assinatura eletrônica de documento",
            "document",
            doc.id,
            f"{doc.employee.full_name}; {doc.title}; código {flow.signature_code}",
        )
        log_security_event(
            "document_signature_finalized",
            severity="info",
            user=current_user,
            employee=doc.employee,
            details=f"Documento #{doc.id} validado e arquivado pelo RH.",
        )
        db.session.commit()

    flash("Assinatura validada e documento arquivado.", "success")
    return redirect(url_for("rh.document_signatures_manage"))


@bp.route("/documents/<int:document_id>/signature-proof.pdf")
@login_required
def document_signature_proof(document_id):
    doc = db.get_or_404(Document, document_id)
    flow = doc.signature_flow
    emp = doc.employee

    if current_user.role != ROLE_ADMIN and (
        not current_user.employee or current_user.employee.id != emp.id
    ):
        abort(403)
    if not flow or not flow.signed_at:
        abort(404)

    file_path = os.path.join(current_app.config["UPLOAD_FOLDER"], doc.stored_name)
    document_hash = "arquivo indisponível"
    if os.path.isfile(file_path):
        sha = hashlib.sha256()
        with open(file_path, "rb") as handle:
            for chunk in iter(lambda: handle.read(1024 * 1024), b""):
                sha.update(chunk)
        document_hash = sha.hexdigest().upper()

    buffer = BytesIO()
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ProofTitle",
        parent=styles["Heading1"],
        fontSize=16,
        leading=19,
        alignment=TA_CENTER,
        spaceAfter=10,
    )
    body_style = ParagraphStyle(
        "ProofBody",
        parent=styles["BodyText"],
        fontSize=9,
        leading=13,
    )
    story = [
        Paragraph("COMPROVANTE DE ASSINATURA ELETRÔNICA", title_style),
        Paragraph("Portal RH — Associação Alecrim Dourado", styles["Heading3"]),
        Spacer(1, 8),
        Table([
            ["Colaborador", emp.full_name],
            ["Documento", doc.title],
            ["Categoria", doc.category],
            ["Arquivo original", doc.original_name],
            ["Enviado em", doc.uploaded_at.strftime("%d/%m/%Y %H:%M")],
            ["Visualizado em", flow.employee_viewed_at.strftime("%d/%m/%Y %H:%M:%S") if flow.employee_viewed_at else "—"],
            ["Assinado em", flow.signed_at.strftime("%d/%m/%Y %H:%M:%S")],
            ["Código da assinatura", flow.signature_code or "—"],
            ["IP registrado", flow.signer_ip or "—"],
            ["Hash SHA-256 do documento", document_hash],
            ["Validação do RH", flow.finalized_at.strftime("%d/%m/%Y %H:%M:%S") if flow.finalized_at else "Aguardando validação"],
            ["Responsável RH", flow.finalizer.email if flow.finalizer else "—"],
            ["Observação final", flow.final_note or "—"],
        ], colWidths=[45*mm, 135*mm], style=[
            ("GRID", (0,0), (-1,-1), .35, colors.HexColor("#cfd8d4")),
            ("BACKGROUND", (0,0), (0,-1), colors.HexColor("#f2f5f4")),
            ("FONTNAME", (0,0), (0,-1), "Helvetica-Bold"),
            ("VALIGN", (0,0), (-1,-1), "TOP"),
            ("FONTSIZE", (0,0), (-1,-1), 8),
            ("LEFTPADDING", (0,0), (-1,-1), 6),
            ("RIGHTPADDING", (0,0), (-1,-1), 6),
            ("TOPPADDING", (0,0), (-1,-1), 6),
            ("BOTTOMPADDING", (0,0), (-1,-1), 6),
        ]),
        Spacer(1, 12),
        Paragraph(
            "Este comprovante registra a ciência e a assinatura eletrônica realizada dentro do Portal RH. "
            "O código de assinatura e o hash SHA-256 vinculam este comprovante à versão do arquivo disponibilizada ao colaborador. "
            "Este mecanismo representa aceite eletrônico interno do Portal e não corresponde a certificado digital ICP-Brasil.",
            body_style,
        ),
    ]
    pdf = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=15*mm,
        leftMargin=15*mm,
        topMargin=15*mm,
        bottomMargin=15*mm,
    )
    pdf.build(story)
    buffer.seek(0)

    log_action("gerou comprovante de assinatura de documento", "document", doc.id, doc.title)
    db.session.commit()
    return send_file(
        buffer,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=f"comprovante-assinatura-documento-{doc.id}.pdf",
    )

@bp.route("/files/<string:kind>/<int:item_id>")
@login_required
def file_download(kind, item_id):
    if kind == "certificate":
        item = db.get_or_404(MedicalCertificate, item_id); emp=item.employee
        # Gestores não recebem acesso ao conteúdo médico; apenas RH e titular.
        if current_user.role != ROLE_ADMIN and (not current_user.employee or current_user.employee.id != emp.id): abort(403)
    elif kind == "document":
        item = db.get_or_404(Document, item_id); emp=item.employee
        if current_user.role != ROLE_ADMIN and (not current_user.employee or current_user.employee.id != emp.id):
            abort(403)
        if (
            current_user.role != ROLE_ADMIN
            and item.signature_flow
            and not item.signature_flow.cancelled_at
            and not item.signature_flow.employee_viewed_at
        ):
            item.signature_flow.employee_viewed_at = now_local()
            log_action(
                "visualizou documento para assinatura",
                "document",
                item.id,
                f"{emp.full_name}; {item.title}",
            )
            db.session.commit()
    elif kind == "payslip":
        item = db.get_or_404(Payslip, item_id); emp=item.employee
        # Holerite contém informação remuneratória: acesso apenas do RH e do titular.
        if current_user.role != ROLE_ADMIN and (not current_user.employee or current_user.employee.id != emp.id):
            abort(403)
        if current_user.role != ROLE_ADMIN and not item.employee_viewed_at:
            item.employee_viewed_at = now_local()
            log_action(
                "visualizou holerite",
                "payslip",
                item.id,
                f"{emp.full_name}; competência {item.month:02d}/{item.year}"
            )
            db.session.commit()
    else:
        abort(404)

    access_label = {
        "certificate": "Atestado",
        "document": "Documento",
        "payslip": "Holerite",
    }.get(kind, "Arquivo")
    log_security_event(
        "sensitive_file_access",
        severity="info",
        user=current_user,
        employee=emp,
        details=f"{access_label} #{item.id} acessado; arquivo: {item.original_name}.",
    )
    db.session.commit()

    response = send_from_directory(
        current_app.config["UPLOAD_FOLDER"],
        item.stored_name,
        as_attachment=False,
        download_name=item.original_name,
        conditional=False,
    )
    response.headers["Cache-Control"] = "no-store, private, max-age=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["X-Content-Type-Options"] = "nosniff"
    return response



@bp.route("/employees/<int:employee_id>/vacations", methods=["POST"])
@login_required
@roles_required(ROLE_ADMIN)
def employee_vacation_add(employee_id):
    emp = db.get_or_404(Employee, employee_id)
    start = parse_date(request.form.get("start_date"))
    days = int(request.form.get("days") or 0)
    note = (request.form.get("note") or "").strip() or None
    if not start or days <= 0:
        flash("Informe a data de início e a quantidade de dias das férias.", "danger")
        return redirect(url_for("rh.employee_detail", employee_id=employee_id) + "#ferias")
    item = Vacation(employee_id=emp.id, start_date=start, days=days, note=note, created_by=current_user.id)
    db.session.add(item); db.session.flush()
    log_action("registrou férias", "vacation", item.id, f"{emp.full_name}; {start.strftime('%d/%m/%Y')}; {days} dias")
    db.session.commit()
    flash("Férias registradas com sucesso.", "success")
    return redirect(url_for("rh.employee_detail", employee_id=employee_id) + "#ferias")


@bp.route("/employees/<int:employee_id>/vacation-schedule", methods=["POST"])
@login_required
@roles_required(ROLE_ADMIN)
def employee_vacation_schedule_add(employee_id):
    emp = db.get_or_404(Employee, employee_id)
    planned_start = parse_date(request.form.get("planned_start"))
    planned_return = parse_date(request.form.get("planned_return"))
    days = int(request.form.get("days") or 0)
    note = (request.form.get("note") or "").strip() or None
    if not planned_start or days <= 0 or days > 30:
        flash("Informe uma data válida e entre 1 e 30 dias para a programação de férias.", "danger")
        return redirect(url_for("rh.employee_detail", employee_id=employee_id) + "#ferias")
    if not planned_return:
        planned_return = planned_start + timedelta(days=days)
    if planned_return <= planned_start:
        flash("A data prevista de retorno deve ser posterior ao início das férias.", "danger")
        return redirect(url_for("rh.employee_detail", employee_id=employee_id) + "#ferias")
    item = VacationSchedule(
        employee_id=emp.id,
        planned_start=planned_start,
        planned_return=planned_return,
        days=days,
        note=note,
        created_by=current_user.id,
        status="planned",
    )
    db.session.add(item); db.session.flush()
    log_action("programou férias", "vacation_schedule", item.id,
               f"{emp.full_name}; início {planned_start.strftime('%d/%m/%Y')}; retorno {planned_return.strftime('%d/%m/%Y')}; {days} dias")
    db.session.commit()
    flash("Previsão de férias registrada. Os dias ainda não foram descontados como férias realizadas.", "success")
    return redirect(url_for("rh.employee_detail", employee_id=employee_id) + "#ferias")

@bp.route("/vacation-schedules/<int:schedule_id>/complete", methods=["POST"])
@login_required
@roles_required(ROLE_ADMIN)
def vacation_schedule_complete(schedule_id):
    item = db.get_or_404(VacationSchedule, schedule_id)
    if item.status != "planned":
        flash("Essa programação já foi finalizada ou cancelada.", "danger")
        return redirect(url_for("rh.employee_detail", employee_id=item.employee_id) + "#ferias")
    vacation = Vacation(
        employee_id=item.employee_id,
        start_date=item.planned_start,
        days=item.days,
        note=item.note or "Férias realizadas conforme programação",
        created_by=current_user.id,
    )
    db.session.add(vacation)
    item.status = "completed"
    item.completed_at = now_local()
    log_action("confirmou férias programadas como realizadas", "vacation_schedule", item.id,
               f"{item.employee.full_name}; {item.planned_start.strftime('%d/%m/%Y')}; {item.days} dias")
    db.session.commit()
    flash("Férias marcadas como realizadas e descontadas do saldo disponível.", "success")
    return redirect(url_for("rh.employee_detail", employee_id=item.employee_id) + "#ferias")

@bp.route("/vacation-schedules/<int:schedule_id>/cancel", methods=["POST"])
@login_required
@roles_required(ROLE_ADMIN)
def vacation_schedule_cancel(schedule_id):
    item = db.get_or_404(VacationSchedule, schedule_id)
    if item.status == "planned":
        item.status = "cancelled"
        log_action("cancelou programação de férias", "vacation_schedule", item.id,
                   f"{item.employee.full_name}; previsão {item.planned_start.strftime('%d/%m/%Y')}")
        db.session.commit()
        flash("Programação de férias cancelada.", "success")
    return redirect(url_for("rh.employee_detail", employee_id=item.employee_id) + "#ferias")

@bp.route("/bank-statement")
@login_required
@roles_required(ROLE_ADMIN)
def bank_statement():
    employee_id = request.args.get("employee_id", type=int)
    employees = Employee.query.order_by(Employee.full_name).all()
    emp = db.session.get(Employee, employee_id) if employee_id else None
    movements = []
    if emp:
        adjustments = BankHourAdjustment.query.filter_by(employee_id=emp.id).all()
        duties = WeekendDuty.query.filter_by(employee_id=emp.id).all()
        requests_rows = Request.query.filter(
            Request.employee_id == emp.id, Request.status == "approved",
            Request.request_type.in_(["overtime","bank_use"])
        ).all()
        for a in adjustments:
            movements.append({"date": a.created_at, "label": a.reason, "minutes": int(a.minutes or 0), "source": "Ajuste RH"})
        for duty in duties:
            duty_dt = datetime.combine(duty.duty_date, datetime.min.time())
            movements.append({
                "date": duty_dt,
                "label": f"Plantão executado - {duty.duty_date.strftime('%d/%m/%Y')}" + (f" - {duty.note}" if duty.note else ""),
                "minutes": int(duty.minutes or 0),
                "source": "Plantão executado",
            })
        for q in requests_rows:
            sign = 1 if q.request_type == "overtime" else -1
            movements.append({"date": datetime.combine(q.request_date, q.start_time or datetime.min.time()),
                              "label": "Hora extra aprovada" if sign > 0 else "Utilização de banco",
                              "minutes": sign * int(q.minutes or 0), "source": "Solicitação aprovada"})
        movements.sort(key=lambda x:x["date"])
        running = 0
        for x in movements:
            running += x["minutes"]; x["running"] = running
    return render_template("bank_statement.html", employees=employees, emp=emp, movements=movements)


@bp.route("/my-time-reports")
@login_required
def my_time_reports():
    if current_user.role == ROLE_ADMIN:
        return redirect(url_for("rh.time_closing"))

    emp = current_user.employee
    if not emp:
        abort(403)

    rows = (
        TimeReportFinalization.query
        .filter_by(employee_id=emp.id)
        .order_by(TimeReportFinalization.year.desc(), TimeReportFinalization.month.desc())
        .all()
    )
    return render_template("my_time_reports.html", emp=emp, rows=rows)


@bp.route("/time-closing")
@login_required
@roles_required(ROLE_ADMIN)
def time_closing():
    month_value = request.args.get("month") or today_local().strftime("%Y-%m")
    try: year, month = map(int, month_value.split("-"))
    except Exception: year, month = today_local().year, today_local().month
    employees = Employee.query.filter_by(is_active=True).order_by(Employee.full_name).all()
    summaries=[]
    for emp in employees:
        s=_month_clock_summary(emp,year,month)
        closure=TimePeriodClosure.query.filter_by(employee_id=emp.id,year=year,month=month).first()
        ack=TimeReportAcknowledgement.query.filter_by(employee_id=emp.id,year=year,month=month).first()
        finalization=TimeReportFinalization.query.filter_by(employee_id=emp.id,year=year,month=month).first()
        signature_log = None
        if ack:
            signature_log = (AuditLog.query
                .filter_by(entity="time_report_acknowledgement", entity_id=ack.id)
                .order_by(AuditLog.created_at.desc())
                .first())
        summaries.append({
            "emp": emp,
            "summary": s,
            "closure": closure,
            "ack": ack,
            "finalization": finalization,
            "signature_log": signature_log,
        })
    return render_template("time_closing.html", rows=summaries, month_value=month_value, year=year, month=month)

@bp.route("/employees/<int:employee_id>/time-closing", methods=["POST"])
@login_required
@roles_required(ROLE_ADMIN)
def employee_time_close(employee_id):
    emp=db.get_or_404(Employee,employee_id)
    year=int(request.form["year"]); month=int(request.form["month"])
    s=_month_clock_summary(emp,year,month)
    if s["incomplete"] and request.form.get("force")!="1":
        flash(f"Não foi possível fechar: existem {s['incomplete']} dia(s) com ponto ausente/incompleto. Revise ou use o fechamento justificado.", "danger")
        return redirect(url_for("rh.time_closing",month=f"{year:04d}-{month:02d}"))
    closure=TimePeriodClosure.query.filter_by(employee_id=emp.id,year=year,month=month).first()
    if not closure:
        closure=TimePeriodClosure(employee_id=emp.id,year=year,month=month,closed_by=current_user.id,
                                  reason=(request.form.get("reason") or "").strip() or None)
        db.session.add(closure)
    else:
        closure.status="closed"; closure.closed_at=now_local(); closure.closed_by=current_user.id
        closure.reason=(request.form.get("reason") or "").strip() or closure.reason
        closure.reopened_at = None
        closure.reopened_by = None
        closure.employee_viewed_at = None
    log_action("fechou competência de ponto","time_period_closure",employee_id,f"{month:02d}/{year}")
    db.session.commit(); flash("Competência fechada.", "success")
    return redirect(url_for("rh.time_closing",month=f"{year:04d}-{month:02d}"))


@bp.route("/employees/<int:employee_id>/time-closing/finalize", methods=["POST"])
@login_required
@roles_required(ROLE_ADMIN)
def employee_time_finalize(employee_id):
    emp = db.get_or_404(Employee, employee_id)
    year = int(request.form["year"])
    month = int(request.form["month"])
    note = (request.form.get("note") or "").strip() or None

    closure = TimePeriodClosure.query.filter_by(
        employee_id=emp.id, year=year, month=month, status="closed"
    ).first()
    ack = TimeReportAcknowledgement.query.filter_by(
        employee_id=emp.id, year=year, month=month
    ).first()

    if not closure or not ack:
        flash("O fechamento precisa estar fechado e assinado pelo colaborador antes da validação final do RH.", "danger")
        return redirect(url_for("rh.time_closing", month=f"{year:04d}-{month:02d}"))

    item = TimeReportFinalization.query.filter_by(
        employee_id=emp.id, year=year, month=month
    ).first()

    if not item:
        item = TimeReportFinalization(
            employee_id=emp.id,
            year=year,
            month=month,
            approved_by=current_user.id,
            note=note,
        )
        db.session.add(item)
        db.session.flush()
    else:
        item.approved_at = now_local()
        item.approved_by = current_user.id
        item.note = note

    log_action(
        "validou fechamento mensal assinado",
        "time_report_finalization",
        item.id,
        f"{emp.full_name}; competência {month:02d}/{year}; via assinada liberada ao colaborador",
    )
    log_security_event(
        "time_report_finalized_by_rh",
        severity="info",
        user=current_user,
        employee=emp,
        details=f"RH validou e arquivou o espelho assinado de {month:02d}/{year}.",
    )
    db.session.commit()

    flash("Fechamento validado pelo RH. A via assinada agora permanece disponível para o colaborador e para o RH.", "success")
    return redirect(url_for("rh.time_closing", month=f"{year:04d}-{month:02d}"))


@bp.route("/employees/<int:employee_id>/time-closing/reopen", methods=["POST"])
@login_required
@roles_required(ROLE_ADMIN)
def employee_time_reopen(employee_id):
    emp = db.get_or_404(Employee, employee_id)
    year = int(request.form["year"])
    month = int(request.form["month"])
    reason = (request.form.get("reason") or "").strip()

    if not reason:
        flash("Informe a justificativa para reabrir a competência.", "danger")
        return redirect(url_for("rh.time_closing", month=f"{year:04d}-{month:02d}"))

    closure = TimePeriodClosure.query.filter_by(
        employee_id=emp.id, year=year, month=month
    ).first()
    if not closure or closure.status != "closed":
        flash("Essa competência não está fechada.", "danger")
        return redirect(url_for("rh.time_closing", month=f"{year:04d}-{month:02d}"))

    # Ao reabrir, qualquer ciência/validação anterior deixa de representar
    # o conteúdo que poderá ser alterado. Exigimos novo ciclo após o próximo fechamento.
    ack = TimeReportAcknowledgement.query.filter_by(
        employee_id=emp.id, year=year, month=month
    ).first()
    finalization = TimeReportFinalization.query.filter_by(
        employee_id=emp.id, year=year, month=month
    ).first()

    if finalization:
        db.session.delete(finalization)
    if ack:
        db.session.delete(ack)

    closure.status = "reopened"
    closure.reopened_at = now_local()
    closure.reopened_by = current_user.id
    closure.employee_viewed_at = None

    log_action(
        "reabriu competência de ponto",
        "time_period_closure",
        closure.id,
        f"{emp.full_name}; {month:02d}/{year}; motivo: {reason}",
    )
    log_security_event(
        "time_period_reopened",
        severity="warning",
        user=current_user,
        employee=emp,
        details=f"Competência {month:02d}/{year} reaberta pelo RH. Motivo: {reason}",
    )
    db.session.commit()

    flash("Competência reaberta. Ajustes podem ser realizados novamente; será necessária nova assinatura após o próximo fechamento.", "success")
    return redirect(url_for("rh.time_closing", month=f"{year:04d}-{month:02d}"))


@bp.route("/time-report/acknowledge", methods=["POST"])
@login_required
def time_report_acknowledge():
    emp = current_user.employee
    if not emp:
        abort(403)

    year = int(request.form["year"])
    month = int(request.form["month"])
    closure = TimePeriodClosure.query.filter_by(
        employee_id=emp.id,
        year=year,
        month=month,
        status="closed",
    ).first()
    if not closure:
        abort(400)

    if not closure.employee_viewed_at:
        flash("Abra e confira o relatório mensal antes de assinar.", "danger")
        return redirect(url_for("main.dashboard"))

    if request.form.get("confirm_ack") != "1":
        flash("Confirme que você leu e conferiu o espelho mensal.", "danger")
        return redirect(url_for("main.dashboard"))

    pin = (request.form.get("point_pin") or "").strip()
    if _pin_is_blocked(emp):
        log_security_event(
            "pin_attempt_while_blocked",
            severity="critical",
            user=current_user,
            employee=emp,
            details="Tentativa de assinatura eletrônica recebida durante bloqueio de PIN.",
        )
        db.session.commit()
        flash("Muitas tentativas de PIN. Aguarde 15 minutos antes de tentar novamente.", "danger")
        return redirect(url_for("main.dashboard"))

    if not re.fullmatch(r"\d{6}", pin) or not emp.check_point_pin(pin):
        _pin_failure(emp)
        flash("Senha de ponto inválida. A assinatura não foi registrada.", "danger")
        return redirect(url_for("main.dashboard"))

    _pin_success(emp)
    ack = TimeReportAcknowledgement.query.filter_by(
        employee_id=emp.id,
        year=year,
        month=month,
    ).first()

    if not ack:
        ack = TimeReportAcknowledgement(
            employee_id=emp.id,
            year=year,
            month=month,
        )
        db.session.add(ack)
        db.session.flush()

        signature_code = _time_report_signature_code(ack, emp)
        log_action(
            "assinou eletronicamente o espelho mensal de ponto e banco de horas",
            "time_report_acknowledgement",
            ack.id,
            (
                f"{emp.full_name}; competência {month:02d}/{year}; "
                f"assinatura {signature_code}; autenticação por sessão e PIN pessoal"
            ),
        )
        log_security_event(
            "electronic_signature",
            severity="info",
            user=current_user,
            employee=emp,
            details=(
                f"Espelho mensal {month:02d}/{year} assinado eletronicamente. "
                f"Identificador: {signature_code}."
            ),
        )
        db.session.commit()

    flash(
        "Espelho mensal conferido e assinado eletronicamente com sucesso. "
        "O documento assinado já está disponível para o RH.",
        "success",
    )
    return redirect(url_for("main.dashboard"))

@bp.route("/time-records")
@login_required
@roles_required(ROLE_ADMIN, ROLE_MANAGER)
def time_records():
    emps = visible_employees()
    ids = [e.id for e in emps]
    q = TimeClock.query.filter(TimeClock.employee_id.in_(ids)) if ids else TimeClock.query.filter(db.text("0=1"))

    employee_id = request.args.get("employee_id", type=int)
    start_date_raw = request.args.get("start_date")
    end_date_raw = request.args.get("end_date")
    start_date = parse_date(start_date_raw) if start_date_raw else None
    end_date = parse_date(end_date_raw) if end_date_raw else None

    if employee_id and employee_id in ids:
        q = q.filter(TimeClock.employee_id == employee_id)
    if start_date:
        q = q.filter(func.date(TimeClock.punched_at) >= start_date)
    if end_date:
        q = q.filter(func.date(TimeClock.punched_at) <= end_date)

    rows = q.order_by(TimeClock.punched_at.desc()).limit(500).all()
    return render_template(
        "time_records.html",
        rows=rows,
        employees=emps,
        selected_employee=employee_id,
        start_date=start_date_raw or "",
        end_date=end_date_raw or "",
    )


@bp.route("/employees/<int:employee_id>/time-report.pdf")
@login_required
def employee_time_report_pdf(employee_id):
    emp = db.get_or_404(Employee, employee_id)

    month_value = (request.args.get("month") or "").strip()
    try:
        if month_value:
            year, month = [int(x) for x in month_value.split("-", 1)]
            if month < 1 or month > 12:
                raise ValueError
        else:
            local_today = today_local()
            year, month = local_today.year, local_today.month
    except (ValueError, TypeError):
        flash("Informe um mês válido para gerar o PDF.", "danger")
        return redirect(url_for("rh.employee_detail", employee_id=employee_id) + "#controle-ponto")

    # RH pode consultar qualquer colaborador. O próprio colaborador só acessa
    # o seu relatório quando a competência já estiver formalmente fechada.
    closure = TimePeriodClosure.query.filter_by(
        employee_id=emp.id, year=year, month=month, status="closed"
    ).first()

    ack = TimeReportAcknowledgement.query.filter_by(
        employee_id=emp.id, year=year, month=month
    ).first()
    finalization = TimeReportFinalization.query.filter_by(
        employee_id=emp.id, year=year, month=month
    ).first()
    report_version = (request.args.get("version") or "original").strip().lower()
    if report_version not in {"original", "signed"}:
        report_version = "original"

    if report_version == "signed" and not ack:
        flash("O colaborador ainda não assinou eletronicamente esta competência.", "danger")
        if current_user.role == ROLE_ADMIN:
            return redirect(url_for("rh.time_closing", month=f"{year:04d}-{month:02d}"))
        return redirect(url_for("main.dashboard"))

    if current_user.role != ROLE_ADMIN:
        if not current_user.employee or current_user.employee.id != emp.id:
            abort(403)
        if not closure:
            flash("Este relatório ainda não foi fechado pelo RH.", "danger")
            return redirect(url_for("main.dashboard"))

    first_day = date(year, month, 1)
    last_day = date(year, month, calendar.monthrange(year, month)[1])
    rows = (TimeClock.query
            .filter(TimeClock.employee_id == emp.id,
                    func.date(TimeClock.punched_at) >= first_day,
                    func.date(TimeClock.punched_at) <= last_day)
            .order_by(TimeClock.punched_at.asc())
            .all())

    grouped = {}
    for row in rows:
        grouped.setdefault(row.punched_at.date(), []).append(row)

    allowance_by_day = _certificate_allowance_by_day(emp, first_day, last_day)

    # Movimentações aprovadas no período: somente solicitações efetivamente aprovadas
    # entram no espelho mensal e no saldo apresentado ao RH.
    approved_requests = (Request.query
        .filter(Request.employee_id == emp.id,
                Request.status == "approved",
                Request.request_date >= first_day,
                Request.request_date <= last_day,
                Request.request_type.in_(["overtime", "bank_use"]))
        .order_by(Request.request_date.asc())
        .all())
    overtime_by_day = {}
    bank_use_by_day = {}
    for req_item in approved_requests:
        target = overtime_by_day if req_item.request_type == "overtime" else bank_use_by_day
        target[req_item.request_date] = target.get(req_item.request_date, 0) + int(req_item.minutes or 0)

    month_adjustments = (BankHourAdjustment.query
        .filter(BankHourAdjustment.employee_id == emp.id,
                func.date(BankHourAdjustment.created_at) >= first_day,
                func.date(BankHourAdjustment.created_at) <= last_day)
        .all())

    month_duties = (WeekendDuty.query
        .filter(
            WeekendDuty.employee_id == emp.id,
            WeekendDuty.duty_date >= first_day,
            WeekendDuty.duty_date <= last_day,
        )
        .order_by(WeekendDuty.duty_date.asc())
        .all())

    # Extrato mensal completo do banco de horas.
    bank_movements = []
    for item in approved_requests:
        minutes_value = int(item.minutes or 0)
        signed = minutes_value if item.request_type == "overtime" else -minutes_value
        bank_movements.append({
            "date": datetime.combine(item.request_date, item.start_time or datetime.min.time()),
            "description": "Hora extra aprovada" if item.request_type == "overtime" else "Utilização de banco aprovada",
            "source": "Solicitação",
            "minutes": signed,
        })
    for item in month_adjustments:
        bank_movements.append({
            "date": item.created_at,
            "description": item.reason or "Ajuste administrativo",
            "source": "Ajuste RH",
            "minutes": int(item.minutes or 0),
        })
    for duty in month_duties:
        bank_movements.append({
            "date": datetime.combine(duty.duty_date, datetime.min.time()),
            "description": "Plantão executado" + (f" - {duty.note}" if duty.note else ""),
            "source": "Plantão executado",
            "minutes": int(duty.minutes or 0),
        })
    bank_movements.sort(key=lambda x: x["date"])

    output = BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        leftMargin=15 * mm, rightMargin=15 * mm, topMargin=14 * mm, bottomMargin=14 * mm,
        title=f"Ponto - {emp.full_name} - {month:02d}/{year}",
        author="Portal RH",
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("ReportTitle", parent=styles["Title"], fontName="Helvetica-Bold",
                                 fontSize=16, leading=19, alignment=TA_CENTER, spaceAfter=4 * mm)
    small_style = ParagraphStyle("Small", parent=styles["Normal"], fontName="Helvetica", fontSize=8.5, leading=11)
    small_bold = ParagraphStyle("SmallBold", parent=small_style, fontName="Helvetica-Bold")
    note_style = ParagraphStyle("Note", parent=styles["Normal"], fontName="Helvetica", fontSize=8, leading=10, textColor=colors.HexColor("#555555"))

    story = [
        Paragraph("ESPELHO MENSAL DE REGISTROS DE PONTO", title_style),
        Paragraph(f"<b>Competência:</b> {_month_name_pt(month)} de {year}", small_style),
        Spacer(1, 2 * mm),
    ]

    info_data = [
        [Paragraph("Colaborador", small_bold), Paragraph(_pdf_text(emp.full_name), small_style),
         Paragraph("CPF", small_bold), Paragraph(_pdf_text(emp.cpf), small_style)],
        [Paragraph("Cargo", small_bold), Paragraph(_pdf_text(emp.job_title), small_style),
         Paragraph("Setor/Projeto", small_bold), Paragraph(_pdf_text(f"{emp.department} / {emp.project}"), small_style)],
        [Paragraph("Admissão", small_bold), Paragraph(emp.admission_date.strftime("%d/%m/%Y"), small_style),
         Paragraph("Carga semanal", small_bold), Paragraph(f"{emp.weekly_hours:g}h", small_style)],
        [Paragraph("Jornada padrão", small_bold),
         Paragraph(
             _pdf_text(
                 f"{emp.standard_start.strftime('%H:%M') if emp.standard_start else '-'} às "
                 f"{emp.standard_end.strftime('%H:%M') if emp.standard_end else '-'}"
             ),
             small_style,
         ),
         Paragraph("Intervalo previsto", small_bold),
         Paragraph(
             _pdf_text(
                 (
                     f"{emp.work_schedule.interval_start.strftime('%H:%M')} às "
                     f"{emp.work_schedule.interval_end.strftime('%H:%M')}"
                 )
                 if emp.work_schedule and emp.work_schedule.interval_start and emp.work_schedule.interval_end
                 else "Padrão do sistema"
             ),
             small_style,
         )],
    ]
    info_table = Table(info_data, colWidths=[27*mm, 93*mm, 30*mm, 100*mm])
    info_table.setStyle(TableStyle([
        ("GRID", (0,0), (-1,-1), 0.35, colors.HexColor("#CCCCCC")),
        ("BACKGROUND", (0,0), (0,-1), colors.HexColor("#F1F3F5")),
        ("BACKGROUND", (2,0), (2,-1), colors.HexColor("#F1F3F5")),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("LEFTPADDING", (0,0), (-1,-1), 5), ("RIGHTPADDING", (0,0), (-1,-1), 5),
        ("TOPPADDING", (0,0), (-1,-1), 4), ("BOTTOMPADDING", (0,0), (-1,-1), 4),
    ]))
    story.extend([info_table, Spacer(1, 4 * mm)])

    header = ["Data", "Entrada", "Saída intervalo", "Retorno", "Saída", "Horas trabalhadas", "Abono atestado", "Horas extras", "Outras marcações / observações"]
    data = [[Paragraph(f"<b>{h}</b>", small_style) for h in header]]
    total_worked_minutes = 0
    total_excused_minutes = 0

    for day in range(1, last_day.day + 1):
        current_day = date(year, month, day)
        day_rows = grouped.get(current_day, [])
        by_kind = {"entrada": [], "saida_intervalo": [], "retorno": [], "saida": []}
        other = []
        for r in day_rows:
            if r.kind in by_kind:
                by_kind[r.kind].append(r)
            else:
                other.append(r)

        def first_or_dash(kind):
            values = by_kind[kind]
            return values[0].punched_at.strftime("%H:%M:%S") if values else "-"

        extras = []
        for kind, values in by_kind.items():
            if len(values) > 1:
                extras.append(f"{_kind_label(kind)} adicional: " + ", ".join(v.punched_at.strftime("%H:%M:%S") for v in values[1:]))
        for r in other:
            extras.append(f"{_kind_label(r.kind)}: {r.punched_at.strftime('%H:%M:%S')}")
        if day_rows and any(r.source != "portal" for r in day_rows):
            adjusted = [r for r in day_rows if r.source != "portal"]
            extras.append("Ajuste administrativo: " + ", ".join(f"{_kind_label(r.kind)} {r.punched_at.strftime('%H:%M:%S')}" for r in adjusted))

        worked_minutes = _worked_minutes_for_day(day_rows)
        total_worked_minutes += worked_minutes

        excused_minutes = int(allowance_by_day.get(current_day, 0) or 0)
        total_excused_minutes += excused_minutes
        if excused_minutes:
            extras.append(f"Abono por atestado: {_format_minutes(excused_minutes)}")

        overtime_minutes = overtime_by_day.get(current_day, 0)
        bank_use_minutes = bank_use_by_day.get(current_day, 0)
        if bank_use_minutes:
            extras.append(f"Banco de horas utilizado: {_format_minutes(bank_use_minutes)}")

        weekday = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sáb", "Dom"][current_day.weekday()]
        data.append([
            f"{current_day.strftime('%d/%m/%Y')} ({weekday})",
            first_or_dash("entrada"), first_or_dash("saida_intervalo"),
            first_or_dash("retorno"), first_or_dash("saida"),
            _format_minutes(worked_minutes) if worked_minutes else "-",
            _format_minutes(excused_minutes) if excused_minutes else "-",
            _format_minutes(overtime_minutes) if overtime_minutes else "-",
            Paragraph(_pdf_text("; ".join(extras) if extras else ""), note_style),
        ])

    table = Table(data, repeatRows=1, colWidths=[27*mm, 22*mm, 25*mm, 22*mm, 22*mm, 27*mm, 27*mm, 24*mm, 54*mm])
    table_style = [
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#E9ECEF")),
        ("GRID", (0,0), (-1,-1), 0.3, colors.HexColor("#C8CDD2")),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("ALIGN", (1,1), (7,-1), "CENTER"),
        ("FONTNAME", (0,1), (7,-1), "Helvetica"),
        ("FONTSIZE", (0,0), (-1,-1), 7.5),
        ("TOPPADDING", (0,0), (-1,-1), 3.2),
        ("BOTTOMPADDING", (0,0), (-1,-1), 3.2),
    ]
    for idx in range(1, len(data)):
        day_date = date(year, month, idx)
        if day_date.weekday() >= 5:
            table_style.append(("BACKGROUND", (0,idx), (-1,idx), colors.HexColor("#FAFAFA")))
    table.setStyle(TableStyle(table_style))
    story.append(table)

    total_overtime = sum(overtime_by_day.values())
    total_bank_use = sum(bank_use_by_day.values())
    positive_adjustments = sum(a.minutes for a in month_adjustments if a.minutes > 0)
    negative_adjustments = abs(sum(a.minutes for a in month_adjustments if a.minutes < 0))
    total_duties = sum(int(d.minutes or 0) for d in month_duties)
    total_credits = total_overtime + positive_adjustments + total_duties
    total_debits = total_bank_use + negative_adjustments
    month_balance = total_credits - total_debits

    story.append(Spacer(1, 4 * mm))
    story.append(Paragraph("<b>RESUMO MENSAL</b>", small_style))
    summary_data = [
        [Paragraph("Horas trabalhadas registradas", small_bold), _format_minutes(total_worked_minutes),
         Paragraph("Horas abonadas por atestado", small_bold), _format_minutes(total_excused_minutes)],
        [Paragraph("Horas extras aprovadas", small_bold), _format_minutes(total_overtime),
         Paragraph("Plantões executados", small_bold), _format_minutes(total_duties)],
        [Paragraph("Créditos manuais de banco", small_bold), _format_minutes(positive_adjustments),
         Paragraph("Horas descontadas / banco utilizado", small_bold), _format_minutes(total_debits)],
        [Paragraph("Total de créditos no mês", small_bold), _format_minutes(total_credits),
         Paragraph("Saldo do banco no mês", small_bold), _format_minutes(month_balance)],
    ]
    summary_table = Table(summary_data, colWidths=[60*mm, 35*mm, 70*mm, 35*mm])
    summary_table.setStyle(TableStyle([
        ("GRID", (0,0), (-1,-1), 0.35, colors.HexColor("#C8CDD2")),
        ("BACKGROUND", (0,0), (0,-1), colors.HexColor("#F1F3F5")),
        ("BACKGROUND", (2,0), (2,-1), colors.HexColor("#F1F3F5")),
        ("ALIGN", (1,0), (1,-1), "CENTER"),
        ("ALIGN", (3,0), (3,-1), "CENTER"),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("FONTSIZE", (0,0), (-1,-1), 8.5),
        ("TOPPADDING", (0,0), (-1,-1), 5),
        ("BOTTOMPADDING", (0,0), (-1,-1), 5),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 5 * mm))

    story.append(Paragraph("<b>EXTRATO MENSAL DO BANCO DE HORAS</b>", small_style))
    bank_header = [
        Paragraph("<b>Data</b>", small_style),
        Paragraph("<b>Movimentação</b>", small_style),
        Paragraph("<b>Origem</b>", small_style),
        Paragraph("<b>Crédito</b>", small_style),
        Paragraph("<b>Débito</b>", small_style),
        Paragraph("<b>Saldo do mês</b>", small_style),
    ]
    bank_data = [bank_header]
    running_month = 0
    if bank_movements:
        for movement in bank_movements:
            running_month += movement["minutes"]
            bank_data.append([
                movement["date"].strftime("%d/%m/%Y"),
                Paragraph(_pdf_text(movement["description"]), note_style),
                movement["source"],
                _format_minutes(movement["minutes"]) if movement["minutes"] > 0 else "-",
                _format_minutes(abs(movement["minutes"])) if movement["minutes"] < 0 else "-",
                _format_minutes(running_month),
            ])
    else:
        bank_data.append([
            "-", Paragraph("Nenhuma movimentação de banco de horas nesta competência.", note_style),
            "-", "-", "-", _format_minutes(0),
        ])

    bank_table = Table(
        bank_data,
        repeatRows=1,
        colWidths=[28*mm, 88*mm, 32*mm, 30*mm, 30*mm, 34*mm],
    )
    bank_table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#E9ECEF")),
        ("GRID", (0,0), (-1,-1), 0.3, colors.HexColor("#C8CDD2")),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("ALIGN", (0,1), (0,-1), "CENTER"),
        ("ALIGN", (3,1), (-1,-1), "CENTER"),
        ("FONTSIZE", (0,0), (-1,-1), 8),
        ("TOPPADDING", (0,0), (-1,-1), 4),
        ("BOTTOMPADDING", (0,0), (-1,-1), 4),
    ]))
    story.append(bank_table)
    story.append(Spacer(1, 4 * mm))

    if closure:
        closure_text = (
            f"<b>Fechamento da competência:</b> {closure.closed_at.strftime('%d/%m/%Y às %H:%M:%S')}"
        )
        if closure.reason:
            closure_text += f" · <b>Observação do RH:</b> {_pdf_text(closure.reason)}"
        story.append(Paragraph(closure_text, note_style))
        story.append(Spacer(1, 2 * mm))

    story.append(Spacer(1, 5 * mm))
    story.append(Paragraph("<b>ASSINATURA ELETRÔNICA DO COLABORADOR</b>", small_style))

    if report_version == "signed" and ack:
        signature_log = (AuditLog.query
            .filter_by(entity="time_report_acknowledgement", entity_id=ack.id)
            .order_by(AuditLog.created_at.desc())
            .first())
        signature_code = _time_report_signature_code(ack, emp)
        signature_ip = signature_log.ip_address if signature_log and signature_log.ip_address else "-"
        signer_email = emp.user.email if emp.user else "-"
        signature_data = [
            [Paragraph("Status", small_bold), Paragraph("ASSINADO ELETRONICAMENTE", small_bold)],
            [Paragraph("Colaborador", small_bold), Paragraph(_pdf_text(emp.full_name), small_style)],
            [Paragraph("Usuário autenticado", small_bold), Paragraph(_pdf_text(signer_email), small_style)],
            [Paragraph("Data e hora do aceite", small_bold), Paragraph(ack.acknowledged_at.strftime("%d/%m/%Y às %H:%M:%S"), small_style)],
            [Paragraph("Identificador da assinatura", small_bold), Paragraph(signature_code, small_style)],
            [Paragraph("Registro de acesso", small_bold), Paragraph(_pdf_text(signature_ip), small_style)],
            [Paragraph("Método de confirmação", small_bold), Paragraph("Sessão autenticada no Portal RH + senha pessoal de ponto (6 dígitos)", small_style)],
        ]
        if finalization:
            signature_data.extend([
                [Paragraph("Validação final do RH", small_bold), Paragraph("VALIDADO E ARQUIVADO PELO RH", small_bold)],
                [Paragraph("Validado em", small_bold), Paragraph(finalization.approved_at.strftime("%d/%m/%Y às %H:%M:%S"), small_style)],
                [Paragraph("Responsável RH", small_bold), Paragraph(_pdf_text(finalization.approver.email if finalization.approver else "-"), small_style)],
                [Paragraph("Observação RH", small_bold), Paragraph(_pdf_text(finalization.note or "Sem observações."), small_style)],
            ])
        signature_table = Table(signature_data, colWidths=[55*mm, 145*mm])
        signature_table.setStyle(TableStyle([
            ("GRID", (0,0), (-1,-1), 0.45, colors.HexColor("#D4B13F")),
            ("BACKGROUND", (0,0), (0,-1), colors.HexColor("#FFF4CF")),
            ("BACKGROUND", (1,0), (1,0), colors.HexColor("#EAF6EC")),
            ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
            ("TOPPADDING", (0,0), (-1,-1), 5),
            ("BOTTOMPADDING", (0,0), (-1,-1), 5),
        ]))
        story.append(signature_table)
        story.append(Spacer(1, 2 * mm))
        final_text = (
            "O colaborador declarou ter visualizado e conferido este espelho mensal, "
            "registrando seu aceite eletrônico no Portal RH."
        )
        if finalization:
            final_text += (
                " O RH conferiu a assinatura e validou o fechamento, mantendo esta via "
                "disponível tanto para o colaborador quanto para o RH."
            )
        story.append(Paragraph(final_text, note_style))
    else:
        pending_signature = [
            [Paragraph("Status", small_bold), Paragraph("AGUARDANDO CIÊNCIA E ASSINATURA DO COLABORADOR", small_style)],
            [Paragraph("Colaborador", small_bold), Paragraph(_pdf_text(emp.full_name), small_style)],
            [Paragraph("Competência", small_bold), Paragraph(f"{month:02d}/{year}", small_style)],
            [Paragraph("Assinatura", small_bold), Paragraph("____________________________________________________________", small_style)],
        ]
        pending_table = Table(pending_signature, colWidths=[55*mm, 145*mm])
        pending_table.setStyle(TableStyle([
            ("GRID", (0,0), (-1,-1), 0.4, colors.HexColor("#C8CDD2")),
            ("BACKGROUND", (0,0), (0,-1), colors.HexColor("#F1F3F5")),
            ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
            ("TOPPADDING", (0,0), (-1,-1), 5),
            ("BOTTOMPADDING", (0,0), (-1,-1), 5),
        ]))
        story.append(pending_table)

    story.append(Spacer(1, 3 * mm))
    story.append(Paragraph(
        f"Documento gerado em {now_local().strftime('%d/%m/%Y às %H:%M:%S')} pelo Portal RH. "
        "Marcações ajustadas administrativamente permanecem registradas na trilha de Auditoria do sistema.",
        note_style,
    ))

    doc.build(story, onFirstPage=_draw_pdf_footer, onLaterPages=_draw_pdf_footer)
    output.seek(0)

    safe_name = secure_filename(emp.full_name).replace("_", "-") or f"colaborador-{emp.id}"
    suffix = "-assinado" if report_version == "signed" and ack else ""
    filename = f"ponto-banco-{safe_name}-{year}-{month:02d}{suffix}.pdf"
    if current_user.role != ROLE_ADMIN and closure and report_version == "original":
        closure.employee_viewed_at = now_local()
        log_action(
            "visualizou relatório mensal para ciência",
            "time_period_closure",
            closure.id,
            f"{emp.full_name}; competência {month:02d}/{year}"
        )
    else:
        log_action("gerou PDF mensal de ponto e banco", "employee", emp.id, f"competência {month:02d}/{year}")
    db.session.commit()

    # Para o colaborador, abre no navegador para leitura antes da ciência.
    # Para o RH, mantém o comportamento de download.
    return send_file(
        output,
        as_attachment=(current_user.role == ROLE_ADMIN),
        download_name=filename,
        mimetype="application/pdf"
    )

@bp.route("/report/time.xlsx")
@login_required
@roles_required(ROLE_ADMIN, ROLE_MANAGER)
def report_time():
    emps = visible_employees(); ids=[e.id for e in emps]
    rows=TimeClock.query.filter(TimeClock.employee_id.in_(ids)).order_by(TimeClock.punched_at.desc()).all()
    wb=Workbook(); ws=wb.active; ws.title="Ponto"
    ws.append(["Colaborador","Projeto","Setor","Data","Hora","Tipo","Origem"])
    for r in rows: ws.append([r.employee.full_name,r.employee.project,r.employee.department,r.punched_at.strftime("%d/%m/%Y"),r.punched_at.strftime("%H:%M:%S"),r.kind,r.source])
    bio=BytesIO(); wb.save(bio); bio.seek(0)
    return send_file(bio, as_attachment=True, download_name="relatorio_ponto.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@bp.route("/security-center")
@login_required
@roles_required(ROLE_ADMIN)
def security_center():
    """
    Painel operacional de segurança do RH.
    Consolida eventos de autenticação e ações sensíveis sem depender dos logs do Render.
    """
    now = now_local()

    try:
        days = int(request.args.get("days") or 7)
    except (TypeError, ValueError):
        days = 7
    days = min(max(days, 1), 90)

    severity = (request.args.get("severity") or "").strip().lower()
    event_type = (request.args.get("event_type") or "").strip()
    start = now - timedelta(days=days)

    security_query = SecurityEvent.query.filter(SecurityEvent.created_at >= start)
    if severity in {"info", "warning", "critical"}:
        security_query = security_query.filter(SecurityEvent.severity == severity)
    if event_type:
        security_query = security_query.filter(SecurityEvent.event_type == event_type)

    security_events = (
        security_query
        .order_by(SecurityEvent.created_at.desc())
        .limit(500)
        .all()
    )

    # Auditoria funcional relevante para segurança/privacidade.
    security_actions = [
        "login realizado",
        "alterou senha no primeiro acesso",
        "redefiniu senha provisória",
        "redefiniu senha de ponto",
        "tentativa inválida de registro de ponto",
        "adicionou marcação de ponto pelo RH",
        "alterou marcação de ponto pelo RH",
        "excluiu marcação de ponto pelo RH",
        "assinou eletronicamente o espelho mensal de ponto e banco de horas",
        "validou fechamento mensal assinado",
        "reabriu competência de ponto",
        "anexou documento",
        "enviou atestado",
        "anexou holerite manualmente",
        "separou holerite de PDF consolidado",
        "removeu holerite",
    ]
    audit_events = (
        AuditLog.query
        .filter(
            AuditLog.created_at >= start,
            AuditLog.action.in_(security_actions),
        )
        .order_by(AuditLog.created_at.desc())
        .limit(500)
        .all()
    )

    last_24h = now - timedelta(hours=24)
    login_failures_24h = SecurityEvent.query.filter(
        SecurityEvent.created_at >= last_24h,
        SecurityEvent.event_type.in_([
            "login_failed",
            "login_blocked",
            "login_attempt_while_blocked",
        ]),
    ).count()

    pin_failures_24h = SecurityEvent.query.filter(
        SecurityEvent.created_at >= last_24h,
        SecurityEvent.event_type.in_([
            "pin_failed",
            "pin_blocked",
            "pin_attempt_while_blocked",
        ]),
    ).count()

    sensitive_access_24h = SecurityEvent.query.filter(
        SecurityEvent.created_at >= last_24h,
        SecurityEvent.event_type == "sensitive_file_access",
    ).count()

    successful_logins_24h = SecurityEvent.query.filter(
        SecurityEvent.created_at >= last_24h,
        SecurityEvent.event_type == "login_success",
    ).count()

    blocked_rows = (
        AuthThrottle.query
        .filter(
            AuthThrottle.blocked_until.isnot(None),
            AuthThrottle.blocked_until > now,
        )
        .order_by(AuthThrottle.blocked_until.desc())
        .all()
    )

    active_users = User.query.filter_by(active=True).count()
    inactive_users = User.query.filter_by(active=False).count()

    critical_7d = SecurityEvent.query.filter(
        SecurityEvent.created_at >= now - timedelta(days=7),
        SecurityEvent.severity == "critical",
    ).count()

    event_types = [
        row[0]
        for row in (
            db.session.query(SecurityEvent.event_type)
            .filter(SecurityEvent.created_at >= now - timedelta(days=90))
            .distinct()
            .order_by(SecurityEvent.event_type.asc())
            .all()
        )
    ]

    return render_template(
        "security_center.html",
        security_events=security_events,
        audit_events=audit_events,
        blocked_rows=blocked_rows,
        days=days,
        severity=severity,
        selected_event_type=event_type,
        event_types=event_types,
        login_failures_24h=login_failures_24h,
        pin_failures_24h=pin_failures_24h,
        sensitive_access_24h=sensitive_access_24h,
        successful_logins_24h=successful_logins_24h,
        active_users=active_users,
        inactive_users=inactive_users,
        critical_7d=critical_7d,
        now=now,
    )


@bp.route("/audit")
@login_required
@roles_required(ROLE_ADMIN)
def audit():
    start_raw = (request.args.get("start_date") or "").strip()
    end_raw = (request.args.get("end_date") or "").strip()

    start_date = parse_date(start_raw) if start_raw else None
    end_date = parse_date(end_raw) if end_raw else None

    if start_date and end_date and end_date < start_date:
        flash("A data final não pode ser anterior à data inicial.", "danger")
        return redirect(url_for("rh.audit"))

    query = AuditLog.query

    if start_date:
        query = query.filter(func.date(AuditLog.created_at) >= start_date)
    if end_date:
        query = query.filter(func.date(AuditLog.created_at) <= end_date)

    rows = query.order_by(AuditLog.created_at.desc()).limit(2000).all()

    return render_template(
        "audit.html",
        rows=rows,
        start_date=start_raw,
        end_date=end_raw,
    )


@bp.route("/audit.pdf")
@login_required
@roles_required(ROLE_ADMIN)
def audit_pdf():
    start_raw = (request.args.get("start_date") or "").strip()
    end_raw = (request.args.get("end_date") or "").strip()

    start_date = parse_date(start_raw) if start_raw else None
    end_date = parse_date(end_raw) if end_raw else None

    if start_date and end_date and end_date < start_date:
        flash("A data final não pode ser anterior à data inicial.", "danger")
        return redirect(url_for("rh.audit"))

    query = AuditLog.query

    if start_date:
        query = query.filter(func.date(AuditLog.created_at) >= start_date)
    if end_date:
        query = query.filter(func.date(AuditLog.created_at) <= end_date)

    rows = query.order_by(AuditLog.created_at.asc()).all()

    output = BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title="Relatório de Auditoria - Portal RH",
        author="Portal RH - Associação Alecrim Dourado",
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "AuditTitle",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=16,
        leading=19,
        alignment=TA_CENTER,
        spaceAfter=4 * mm,
    )
    small = ParagraphStyle(
        "AuditSmall",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=7.2,
        leading=9,
    )
    small_bold = ParagraphStyle(
        "AuditSmallBold",
        parent=small,
        fontName="Helvetica-Bold",
    )

    period_start = start_date.strftime("%d/%m/%Y") if start_date else "Início dos registros"
    period_end = end_date.strftime("%d/%m/%Y") if end_date else "Data atual"

    story = [
        Paragraph("RELATÓRIO DE AUDITORIA DO PORTAL RH", title_style),
        Paragraph(
            f"<b>Período:</b> {period_start} até {period_end}",
            small,
        ),
        Paragraph(
            f"<b>Gerado em:</b> {now_local().strftime('%d/%m/%Y às %H:%M:%S')} "
            f"&nbsp;&nbsp;&nbsp; <b>Responsável pela exportação:</b> {_pdf_text(current_user.email)}",
            small,
        ),
        Spacer(1, 4 * mm),
    ]

    headers = ["Data/Hora", "Usuário", "Ação", "Recurso", "IP", "Detalhes"]
    data = [[Paragraph(f"<b>{h}</b>", small_bold) for h in headers]]

    for log in rows:
        user_label = log.user.email if log.user else "Sistema / não autenticado"
        resource = f"{log.entity}"
        if log.entity_id:
            resource += f" #{log.entity_id}"

        data.append([
            Paragraph(log.created_at.strftime("%d/%m/%Y %H:%M:%S"), small),
            Paragraph(_pdf_text(user_label), small),
            Paragraph(_pdf_text(log.action), small),
            Paragraph(_pdf_text(resource), small),
            Paragraph(_pdf_text(log.ip_address or "—"), small),
            Paragraph(_pdf_text(log.details or "—"), small),
        ])

    if len(data) == 1:
        data.append([
            Paragraph("—", small),
            Paragraph("—", small),
            Paragraph("Nenhum registro encontrado no período.", small),
            Paragraph("—", small),
            Paragraph("—", small),
            Paragraph("—", small),
        ])

    table = Table(
        data,
        repeatRows=1,
        colWidths=[34*mm, 48*mm, 48*mm, 35*mm, 34*mm, 74*mm],
    )
    table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#E9ECEF")),
        ("GRID", (0,0), (-1,-1), 0.3, colors.HexColor("#C7CDD2")),
        ("VALIGN", (0,0), (-1,-1), "TOP"),
        ("LEFTPADDING", (0,0), (-1,-1), 4),
        ("RIGHTPADDING", (0,0), (-1,-1), 4),
        ("TOPPADDING", (0,0), (-1,-1), 3),
        ("BOTTOMPADDING", (0,0), (-1,-1), 3),
    ]))

    story.append(table)
    story.append(Spacer(1, 4 * mm))
    story.append(Paragraph(
        f"Total de registros no período: <b>{len(rows)}</b>. "
        "Este relatório é uma exportação da trilha de auditoria do Portal RH.",
        small,
    ))

    doc.build(story)
    output.seek(0)

    file_start = start_date.strftime("%Y%m%d") if start_date else "inicio"
    file_end = end_date.strftime("%Y%m%d") if end_date else today_local().strftime("%Y%m%d")
    filename = f"auditoria-portal-rh-{file_start}-{file_end}.pdf"

    log_action(
        "exportou auditoria em PDF",
        "audit",
        None,
        f"Período {period_start} até {period_end}; {len(rows)} registros exportados.",
    )
    db.session.commit()

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/pdf",
    )
